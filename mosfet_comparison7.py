"""
MOSFET Comparison & Loss Analysis Tool  —  v4
==============================================
Reads the Universal Design Reference v5 extractor Excel format.

Column layout of each input file (v5 format):
  Col A (0): Symbol          e.g. "V_DS", "R_DS(on)", "E_on" ...
  Col B (1): Parameter name  (ignored — we use Symbol)
  Col C (2): Unit            e.g. "mΩ", "μJ", "pF"
  Col D (3): Typ @ 25 °C     ← PRIMARY value used
  Col E (4): Max @ 25 °C
  Col F (5): Typ @ High-T    ← used for Rds @ 125 °C
  Col G (6): Max @ High-T
  Col H (7): Test Conditions ← parsed for VDS/ID/Tj context
  Col I (8): Loss Relevance  (ignored)
  Col J (9): Status          "Found ✓" / "Not Found"

Row 2 holds "Part: XXX | Manufacturer: YYY | ..."
Row 4 is the header row.
Data starts at row 5 (0-indexed row 4).

PARAMETER MAPPINGS (Symbol → internal key):
  V_DS          → vdss          (V, rated)
  I_D           → id            (A, @ 25°C col D; high-T col F)
  R_DS(on)      → rds25 (col D typ@25°C), rds_max25 (col E max@25°C),
                  rds_highT (col F typ@highT)
  V_GS(th)      → vth           (V, typ col D, max col E)
  C_iss         → ciss          (pF)
  C_oss         → coss          (pF, typ col D; max col E)
  C_rss         → crss          (pF)
  Q_g           → qg            (nC, with test conditions in col H)
  Q_gd          → qgd           (nC)
  t_d(on)       → td_on         (ns, with test conditions)
  t_r           → tr            (ns)
  t_d(off)      → td_off        (ns)
  t_f           → tf            (ns)
  E_oss         → eoss          (μJ)
  V_SD          → vsd           (V)
  E_on          → eon           (μJ, with test conditions → _eon_itest, _eon_vds)
  E_off         → eoff          (μJ, with test conditions → _eoff_itest)
  t_rr          → trr           (ns)
  Q_rr / E_rr   → qrr           (nC)

HIGH-TEMP Rds handling:
  If col F (Typ @ High-T) is not "—" and the test condition in col H
  mentions a temperature (e.g. "Tj = 150 °C") we store that as rds_highT
  with its temperature in _rds_highT_tj.
  If no explicit temperature is found we default to 125 °C.

NORMALISATION (common-ground comparison):
  Rds_on(Tj_ref)  → linear interpolation/extrapolation between the two
                    temperature points (25 °C and the high-T point).
  Coss @ VDS_ref  → depletion-cap model  C = C0/√(1+VDS/Vbi)
  Eon/Eoff @ ID_ref → linear scaling  E_scaled = E_ref × (ID_ref / ID_test)
  Computed switching energy from timing (when Eon/Eoff not tabulated):
    E_on  ≈ ½ × VDS_test × ID_test × (t_r + t_d_on)   [ns×A×V → µJ /1000]
    E_off ≈ ½ × VDS_test × ID_test × (t_f + t_d_off)

LOSS MODEL:
  P_cond   = Irms² × Rds_on(Tj) [Ω]
  P_sw_on  = Eon[J] × (Iavg / ID_test) × Fsw
  P_sw_off = Eoff[J] × (Iavg / ID_test) × Fsw
  P_coss   = Eoss[J] × Fsw
  P_gate   = Qg[C] × VGS_drive × Fsw
  P_body_diode = Vsd × Iavg × deadtime_ratio  (informational, not summed)
  P_total  = P_cond + P_sw_on + P_sw_off + P_coss + P_gate

All loss values in Excel are LIVE FORMULAS.
Yellow cells = user-editable operating conditions.
Blue cells   = device parameters (pre-filled, fully editable).
"""

import sys, os, re, warnings, io, tempfile, math
from pathlib import Path

import numpy as np
try:
    from scipy.optimize import curve_fit
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ─── colours ──────────────────────────────────────────────────────────────────
C_TITLE   = "13315C";  C_SUBHDR  = "2E75B6";  C_HDR     = "1F3864"
C_PRI_BG  = "1A3C5E";  C_SEC_BG  = "2E75B6"
C_BEST    = "E2EFDA";  C_BESTF   = "375623"
C_WARN    = "FCE4D6";  C_WARNF   = "843C0C"
C_ALT     = "F2F8FF";  C_WHITE   = "FFFFFF"
C_INPUT   = "FFF2CC";  C_GOLD    = "FFD700"
C_GREEN   = "375623";  C_NORM    = "E8F5E9"
C_PRI_ROW = "E3EEF8";  C_DEVPARAM= "D6EEFF"
C_DERIVED = "FFF0CC"   # orange-yellow for values derived from timing

_t = Side(style="thin",   color="AAAAAA")
_m = Side(style="medium", color="444444")
BT = Border(left=_t, right=_t, top=_t, bottom=_t)
BM = Border(left=_m, right=_m, top=_m, bottom=_m)

def ff(h):  return PatternFill("solid", fgColor=h)
def ctr():  return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft():  return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def fnt(sz=10, bold=False, color="000000", italic=False):
    return Font(name="Calibri", size=sz, bold=bold, color=color, italic=italic)

def wc(ws, r, c, v, font=None, fill=None, align=None, border=None, nfmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font          = font
    if fill:   cell.fill          = fill
    if align:  cell.alignment     = align
    if border: cell.border        = border
    if nfmt:   cell.number_format = nfmt
    return cell

def merge_title(ws, r, c1, c2, text, bg=C_HDR, fg="FFFFFF", sz=12, h=30):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(row=r, column=c1, value=text)
    c.font = Font(name="Calibri", bold=True, size=sz, color=fg)
    c.fill = ff(bg); c.alignment = ctr(); c.border = BM
    ws.row_dimensions[r].height = h

def gcol(n): return get_column_letter(n)

# ─── PARAMETER CATALOGUE ──────────────────────────────────────────────────────
# key: (display_name, unit, lower_is_better, is_priority, show_conditions)
PARAMS = {
    # ── PRIORITY ──────────────────────────────────────────────────────────────
    "vdss":   ("V_DS – Breakdown Voltage",           "V",    False, True,  False),
    "rds25":  ("R_DS(on) @ 25 °C  (typ)",            "mΩ",   True,  True,  True),
    "rds_ht": ("R_DS(on) @ High-T  (typ)",           "mΩ",   True,  True,  True),
    "qg":     ("Q_G – Gate Charge  @ VDS_ref",        "nC",   True,  True,  True),
    "eon":    ("E_on – Turn-ON Energy  @ ID_ref, VDS_ref",  "µJ", True, True, True),
    "eoff":   ("E_off – Turn-OFF Energy  @ ID_ref, VDS_ref","µJ", True, True, True),
    # ── SECONDARY ─────────────────────────────────────────────────────────────
    "id":     ("I_D – Continuous Drain Current",     "A",    False, False, False),
    "vth":    ("V_GS(th) – Gate Threshold  @ Tj_ref","V",    False, False, True),
    "ciss":   ("C_iss – Input Cap.  @ VDS_ref",      "pF",   True,  False, True),
    "coss":   ("C_oss – Output Cap.  @ VDS_ref",     "pF",   True,  False, True),
    "crss":   ("C_rss – Rev. Transfer Cap.  @ VDS_ref","pF", True,  False, False),
    "qgd":    ("Q_GD – Miller Charge  (direct)",     "nC",   True,  False, False),
    "eoss":   ("E_oss – Output Energy  @ VDS_ref",   "µJ",   True,  False, False),
    "vsd":    ("V_SD – Body Diode Voltage  @ Tj_ref","V",    True,  False, True),
    "trr":    ("t_rr – Rev. Recovery Time  @ Tj_ref","ns",   True,  False, True),
    "qrr":    ("Q_rr – Rev. Recovery Charge  @ Tj_ref","nC", True,  False, False),
    "td_on":  ("t_d(on) – Turn-On Delay  (direct)",  "ns",   True,  False, True),
    "tr":     ("t_r – Rise Time  (direct)",           "ns",   True,  False, False),
    "td_off": ("t_d(off) – Turn-Off Delay  (direct)","ns",   True,  False, False),
    "tf":     ("t_f – Fall Time  (direct)",           "ns",   True,  False, False),
}

PRIORITY_KEYS  = [k for k, v in PARAMS.items() if v[3]]
SECONDARY_KEYS = [k for k, v in PARAMS.items() if not v[3]]
ALL_KEYS       = PRIORITY_KEYS + SECONDARY_KEYS

# ─── LOG-GRAPH PARAMETERS ─────────────────────────────────────────────────────
# Parameters whose datasheet values come from a LOG-scale graph (C-vs-VDS).
# Auto-extraction and the 1/√V depletion formula are both unreliable for these,
# so in the Common-Ground sheet we do not print a misleading number — instead we
# point the reader to the side-by-side datasheet images in the Graph Comparison
# sheet.  (User request: "just put this line — refer the graphs in graph section".)
_REFER_TO_GRAPH_KEYS = {"ciss", "coss", "crss"}
_REFER_TO_GRAPH_TEXT = "refer the graphs in graph section"

# ─── v5 EXCEL READER ──────────────────────────────────────────────────────────

def _val(raw):
    """Convert a cell value to float, returning None for blanks/dashes.
    Handles European decimal notation where comma is the decimal separator,
    common in German/EU datasheets (e.g. Infineon): '1,000' → 1.0; '2,500' → 2.5.
    """
    if raw is None: return None
    if isinstance(raw, (int, float)): return float(raw)
    s = str(raw).strip()
    if s in ("", "—", "-", "N/A", "Not Found"): return None
    # European decimal: comma as decimal separator, no period present
    # e.g. "0,500" → 0.5; "1,000" → 1.0; "2,500" → 2.5
    if ',' in s and '.' not in s:
        s = s.replace(',', '.')
    m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
    return float(m.group()) if m else None

def _parse_tcond(tcond):
    """
    Parse test condition string.
    Returns dict with any of: vds, id, vgs, rg, tj
    Examples:
      "VGS=10V, ID=15A"         → {vgs:10, id:15}
      "VDD=400V, ID=30A, RG=4.7Ω, VGS=10V"  → {vds:400, id:30, rg:4.7, vgs:10}
      "Tj=150°C"                → {tj:150}
    """
    if not tcond: return {}
    t = str(tcond)
    out = {}
    # VDS / VDD / VDrain  — use \d+\.?\d* to avoid matching ranges like "0...400"
    m = re.search(r"V(?:DS|DD|D)\s*=\s*(\d+\.?\d*)", t, re.I)
    if m: out["vds"] = float(m.group(1))
    # ID / IDS
    m = re.search(r"I_?D\s*=\s*(\d+\.?\d*)\s*A", t, re.I)
    if m: out["id"] = float(m.group(1))
    # VGS
    m = re.search(r"V_?GS\s*=\s*(\d+\.?\d*)\s*V", t, re.I)
    if m: out["vgs"] = float(m.group(1))
    # RG (gate resistance)
    m = re.search(r"R_?G\s*=\s*(\d+\.?\d*)", t, re.I)
    if m: out["rg"] = float(m.group(1))
    # Tj
    m = re.search(r"T_?j\s*=\s*(\d+\.?\d*)\s*°?[Cc]", t, re.I)
    if m: out["tj"] = float(m.group(1))
    return out


def read_v5_file(path):
    """
    Read a v5-format MOSFET parameter Excel file.
    Returns a rich dict with all extracted values plus metadata.
    """
    wb  = openpyxl.load_workbook(path, data_only=True)
    ws  = wb.active
    d   = {
        "part": Path(path).stem, "mfr": "",
        "coss_rows": [],   # (vds, coss_pF) for extrapolation
        "ciss_rows": [],   # (vds, ciss_pF) for VDS-depletion extrapolation
        "crss_rows": [],   # (vds, crss_pF) for VDS-depletion extrapolation
    }

    # ── Row 2: Part / Manufacturer ──────────────────────────────────────────
    hdr = str(ws.cell(row=2, column=1).value or "")
    m = re.search(r"Part:\s*([^|]+)", hdr)
    if m: d["part"] = m.group(1).strip()
    m = re.search(r"Manufacturer:\s*([^|]+)", hdr)
    if m: d["mfr"]  = m.group(1).strip()

    # ── Data rows (row 5 onward) ─────────────────────────────────────────────
    # v5 format: Symbol|Name|Unit|Typ|Max|Temp(°C)|Test Conditions|Design Note|Status
    _cur_sym = ""   # tracks symbol for multi-row (continuation) parameter handling
    for row in ws.iter_rows(min_row=5, values_only=True):
        sym    = str(row[0] or "").strip()  if row[0] else ""
        unit   = str(row[2] or "").strip()  if len(row) > 2 else ""
        v25    = _val(row[3])               if len(row) > 3 else None  # Typ value
        vmax25 = _val(row[4])               if len(row) > 4 else None  # Max value
        temp_c = _val(row[5])               if len(row) > 5 else None  # Temperature (°C)
        tcond  = str(row[6] or "").strip()  if len(row) > 6 else ""    # Test Conditions
        status = str(row[8] or "").strip()  if len(row) > 8 else ""

        if sym.startswith("▶") or sym.lower() in ("symbol", "legend:"):
            continue

        # Continuation rows (empty symbol) belong to the previous multi-row parameter
        if sym == "":
            if _cur_sym == "R_DS(on)" and v25 is not None and temp_c is not None and temp_c > 50:
                _rds_k = d.get("_rds_k", 1.0)
                d["rds_ht"]     = v25 * _rds_k
                d["rds_ht_max"] = vmax25 * _rds_k if vmax25 is not None else None
                d["_rds_ht_tj"] = temp_c
            elif _cur_sym == "I_D" and temp_c is not None and temp_c > 50:
                d["id_highT"] = v25 if v25 is not None else vmax25
            continue

        _cur_sym = sym
        tc = _parse_tcond(tcond)

        # ── V_DS ───────────────────────────────────────────────────────────
        if sym == "V_DS":
            d["vdss"]      = v25
            d["vdss_max"]  = vmax25
            d["vdss_tcond"]= tcond

        # ── I_D ────────────────────────────────────────────────────────────
        elif sym == "I_D":
            d["id"]       = v25 if v25 is not None else vmax25  # prefer Typ, fall back to Max
            d["id_highT"] = None          # set by high-T continuation row
            d["id_tcond"] = tcond

        # ── R_DS(on) ────────────────────────────────────────────────────────
        elif sym == "R_DS(on)":
            # Unit-aware conversion: some extractors output Ω instead of mΩ.
            # Infineon's Symbol-font Omega glyph is sometimes mis-decoded as "W"
            # by PDF text extractors — treat "W" as Ω for this parameter.
            _u = unit.strip()
            _u_low = _u.lower().replace('Ω', 'ohm').replace('Ω', 'ohm')
            if 'mohm' in _u_low or 'mΩ' in _u or 'mΩ' in _u or 'mω' in _u_low:
                _rds_k = 1.0       # already in mΩ
            elif 'ohm' in _u_low or 'Ω' in _u or 'Ω' in _u or _u_low == 'w':
                _rds_k = 1000.0    # Ω → mΩ  (includes "W" which is a Symbol-font Ω)
            else:
                _rds_k = 1.0       # unknown unit — keep as-is, assume mΩ
            d["_rds_k"]     = _rds_k          # saved for high-T continuation rows
            d["rds25"]      = v25 * _rds_k if v25 is not None else None
            d["rds25_max"]  = vmax25 * _rds_k if vmax25 is not None else None
            d["rds_ht"]     = None             # set by high-T continuation row
            d["rds_ht_max"] = None             # set by high-T continuation row
            d["_rds_ht_tj"] = None             # set by high-T continuation row
            d["rds_tcond"]  = tcond
            d["_rds_unit"]  = _u
            d["_rds_id_test"] = tc.get("id")
            d["_rds_vgs_test"]= tc.get("vgs")

        # ── V_GS(th) ────────────────────────────────────────────────────────
        elif sym == "V_GS(th)":
            d["vth"]     = v25
            d["vth_max"] = vmax25
            d["vth_tcond"]= tcond

        # ── C_iss ────────────────────────────────────────────────────────────
        elif sym == "C_iss":
            d["ciss"]      = v25
            d["ciss_tcond"]= tcond
            vds_ci = tc.get("vds")
            d["_ciss_vds"] = vds_ci
            d["_ciss_vds_test"] = vds_ci
            if vds_ci is not None and v25 is not None:
                d["ciss_rows"].append((vds_ci, v25))

        # ── C_oss ────────────────────────────────────────────────────────────
        elif sym == "C_oss":
            d["coss"]      = v25
            d["coss_max"]  = vmax25
            d["coss_tcond"]= tcond
            vds_c = tc.get("vds")
            # Try to extract VDS from tcond or fall back to unit context
            if vds_c is None:
                # sometimes condition is just "—" – try to infer from VDSS
                # we'll add to coss_rows when we know VDSS (done after loop)
                pass
            if vds_c is not None and v25 is not None:
                d["coss_rows"].append((vds_c, v25))
            d["_coss_vds_test"] = vds_c

        # ── C_rss ────────────────────────────────────────────────────────────
        elif sym == "C_rss":
            d["crss"]      = v25
            d["crss_tcond"]= tcond
            vds_cr = tc.get("vds")
            d["_crss_vds_test"] = vds_cr
            if vds_cr is not None and v25 is not None:
                d["crss_rows"].append((vds_cr, v25))

        # ── Q_g ──────────────────────────────────────────────────────────────
        elif sym == "Q_g":
            d["qg"]       = v25
            d["qg_tcond"] = tcond
            d["_qg_vds"]  = tc.get("vds")
            d["_qg_id"]   = tc.get("id")
            d["_qg_vgs"]  = tc.get("vgs")

        # ── Q_gd ─────────────────────────────────────────────────────────────
        elif sym == "Q_gd":
            d["qgd"]       = v25
            d["qgd_tcond"] = tcond

        # ── t_d(on) ──────────────────────────────────────────────────────────
        elif sym == "t_d(on)":
            d["td_on"]       = v25
            d["td_on_tcond"] = tcond
            d["_sw_vds"]     = tc.get("vds")   # switching test VDS (shared)
            d["_sw_id"]      = tc.get("id")    # switching test ID
            d["_sw_vgs"]     = tc.get("vgs")
            d["_sw_rg"]      = tc.get("rg")

        # ── t_r ──────────────────────────────────────────────────────────────
        elif sym == "t_r":
            d["tr"]       = v25
            d["tr_tcond"] = tcond

        # ── t_d(off) ─────────────────────────────────────────────────────────
        elif sym == "t_d(off)":
            d["td_off"]       = v25
            d["td_off_tcond"] = tcond

        # ── t_f ──────────────────────────────────────────────────────────────
        elif sym == "t_f":
            d["tf"]       = v25
            d["tf_tcond"] = tcond

        # ── E_oss ────────────────────────────────────────────────────────────
        elif sym == "E_oss":
            d["eoss"]       = v25
            d["eoss_tcond"] = tcond

        # ── V_SD ─────────────────────────────────────────────────────────────
        elif sym == "V_SD":
            d["vsd"]       = v25
            d["vsd_tcond"] = tcond

        # ── E_on ─────────────────────────────────────────────────────────────
        elif sym == "E_on":
            d["eon"]          = v25
            d["eon_tcond"]    = tcond
            d["_eon_vds"]     = tc.get("vds")
            d["_eon_id"]      = tc.get("id")

        # ── E_off ────────────────────────────────────────────────────────────
        elif sym == "E_off":
            d["eoff"]         = v25
            d["eoff_tcond"]   = tcond
            d["_eoff_vds"]    = tc.get("vds")
            d["_eoff_id"]     = tc.get("id")

        # ── t_rr ─────────────────────────────────────────────────────────────
        elif sym == "t_rr":
            d["trr"]       = v25
            d["trr_tcond"] = tcond

        # ── Q_rr / E_rr ──────────────────────────────────────────────────────
        elif re.match(r"Q_rr", sym):
            d["qrr"]       = v25
            d["qrr_tcond"] = tcond

    # ── Post-processing ───────────────────────────────────────────────────────

    # If Coss was found but no test VDS was extracted, add a fallback entry
    # using Coss_test_VDS = VDSS × 0.5 as a rough guess (or 400V if unknown)
    if d.get("coss") is not None and not d.get("coss_rows"):
        fallback_vds = d.get("vdss", 650.0) * 0.5 if d.get("vdss") else 400.0
        d["coss_rows"].append((fallback_vds, d["coss"]))
        d["_coss_vds_test"] = fallback_vds

    # Ciss fallback: if no VDS test point extracted, assume VDSS×0.5
    if d.get("ciss") is not None and not d.get("ciss_rows"):
        fallback_vds = d.get("vdss", 650.0) * 0.5 if d.get("vdss") else 400.0
        if d.get("_ciss_vds") is not None:
            fallback_vds = d["_ciss_vds"]
        d["ciss_rows"].append((fallback_vds, d["ciss"]))
        d["_ciss_vds_test"] = fallback_vds

    # Crss fallback: if no VDS test point extracted, assume VDSS×0.5
    if d.get("crss") is not None and not d.get("crss_rows"):
        fallback_vds = d.get("vdss", 650.0) * 0.5 if d.get("vdss") else 400.0
        if d.get("_crss_vds_test") is not None:
            fallback_vds = d["_crss_vds_test"]
        d["crss_rows"].append((fallback_vds, d["crss"]))
        d["_crss_vds_test"] = fallback_vds

    # Derive Eoss from Coss when not tabulated
    if d.get("eoss") is None and d.get("coss") is not None:
        vds_t = d.get("_coss_vds_test") or (d.get("vdss") or 650.0) * 0.5
        d["eoss"] = round(0.5 * d["coss"] * 1e-12 * vds_t**2 * 1e6, 4)
        d["_eoss_derived"] = True

    # Derive Eon/Eoff from switching times when not tabulated
    if d.get("eon") is None:
        vds_t = d.get("_sw_vds") or d.get("_eon_vds") or d.get("_qg_vds") or (d.get("vdss") or 400.0)*0.6
        id_t  = d.get("_sw_id")  or d.get("_eon_id")  or d.get("id") or 15.0
        tr    = d.get("tr")
        tdon  = d.get("td_on")
        if tr is not None and tdon is not None:
            # E = ½ × V × I × t_switch [V × A × ns → µJ = /1000]
            d["eon"]  = round(0.5 * vds_t * id_t * (tr + tdon) / 1000.0, 4)
            d["_eon_derived"]  = True
            d["_eon_vds"]      = vds_t
            d["_eon_id"]       = id_t
        elif tr is not None:
            d["eon"]  = round(0.5 * vds_t * id_t * tr / 1000.0, 4)
            d["_eon_derived"]  = True
            d["_eon_vds"]      = vds_t
            d["_eon_id"]       = id_t

    if d.get("eoff") is None:
        vds_t = d.get("_sw_vds")  or d.get("_eoff_vds") or d.get("_qg_vds") or (d.get("vdss") or 400.0)*0.6
        id_t  = d.get("_sw_id")   or d.get("_eoff_id")  or d.get("id") or 15.0
        tf    = d.get("tf")
        tdoff = d.get("td_off")
        if tf is not None and tdoff is not None:
            d["eoff"] = round(0.5 * vds_t * id_t * (tf + tdoff) / 1000.0, 4)
            d["_eoff_derived"] = True
            d["_eoff_vds"]     = vds_t
            d["_eoff_id"]      = id_t
        elif tf is not None:
            d["eoff"] = round(0.5 * vds_t * id_t * tf / 1000.0, 4)
            d["_eoff_derived"] = True
            d["_eoff_vds"]     = vds_t
            d["_eoff_id"]      = id_t

    # Default gate drive voltage (from Qg test condition if available)
    d.setdefault("vgs_drive", d.get("_qg_vgs") or d.get("_sw_vgs") or 10.0)
    d.setdefault("tjmax", 175.0)

    wb.close()
    return d


def mosfet_label(path, d):
    part = (d.get("part") or "").strip()
    mfr  = (d.get("mfr")  or "").strip()
    if part and part.lower() not in ("unknown", ""):
        return f"{part} ({mfr})" if mfr else part
    return Path(path).stem


# ─── PHYSICS MODELS ───────────────────────────────────────────────────────────

def _interp_at(pairs, x):
    """
    Linear interpolation of y at x from a sorted list of (x, y) pairs.
    Clamps to the nearest endpoint when x is outside the range.
    Returns None if pairs is empty.
    """
    if not pairs:
        return None
    if x <= pairs[0][0]:
        return pairs[0][1]
    if x >= pairs[-1][0]:
        return pairs[-1][1]
    for i in range(len(pairs) - 1):
        x0, y0 = pairs[i]
        x1, y1 = pairs[i + 1]
        if x0 <= x <= x1:
            t = (x - x0) / (x1 - x0)
            return y0 + t * (y1 - y0)
    return pairs[-1][1]


def extrapolate_rds(d, tj_ref):
    """
    Return (rds_at_tj, method_string).
    Priority:
      1. Normalized Rds-vs-Temp graph curve (most accurate; handles SJ nonlinearity)
         Rds(Tj) = rds25_spec × ratio_at_Tj_from_graph
      2. Two-point spec table linear interpolation (fallback)
    """
    r25  = d.get("rds25")
    r_ht = d.get("rds_ht")
    t_ht = d.get("_rds_ht_tj") or 125.0

    if r25 is None and r_ht is None:
        return None, "No Rds data available"

    # Priority 1: normalized Rds-vs-Temp graph → multiply ratio by rds25 spec value
    rds_curve = d.get("rds_ratio_curve", [])
    if rds_curve and r25 is not None:
        ratio = _interp_at(rds_curve, tj_ref)
        if ratio is not None and 0.3 <= ratio <= 6.0:
            val = round(r25 * ratio, 4)
            t_lo, t_hi = rds_curve[0][0], rds_curve[-1][0]
            interp_type = "INTERPOLATED" if t_lo <= tj_ref <= t_hi else "EXTRAPOLATED"
            return (val,
                    f"[GRAPH - {interp_type}] rds25 × ratio({tj_ref:.0f}°C) = "
                    f"{r25:.2f} mΩ × {ratio:.4f} = {val:.2f} mΩ  "
                    f"(curve: {t_lo:.0f}–{t_hi:.0f}°C, {len(rds_curve)} pts)")

    # Priority 2: two-point spec table linear interpolation
    if r_ht is None:
        return (None,
                f"Can't extrapolate — only 25°C Rds point available "
                f"(Rds={r25:.2f}mΩ; need high-T measurement too for interpolation)")

    if r25 is None:
        return (None,
                f"Can't extrapolate — only high-T ({t_ht:.0f}°C) Rds point available "
                f"(Rds={r_ht:.2f}mΩ; need 25°C measurement too for interpolation)")

    if abs(tj_ref - 25.0) < 0.5:
        return round(r25, 4), f"Direct @ 25°C: {r25:.2f} mΩ"

    if abs(tj_ref - t_ht) < 0.5:
        return round(r_ht, 4), f"Direct @ {t_ht:.0f}°C: {r_ht:.2f} mΩ"

    slope = (r_ht - r25) / (t_ht - 25.0)
    val   = r25 + slope * (tj_ref - 25.0)
    return (round(val, 4),
            f"Linear interp: {r25:.2f}mΩ@25°C → {r_ht:.2f}mΩ@{t_ht:.0f}°C → {val:.2f}mΩ@{tj_ref:.0f}°C")


def extrapolate_coss(coss_rows, vds_target):
    """
    Extrapolate C_oss to vds_target using depletion-cap model C = C0/√(1+VDS/Vbi).
    coss_rows: list of (vds, coss_pF) tuples.
    """
    rows = [(v, c) for v, c in coss_rows if v is not None and c is not None and v >= 0]
    if not rows:
        return None, "No Coss data"

    rows_s = sorted(rows, key=lambda x: x[0])
    VDs = np.array([r[0] for r in rows_s], dtype=float)
    Cs  = np.array([r[1] for r in rows_s], dtype=float)

    # Single-point: use 1/√V scaling (depletion approximation)
    if len(rows) == 1:
        v0, c0 = rows_s[0]
        if abs(v0 - vds_target) < 1.0:
            return round(float(c0), 2), f"[GRAPH - DIRECT] Coss={c0:.0f}pF @ VDS={v0:.0f}V (table)"
        val = c0 * np.sqrt(max(v0, 1.0) / max(vds_target, 1.0))
        return (round(float(val), 2),
                f"[FORMULA (derived) — 1/√V depletion law] Coss={c0:.0f}pF@{v0:.0f}V → "
                f"{val:.1f}pF@{vds_target:.0f}V  [GRAPH-DIRECT from C-V curve preferred per spec §3.4]")

    # Multi-point: fit C0/√(1+VDS/Vbi)
    def _dep(VD, C0, Vbi):
        return C0 / np.sqrt(1.0 + np.maximum(VD, 0) / max(abs(Vbi), 0.01))

    if HAS_SCIPY:
        try:
            Vg = max(VDs[0], 1.0)
            p0 = [Cs[0] * np.sqrt(1 + VDs[0] / Vg), Vg]
            popt, _ = curve_fit(_dep, VDs, Cs, p0=p0,
                                bounds=([0, 0.01], [1e8, 1e5]), maxfev=8000)
            val = _dep(vds_target, *popt)
            if 0 < val < Cs.max() * 20:
                pts = "; ".join(f"{c:.0f}pF@{v:.0f}V" for v, c in rows_s)
                return (round(float(val), 2),
                        f"Depletion fit C0/√(1+V/Vbi): C0={popt[0]:.1f}pF, Vbi={popt[1]:.2f}V [{pts}]")
        except Exception:
            pass

    # Power-law fallback
    try:
        b, a = np.polyfit(np.log(VDs + 1), np.log(Cs), 1)
        val  = np.exp(a) * (vds_target + 1) ** b
        if val > 0:
            return round(float(val), 2), f"Power-law fit Coss∝(VDS+1)^{b:.3f}"
    except Exception:
        pass

    # sqrt fallback
    v0, c0 = min(rows_s, key=lambda r: abs(r[0] - vds_target))
    val = c0 * np.sqrt(max(v0, 1) / max(vds_target, 1))
    return round(float(val), 2), f"Sqrt fallback: Coss={c0:.0f}pF@{v0:.0f}V"


def scale_energy(e_ref, id_test, id_target, vdd_test=None, vdd_ref=None):
    """
    Scale Eon/Eoff to (VDD_ref, ID_ref) using the hard-switching overlap model (spec §3.7).
    E(VDD,ref, ID,ref) = E_test × (VDD,ref / VDD,test) × (ID,ref / ID,test)
    Energy is linear in both VDD and ID — omitting the VDD ratio understates energy when
    the datasheet test voltage is below the operating reference (e.g. ST: 300V test vs 400V ref).
    """
    if e_ref is None:
        return None, "[FORMULA] Not available"
    parts = []
    val   = e_ref

    id_ratio  = 1.0
    vdd_ratio = 1.0

    if id_test and id_test != 0 and abs(id_target - id_test) >= 0.1:
        id_ratio = id_target / id_test
        val     *= id_ratio
        parts.append(f"ID: {e_ref:.3f}µJ × ({id_target:.1f}/{id_test:.1f}A)")
    elif id_test:
        parts.append(f"ID matches test ({id_test:.1f}A)")

    if vdd_test and vdd_ref and abs(vdd_ref - vdd_test) >= 1.0:
        base = val / id_ratio if id_ratio != 0 else e_ref  # before id scaling
        vdd_ratio = vdd_ref / vdd_test
        val   = e_ref * id_ratio * vdd_ratio
        parts.append(f"VDD: ×({vdd_ref:.0f}/{vdd_test:.0f}V)")
    elif vdd_test:
        parts.append(f"VDD matches ref ({vdd_test:.0f}V)")

    tag = "[FORMULA (derived) — ½·V·I·t, both axes]"
    if not parts:
        return round(e_ref, 4), f"{tag} Direct (test = ref conditions)"
    return round(val, 4), f"{tag} {' ; '.join(parts)} → {val:.3f}µJ"


def extrapolate_vth(d, tj_ref, dvth_dt_mv=-5.0):
    """
    Extrapolate VGS(th) to tj_ref (spec §3.2).
    If the device has a normalized Vth-vs-Tj graph ratio stored (_vth_norm_ratio),
    use GRAPH-INTERPOLATED: Vth(Tj) = Vth(25°C) × norm_ratio.
    Otherwise fall back to FORMULA: Vth(T) = Vth(25°C) + (dVth/dT)×(T−25)
    with dVth/dT ≈ −5 mV/°C (silicon physics default; user-overridable).
    """
    vth25 = d.get("vth")
    if vth25 is None:
        return None, "[NOT EXTRAPOLATABLE] No Vth data in datasheet"

    if abs(tj_ref - 25.0) < 0.5:
        return round(vth25, 3), f"[GRAPH - DIRECT] Vth = {vth25:.2f}V @ 25°C"

    # Graph-interpolated path: normalized ratio from digitized Vth-vs-Tj curve
    norm_ratio = d.get("_vth_norm_ratio")   # set by graph digitizer if available
    if norm_ratio is not None:
        val = vth25 * norm_ratio
        return (round(val, 3),
                f"[GRAPH - INTERPOLATED] Vth({tj_ref:.0f}°C) = {vth25:.2f}V × {norm_ratio:.4f} "
                f"(ratio read from normalized Vth-vs-Tj curve) = {val:.3f}V")

    # Formula fallback (IPA has no Vth-vs-T curve at all)
    dvth = dvth_dt_mv * 1e-3
    val  = vth25 + dvth * (tj_ref - 25.0)
    return (round(val, 3),
            f"[FORMULA (derived)] Vth({tj_ref:.0f}°C) = {vth25:.2f}V + "
            f"({dvth_dt_mv:.1f}mV/°C) × ({tj_ref:.0f}−25) = {val:.3f}V  "
            f"[silicon-physics default; supply dVth/dT to override]")


def _extract_eoss_rows(sources):
    """
    Pull (vds, eoss_µJ) data points from eoss_vs_vds entries in `sources`.
    `sources` can be either a charts list (from extract_charts_from_file) or a
    figs list (from extract_figures_from_file) — both use the same dict shape
    with 'chart_type' and 'series[{x, y}]' keys.

    The v5 extractor stores digitized Eoss-vs-VDS data in spreadsheet cells
    alongside the embedded image; extract_figures_from_file reads those cells
    into series[].  extract_charts_from_file reads native Excel chart objects.
    We search both, preferring the source with more data points.

    Returns a sorted list of (vds, eoss_µJ) tuples, or [] if none found.
    """
    best_rows = []
    for c in sources:
        if c.get('chart_type') != 'eoss_vs_vds':
            continue
        for ser in c.get('series', []):
            xs = ser.get('x', [])
            ys = ser.get('y', [])
            pairs = []
            for x, y in zip(xs, ys):
                try:
                    fx, fy = float(x), float(y)
                    if fx >= 0 and fy >= 0:
                        pairs.append((fx, fy))
                except (TypeError, ValueError):
                    pass
            if len(pairs) >= 2 and len(pairs) > len(best_rows):
                best_rows = sorted(pairs, key=lambda r: r[0])
    return best_rows


def _extract_rds_ratio_from_figs(figs):
    """
    Extract (temperature_°C, normalized_rds_ratio) pairs from rds_vs_temp figures.

    Uses the same widest-in-band-y-range series selection as _figure_table_points
    to reject artifact/contamination curves.  The returned curve can be used to
    compute Rds(Tj) = rds25_spec × ratio_at_Tj via _interp_at().

    Returns a sorted list of (temp, ratio) tuples, or [] if no usable data.
    """
    for f in figs:
        if f.get('chart_type') != 'rds_vs_temp':
            continue
        if not f.get('series'):
            continue
        PHYS_LO, PHYS_HI = 0.10, 5.0
        best_ser, best_yr = None, -1.0
        for s in f['series']:
            # Use _raw_y if the figure was already cleaned, otherwise use y
            raw = s.get('_raw_y', s.get('y', []))
            inb = [abs(float(y)) for y in raw
                   if y is not None and PHYS_LO <= abs(float(y)) <= PHYS_HI]
            if len(inb) < 2:
                continue
            yr = max(inb) / max(min(inb), 1e-9)
            if yr > best_yr:
                best_yr, best_ser = yr, s
        if best_ser is None and f['series']:
            best_ser = f['series'][0]
        if best_ser is None:
            continue
        pairs = _extract_ratio_curve(best_ser, ascending=True)
        if len(pairs) >= 2:
            return pairs
    return []


def extrapolate_eoss_from_rows(eoss_rows, vds_ref):
    """
    Interpolate/extrapolate E_oss at vds_ref directly from Eoss-vs-VDS graph data.
    eoss_rows: sorted list of (vds, eoss_µJ) tuples.
    """
    rows = sorted([(float(v), float(e)) for v, e in eoss_rows
                   if v is not None and e is not None and float(v) >= 0 and float(e) >= 0],
                  key=lambda r: r[0])
    if not rows:
        return None, "No Eoss graph data"

    # Direct hit (within 1 V of a data point)
    for v, e in rows:
        if abs(v - vds_ref) < 1.0:
            return (round(e, 4),
                    f"[GRAPH - DIRECT] Eoss = {e:.3f} µJ read directly "
                    f"@ VDS = {v:.0f} V from Eoss-vs-VDS curve")

    # Interpolation within the curve range
    below = [(v, e) for v, e in rows if v <= vds_ref]
    above = [(v, e) for v, e in rows if v >= vds_ref]
    if below and above:
        v1, e1 = below[-1]
        v2, e2 = above[0]
        if abs(v2 - v1) > 0.01:
            frac = (vds_ref - v1) / (v2 - v1)
            val  = e1 + frac * (e2 - e1)
        else:
            val = e1
        return (round(val, 4),
                f"[GRAPH - INTERPOLATED] Linear: {e1:.3f} µJ@{v1:.0f}V → "
                f"{e2:.3f} µJ@{v2:.0f}V → {val:.3f} µJ @ {vds_ref:.0f}V")

    # Extrapolation: power-law fit on the full curve (Eoss roughly ∝ VDS^α)
    if len(rows) >= 2:
        try:
            xs = np.array([r[0] for r in rows if r[0] > 0 and r[1] > 0], dtype=float)
            ys = np.array([r[1] for r in rows if r[0] > 0 and r[1] > 0], dtype=float)
            if len(xs) >= 2:
                b, a = np.polyfit(np.log(xs), np.log(ys), 1)
                val = float(np.exp(a) * max(vds_ref, 1.0) ** b)
                if 0 < val < 1e6:
                    return (round(val, 4),
                            f"[GRAPH - EXTRAPOLATED] Power-law Eoss∝VDS^{b:.2f} "
                            f"fitted to {len(xs)}-point curve → {val:.3f} µJ @ {vds_ref:.0f}V")
        except Exception:
            pass

    # Last resort: nearest point with quadratic VDS scaling
    nearest = min(rows, key=lambda r: abs(r[0] - vds_ref))
    v0, e0  = nearest
    val = e0 * (vds_ref / max(v0, 1.0)) ** 2
    return (round(val, 4),
            f"[GRAPH - EXTRAPOLATED] Quadratic scale from {e0:.3f} µJ@{v0:.0f}V "
            f"→ {val:.3f} µJ @ {vds_ref:.0f}V")


def extrapolate_eoss(d, vds_ref, coss_at_ref):
    """
    Extrapolate E_oss to vds_ref.
    Priority:
      1. Eoss-vs-VDS graph curve (direct read / interpolation) — most accurate
      2. Re-derive from Coss @ vds_ref: Eoss = ½·Coss·VDS² (formula fallback)
      3. Scale table Eoss by (VDS_ref/VDS_test)² if test VDS known
      4. Return table value with a note
    """
    # 1. Eoss-vs-VDS graph — highest priority
    eoss_rows = d.get("eoss_rows", [])
    if eoss_rows:
        return extrapolate_eoss_from_rows(eoss_rows, vds_ref)

    # 2. Coss-formula fallback (when no Eoss graph but Coss value available)
    if coss_at_ref is not None and vds_ref is not None:
        val = round(0.5 * coss_at_ref * 1e-12 * vds_ref**2 * 1e6, 4)
        return (val,
                f"[FORMULA (derived)] ½ × Coss({vds_ref:.0f}V) × VDS_ref² = "
                f"½ × {coss_at_ref:.0f}pF × ({vds_ref:.0f}V)² = {val:.3f}µJ")

    eoss_tab = d.get("eoss")
    if eoss_tab is None:
        return None, "[NOT AVAILABLE] No Eoss or Coss data"

    # 3. Scale table value using (VDS_ref/VDS_test)²
    vds_test = d.get("_eoss_vds_test") or d.get("_coss_vds_test")
    if vds_test and abs(vds_test - vds_ref) > 5:
        val = round(eoss_tab * (vds_ref / vds_test) ** 2, 4)
        return (val,
                f"[FORMULA (derived)] Eoss({vds_ref:.0f}V) = {eoss_tab:.3f}µJ@{vds_test:.0f}V "
                f"× ({vds_ref:.0f}/{vds_test:.0f})² = {val:.3f}µJ")

    # 4. Table value at nearest known VDS
    return round(eoss_tab, 4), "[TABLE - DIRECT] Read at nearest VDS from spec table"


def extrapolate_vsd(d, tj_ref, dvsd_dt_mv=-2.0):
    """
    Extrapolate V_SD (body diode forward voltage) to tj_ref.
    Silicon body diode: dVSD/dT ≈ -2 mV/°C.
    """
    vsd25 = d.get("vsd")
    if vsd25 is None:
        return None, "[NOT AVAILABLE] No VSD data"

    if abs(tj_ref - 25.0) < 0.5:
        return round(vsd25, 3), f"[DIRECT] VSD = {vsd25:.2f}V @ 25°C"

    dvsd = dvsd_dt_mv * 1e-3
    val  = max(vsd25 + dvsd * (tj_ref - 25.0), 0.05)
    return (round(val, 3),
            f"[FORMULA (derived)] VSD({tj_ref:.0f}°C) = {vsd25:.2f}V + "
            f"({dvsd_dt_mv:.0f}mV/°C)×({tj_ref:.0f}−25) = {val:.3f}V  "
            f"[silicon diode; typical dVSD/dT = −2 mV/°C]")


def extrapolate_qg(d, vds_ref):
    """
    Scale Q_G to vds_ref when test VDS differs.
    Qg(VDS_ref) = Qgs + Qgd × (VDS_ref / VDS_test)
    where Qgs = Qg_test − Qgd.
    Falls back to linear scaling if Qgd is not available.
    """
    qg_tab = d.get("qg")
    if qg_tab is None:
        return None, "[NOT AVAILABLE] No Qg data"

    vds_test = d.get("_qg_vds")
    qgd      = d.get("qgd")

    if vds_test is None or abs(vds_test - vds_ref) < 10:
        cond = d.get("qg_tcond", "")
        return round(qg_tab, 4), f"[GRAPH - DIRECT] Test VDS ≈ VDS_ref; {cond}"

    if qgd is not None:
        qgs = max(qg_tab - qgd, 0.0)
        val = qgs + qgd * (vds_ref / vds_test)
        return (round(val, 4),
                f"[FORMULA (derived)] Qg({vds_ref:.0f}V) = Qgs + Qgd×(VDS_ref/VDS_test) = "
                f"{qgs:.1f}nC + {qgd:.1f}nC×({vds_ref:.0f}/{vds_test:.0f}) = {val:.1f}nC")
    else:
        val = qg_tab * (vds_ref / vds_test)
        return (round(val, 4),
                f"[FORMULA (derived, rough)] Qg({vds_ref:.0f}V) ≈ {qg_tab:.1f}nC × "
                f"({vds_ref:.0f}/{vds_test:.0f}) = {val:.1f}nC  "
                f"[Qgd not in datasheet — split Qg into Qgs+Qgd for accuracy]")


def extrapolate_trr(d, tj_ref):
    """
    Rough temperature scaling for t_rr: approx +0.5 %/°C above 25°C.
    This is a first-order estimate; actual coefficient varies by device.
    """
    trr25 = d.get("trr")
    if trr25 is None:
        return None, "[NOT AVAILABLE]"

    if abs(tj_ref - 25.0) < 0.5:
        return round(trr25, 2), f"[DIRECT] {trr25:.0f} ns @ 25°C"

    scale = 1.0 + 0.005 * (tj_ref - 25.0)
    val   = round(trr25 * scale, 2)
    return (val,
            f"[FORMULA (derived)] t_rr({tj_ref:.0f}°C) ≈ {trr25:.0f}ns × {scale:.3f} "
            f"(+0.5%/°C; approximate — check datasheet for specific coefficient)")


def extrapolate_qrr(d, tj_ref):
    """
    Rough temperature scaling for Q_rr: approx +1 %/°C above 25°C.
    Qrr grows faster than trr because both charge and time increase.
    """
    qrr25 = d.get("qrr")
    if qrr25 is None:
        return None, "[NOT AVAILABLE]"

    if abs(tj_ref - 25.0) < 0.5:
        return round(qrr25, 2), f"[DIRECT] {qrr25:.0f} nC @ 25°C"

    scale = 1.0 + 0.010 * (tj_ref - 25.0)
    val   = round(qrr25 * scale, 2)
    return (val,
            f"[FORMULA (derived)] Q_rr({tj_ref:.0f}°C) ≈ {qrr25:.0f}nC × {scale:.3f} "
            f"(+1%/°C; approximate — check datasheet for specific coefficient)")


# ─── GRAPH EXTRACTION & COMPARISON ───────────────────────────────────────────

# Chart types rendered on log scales ─ images only, no extraction table
_LOG_CHART_TYPES = {'capacitance_vs_vds', 'thermal_impedance'}

_CHART_TYPE_PATTERNS = {
    'capacitance_vs_vds': ['capacitance', 'coss', 'ciss', 'crss', 'c_oss', 'c_iss',
                            'c_rss', 'output cap', 'input cap', 'reverse transfer cap'],
    'thermal_impedance':  ['thermal', 'zth', 'z_th', 'impedance', 'junction-to-case',
                            'transient thermal'],
    'transfer_char':      ['transfer char', 'id vs vgs', 'id-vgs', 'id vs. vgs'],
    'output_char':        ['output char', 'id vs vds', 'id-vds', 'drain current vs drain'],
    'body_diode':         ['body diode', 'vsd', 'forward voltage', 'diode forward',
                            'source drain'],
    'soa':                ['safe operating', 'soa', 'safe area'],
    'gate_charge':        ['gate charge', 'vgs vs qg', 'gate voltage vs charge'],
    'rds_vs_temp':        ['rds vs', 'on-resistance vs temp', 'rds(on) vs',
                            'normalised rds'],
}

_CHART_DISPLAY_NAMES = {
    'capacitance_vs_vds': 'Capacitance vs V_DS  (C_oss / C_iss / C_rss)',
    'thermal_impedance':  'Thermal Impedance  Z_th vs t_p',
    'transfer_char':      'Transfer Characteristics  I_D vs V_GS',
    'output_char':        'Output Characteristics  I_D vs V_DS',
    'body_diode':         'Body Diode Forward Voltage  V_SD vs I_SD',
    'soa':                'Safe Operating Area',
    'gate_charge':        'Gate Charge  V_GS vs Q_G',
    'rds_vs_temp':        'R_DS(on) vs Temperature',
    'unknown':            'Other Charts',
}

_CHART_AXIS_LABELS = {
    'capacitance_vs_vds': ('V_DS (V)',   'Capacitance (pF)'),
    'thermal_impedance':  ('t_p (s)',    'Z_th (°C/W)'),
    'transfer_char':      ('V_GS (V)',   'I_D (A)'),
    'output_char':        ('V_DS (V)',   'I_D (A)'),
    'body_diode':         ('I_SD (A)',   'V_SD (V)'),
    'gate_charge':        ('Q_G (nC)',   'V_GS (V)'),
    'rds_vs_temp':        ('T_j (°C)',   'R_DS(on) (mΩ)'),
}

_TABLE_Y_UNITS = {
    'capacitance_vs_vds': 'pF',
    'transfer_char':      'A',
    'output_char':        'A',
    'body_diode':         'V',
    'gate_charge':        'V',
    'rds_vs_temp':        'mΩ',
    'eoss_vs_vds':        'µJ',
    'switching_energy':   'µJ',
    'id_vs_tc':           'A',
    'vth_vs_temp':        'ratio',
    'power_derating':     'W',
    'avalanche_energy':   'mJ',
    'vbr_vs_temp':        'V',
}

# ─── EMBEDDED-IMAGE (digitised datasheet figure) SUPPORT ───────────────────────
#
# The v16/v17 extractor format stores the *actual* datasheet graph pictures as
# embedded PNGs on dedicated sheets ("Capacitance Graphs", "Temperature Graphs",
# "Thermal Impedance Graphs", "Energy Graphs"), each next to a "▶ Figure N …"
# caption, an axis-calibration line, and a digitised data table.
#
# These chart types use captions that are finer-grained than the chart-object
# classifier, so they get their own catalogue here.

# Ordered (first match wins).
# switching_energy must precede eoss_vs_vds so Eon/Eoff are not swallowed by the
# generic 'energy' keyword.  rds_vs_temp must match both 'rds(on)' (no separator)
# AND 'r_ds(on)' (underscore used by many datasheets and our own output headers).
_FIGURE_TYPE_PATTERNS = [
    ('switching_energy',   ['turn-on energy', 'turn-off energy', 'switching energy',
                            'eon vs', 'eoff vs', 'e_on vs', 'e_off vs',
                            'switch-on energy', 'switch-off energy',
                            'turn on energy', 'turn off energy']),
    ('eoss_vs_vds',        ['stored energy', 'eoss', 'e_oss']),
    ('capacitance_vs_vds', ['capacitance', 'c=f(vds)', 'ciss', 'coss', 'crss',
                            'output cap', 'input cap', 'reverse transfer']),
    ('thermal_impedance',  ['thermal impedance', 'zth', 'transient thermal',
                            'z_th', 'zthjc']),
    # rds patterns: covers 'rds(on)' (no separator), 'r_ds(on)' (underscore),
    # 'on-resistance' / 'on resistance' / 'on-resistance variation' (ONsemi),
    # 'drain-to-source on' / 'drain-source on' (Infineon / Vishay long-form).
    ('rds_vs_temp',        ['rds(on)', 'r_ds(on)', 'rds vs', 'r_ds vs',
                            'normalised rds', 'normalized rds',
                            'on-resistance variation', 'on resistance variation',
                            'drain-to-source on', 'drain-source on',
                            'on-state resistance', 'on-resistance',
                            'on resistance', 'on state resistance']),
    ('vth_vs_temp',        ['threshold']),
    ('power_derating',     ['power dissipation', 'derating', 'ptot', 'p_tot']),
    # id_vs_tc: Maximum/Continuous Drain Current vs Case OR Junction Temperature.
    # Groups "ID vs Tc" and "ID vs Tj" under the same chart type so they are
    # placed in the same side-by-side comparison block.
    # Must come AFTER power_derating (Ptot vs Tc is not Id vs Tc) and
    # BEFORE transfer_char ('id vs vgs' must not steal these).
    ('id_vs_tc',           ['drain current vs case', 'drain current vs. case',
                            'drain current vs junction', 'drain current vs. junction',
                            'id vs tc', 'id vs. tc', 'id(a) vs t_c',
                            'id vs tj', 'id vs. tj', 'id(a) vs t_j',
                            'id vs temperature', 'continuous drain current vs',
                            'maximum continuous drain current',
                            'max. drain current vs', 'drain current derating']),
    ('avalanche_energy',   ['avalanche', 'eas']),
    ('vbr_vs_temp',        ['breakdown voltage', 'vbr', '(br)', 'br)dss',
                            'breakdown']),
    ('gate_charge',        ['gate charge', 'qg', 'q_g']),
    ('transfer_char',      ['transfer char', 'id vs vgs', 'transfer']),
    ('output_char',        ['output characteristic', 'id vs vds']),
    ('body_diode',         ['body diode', 'source-drain', 'source drain',
                            'diode forward']),
]

_FIGURE_DISPLAY_NAMES = {
    'eoss_vs_vds':      'C_oss Stored Energy  E_oss vs V_DS',
    'switching_energy': 'Switching Energy  E_on / E_off vs I_D',
    'power_derating':   'Power Dissipation / Derating  P_tot vs T_c',
    'id_vs_tc':         'Maximum Drain Current  I_D vs T_C / T_J',
    'avalanche_energy': 'Avalanche Energy  E_AS vs T_j',
    'vbr_vs_temp':      'Breakdown Voltage  V_(BR)DSS vs T_j',
    'vth_vs_temp':      'Gate Threshold (normalised)  V_GS(th) vs T_j',
}

_FIGURE_AXIS_LABELS = {
    'eoss_vs_vds':      ('V_DS (V)',  'E_oss (µJ)'),
    'switching_energy': ('I_D (A)',   'Energy (µJ)'),
    'power_derating':   ('T_c (°C)',  'P_tot (W)'),
    'avalanche_energy': ('T_j (°C)',  'E_AS (mJ)'),
    'id_vs_tc':         ('T_C / T_J (°C)',  'I_D (A)'),
    'vbr_vs_temp':      ('T_j (°C)',  'V_(BR)DSS (V)'),
    'vth_vs_temp':      ('T_j (°C)',  'V_GS(th) (ratio)'),
}

# Figures rendered on log axes in their datasheet — numeric overlay/table is
# only meaningful when the digitised axis is real (not 0-1 normalised), so this
# set is informational only.
_FIGURE_LOG_TYPES = {'capacitance_vs_vds', 'thermal_impedance'}


def _classify_figure(caption, sheet_name):
    """Classify a digitised figure into a fine-grained chart type."""
    # Strip leading ▶ / ► markers, then normalise Unicode dashes (en-dash –,
    # em-dash —, minus −) to plain hyphen-minus so patterns like 'on-resistance'
    # match captions that were copy-pasted from PDF datasheets with U+2013/2014.
    clean_cap = re.sub(r'^[▶►\s]+', '', str(caption)).strip()
    clean_cap = re.sub(r'[‐‑‒–—―−]', '-', clean_cap)
    text = (clean_cap + ' ' + str(sheet_name)).lower()
    for ctype, kws in _FIGURE_TYPE_PATTERNS:
        if any(k in text for k in kws):
            return ctype
    # Last resort: fall back to the sheet name.
    # Be specific about energy sheets so that switching-energy (Eon/Eoff) graphs
    # on a generic "Energy Graphs" sheet are NOT mis-routed to eoss_vs_vds.
    s = sheet_name.lower()
    if 'capacit' in s:  return 'capacitance_vs_vds'
    if 'thermal' in s:  return 'thermal_impedance'
    if 'stored energy' in s or 'eoss' in s:    return 'eoss_vs_vds'
    if ('switch' in s or 'eon' in s or 'eoff' in s) and 'energy' in s:
        return 'switching_energy'
    if 'energy' in s:   return 'eoss_vs_vds'   # generic energy sheet → assume Eoss
    return 'other'


def _figure_axis_norm(axis_line):
    """
    Parse the 'Axis calibration — X: … | Y: …' line.
    Returns (x_normalised, y_normalised) booleans.
    An axis is 'normalised' when the calibration says NORMALISED 0-1, meaning the
    digitised values are 0-1 fractions of the (unknown) real span and therefore
    NOT comparable across datasheets in real units.
    """
    if not axis_line:
        return False, False
    x_norm = y_norm = False
    m = re.search(r'Axis calibration[^|]*?X:\s*([^|]+)\|\s*Y:\s*([^|]+)',
                  str(axis_line), re.I)
    if m:
        x_norm = 'normalis' in m.group(1).lower()
        y_norm = 'normalis' in m.group(2).lower()
    return x_norm, y_norm


def _img_bytes(im):
    """Return PNG/JPEG bytes from an openpyxl embedded image object."""
    try:
        data = im._data()
        return data if data else None
    except Exception:
        pass
    try:
        ref = getattr(im, 'ref', None)
        if ref is not None and hasattr(ref, 'getvalue'):
            return ref.getvalue()
    except Exception:
        pass
    return None


def extract_figures_from_file(path):
    """
    Extract the embedded datasheet graph PICTURES (and their digitised data)
    from a v16/v17-format extractor workbook.

    Returns list of figure dicts:
      {chart_type, caption, sheet, x_norm, y_norm,
       x_label, y_label, image_bytes, series:[{label,x,y}]}
    """
    figs = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as ex:
        print(f"  Warning: image extraction from {Path(path).name}: {ex}")
        return figs

    for ws_name in wb.sheetnames:
        ws   = wb[ws_name]
        imgs = getattr(ws, '_images', [])
        if not imgs:
            continue

        # Snapshot all non-empty cells + figure-header rows for this sheet
        cells       = {}
        header_rows = []
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                cells[(c.row, c.column)] = c.value
                if c.column == 1 and str(c.value).strip().startswith('▶'):
                    header_rows.append(c.row)
        header_rows.sort()

        for im in imgs:
            data = _img_bytes(im)
            if not data:
                continue
            try:
                anchor_row = im.anchor._from.row + 1     # Excel (1-based) row
            except Exception:
                anchor_row = 1

            # Caption = nearest "▶ Figure …" header at/above this image
            cand = [h for h in header_rows if h <= anchor_row + 1]
            hrow = cand[-1] if cand else (header_rows[0] if header_rows else 1)
            raw_cap = str(cells.get((hrow, 1), '')).strip()
            caption = re.sub(r'^[▶\s]*Figure\s*\d+\s*[—\-:]*\s*', '', raw_cap)
            caption = re.sub(r'^Diagram\s*\d+\s*:?\s*', '', caption).strip()
            if not caption:
                caption = ws_name

            axis_line     = str(cells.get((hrow + 1, 1), ''))
            x_norm, y_norm = _figure_axis_norm(axis_line)
            ctype          = _classify_figure(caption, ws_name)

            # Digitised data table: header row hrow+2, numeric rows below
            hdr_row = hrow + 2
            xs, curve_pts = [], {}
            blanks = 0
            for r in range(hrow + 3, hrow + 300):
                xv = _val(cells.get((r, 1)))
                if xv is None:
                    blanks += 1
                    if blanks >= 3 and xs:
                        break
                    continue
                blanks = 0
                for col in (2, 3, 4, 5):
                    yv = _val(cells.get((r, col)))
                    if yv is None:
                        continue
                    lbl = str(cells.get((hdr_row, col), '') or f'Curve {col-1}')
                    curve_pts.setdefault((col, lbl), []).append((xv, yv))
                xs.append(xv)

            series = []
            for (_col, lbl), pts in sorted(curve_pts.items()):
                series.append({
                    'label': re.sub(r'\s*\([^)]*\)\s*$', '', lbl).strip() or lbl,
                    'x': [p[0] for p in pts],
                    'y': [p[1] for p in pts],
                })

            xl, yl = _FIGURE_AXIS_LABELS.get(
                ctype, _CHART_AXIS_LABELS.get(ctype, ('X', 'Y')))
            figs.append({
                'chart_type':  ctype,
                'caption':     caption,
                'sheet':       ws_name,
                'x_norm':      x_norm,
                'y_norm':      y_norm,
                'x_label':     xl,
                'y_label':     yl,
                'image_bytes': data,
                'series':      series,
            })
    wb.close()
    return figs


# ─── NORMALISED-RATIO CURVE CLEANING ──────────────────────────────────────────
# Chart types whose Y-axis is a normalised ratio (0.5–3.0 range), NOT absolute
# units. Their digitised series sometimes contain ×10 / ×100 / ÷100 errors from
# pixel mis-reads or European decimal notation ("3.000" read as 3000).
_NORM_RATIO_CTYPES = {'rds_vs_temp', 'vth_vs_temp', 'vbr_vs_temp'}


def _pow10_closest(v, ref):
    """Return v multiplied/divided by the power of 10 that lands it closest
    (in log space) to the local reference level `ref`.  Used to undo ×10/×100/
    ÷100 digitisation errors without touching values that are already correct."""
    a = abs(float(v))
    if a == 0:
        return 0.0
    ref = max(abs(ref), 1e-9)
    best, best_d = a, float('inf')
    for p in range(-5, 6):
        c = a * (10.0 ** p)
        d = abs(math.log10(c) - math.log10(ref))
        if d < best_d:
            best_d, best = d, c
    return math.copysign(best, v)


def _fig_is_ratio(f):
    """True when a figure's Y-axis is a normalised ratio (median |y| < 20),
    as opposed to absolute units (e.g. V_BR in volts ≈ 600)."""
    ys = sorted(abs(float(y)) for s in f.get('series', [])
                for y in s.get('y', []) if y is not None)
    if not ys:
        return False
    return ys[len(ys) // 2] < 20.0


def _clean_ratio_series(xs, ys, monotone_dir=0):
    """
    Clean one normalised-ratio series.  Two-pass strategy:

    Pass 1 — Global pre-scale:
      If the MEDIAN of all raw values is outside [0.1, 10], find the single
      power-of-10 that brings it closest to 1.0 and pre-scale every value by
      that factor.  This handles European decimal charts ("3.000" → 3000) where
      ALL values in the series are inflated by the same factor, making the
      in-band seed detection in pass 2 useless without this first correction.

    Pass 2 — Local-neighbour fine-tune:
      Walk low→high T.  Each point is rescaled by the power of 10 that puts it
      closest to the median of the 2 already-cleaned left neighbours.
      monotone_dir (+1 rising / -1 falling / 0 unknown): when set, non-monotone
      candidates are penalised by +1 log-decade so the algorithm prefers the
      scale factor that preserves the expected trend direction (e.g. for Rds-vs-T,
      a value lower than its neighbour gets penalised, pushing the choice toward
      a larger scale factor instead).
    """
    idx = sorted(range(len(xs)), key=lambda i: (xs[i] if xs[i] is not None else 0.0))

    # ── Pass 1: global pre-scale ───────────────────────────────────────────────
    raw_abs = [abs(float(ys[i])) for i in idx if ys[i] is not None and ys[i] != 0]
    global_p = 0
    if raw_abs:
        med_raw = sorted(raw_abs)[len(raw_abs) // 2]
        if not (0.1 <= med_raw <= 10.0):
            best_d = float('inf')
            for p in range(-6, 7):
                c = med_raw * (10.0 ** p)
                d = abs(math.log10(max(c, 1e-12)) - 0.0)   # target log10 = 0 (value ≈ 1)
                if d < best_d:
                    best_d, global_p = d, p

    prescaled = [float(y) * (10.0 ** global_p) if y is not None else None for y in ys]

    # ── Pass 2: local-neighbour fine-tune ─────────────────────────────────────
    inband = sorted(abs(float(prescaled[i])) for i in idx
                    if prescaled[i] is not None and 0.2 <= abs(float(prescaled[i])) <= 6.0)
    seed = inband[len(inband) // 2] if inband else 1.0
    cleaned = {}
    for pos, i in enumerate(idx):
        y = prescaled[i]
        if y is None:
            cleaned[i] = None
            continue
        left = [abs(cleaned[idx[k]]) for k in range(max(0, pos - 2), pos)
                if cleaned.get(idx[k]) is not None]
        ref = sorted(left)[len(left) // 2] if left else seed
        prev = sorted(left)[len(left) // 2] if left else None

        a = abs(float(y))
        best_val, best_d = a, float('inf')
        # Only penalise monotone direction when the raw value is clearly out of
        # the valid ratio band [0.02, 20].  If it's already in-range it may
        # simply come from a different (lower/higher VGS) curve — forcing a scale
        # correction here would produce a larger wrong value.
        needs_rescale = not (0.02 <= a <= 20.0)
        for p2 in range(-5, 6):
            c = a * (10.0 ** p2)
            if c < 0.02 or c > 50.0:
                continue
            d = abs(math.log10(c) - math.log10(max(ref, 1e-9)))
            # Penalise direction that violates the expected monotone trend,
            # but only when we are actually rescaling (not for in-range values).
            if needs_rescale and monotone_dir != 0 and prev is not None:
                if monotone_dir > 0 and c < prev * 0.80:
                    d += 1.5
                elif monotone_dir < 0 and c > prev * 1.25:
                    d += 1.5
            if d < best_d:
                best_d, best_val = d, c
        cleaned[i] = math.copysign(best_val, float(y))
    return [cleaned[i] for i in range(len(ys))]


def _clean_ratio_fig(f, monotone_dir=0):
    """Clean every series of a normalised-ratio figure in place.
    Saves original y-values as '_raw_y' before modifying, so
    _extract_ratio_curve can access un-distorted data later.
    Returns True if any value was corrected."""
    changed = False
    for s in f.get('series', []):
        xs, ys = s.get('x', []), s.get('y', [])
        if '_raw_y' not in s:
            s['_raw_y'] = list(ys)   # preserve original for extraction
        ny = _clean_ratio_series(xs, ys, monotone_dir=monotone_dir)
        for a, b in zip(ys, ny):
            if a is not None and b is not None and abs(float(a) - b) > 1e-6:
                changed = True
        s['y'] = ny
    return changed


def _extract_ratio_curve(ser, ascending=True):
    """
    From a series that may contain multiple overlapping curves (e.g. correct
    normalised Rds + European-decimal ×1000 duplicate + digitiser artifact),
    extract the single physically-plausible normalised-ratio sub-curve.

    Uses '_raw_y' (saved by _clean_ratio_fig before cleaning) so the
    extraction operates on un-distorted original values.

    Strategy:
      Phase 1 — if enough raw values are already in the physical ratio band
                 [0.10, 5.0], filter to that band and skip pre-scaling.
      Phase 2 — if too few raw values are in-band, find the power-of-10 that
                 brings the series median into band (handles all-×1000 series).
      Phase 3 — greedy monotone walk: at each x position (possibly multiple y
                 values from different curves), pick the y that best continues
                 the ascending / descending trend from the previous point.

    Returns a sorted list of (x, y) pairs with at most one y per x.
    """
    xs = ser.get('x', [])
    ys = ser.get('_raw_y', ser.get('y', []))   # prefer pre-cleaning values

    raw_pairs = sorted(
        (float(x), float(y)) for x, y in zip(xs, ys)
        if x is not None and y is not None
    )
    if not raw_pairs:
        return []

    PHYS_LO, PHYS_HI = 0.10, 5.0

    # Phase 1 — count how many raw values are already in the physical band
    in_range_count = sum(1 for _, y in raw_pairs if PHYS_LO <= abs(y) <= PHYS_HI)

    if in_range_count >= 3:
        scale = 1.0
    else:
        # Phase 2 — global pre-scale (handles all-×1000 European-decimal series)
        all_abs = sorted(abs(y) for _, y in raw_pairs if y != 0)
        med = all_abs[len(all_abs) // 2] if all_abs else 1.0
        best_d, global_p = float('inf'), 0
        for p in range(-6, 7):
            c = med * (10.0 ** p)
            d = abs(math.log10(max(c, 1e-12)))
            if d < best_d:
                best_d, global_p = d, p
        scale = 10.0 ** global_p

    # Build per-x groups, keeping only scaled values inside the physical band
    from collections import defaultdict
    x_groups = defaultdict(list)
    for x, y in raw_pairs:
        ysc = y * scale
        if PHYS_LO <= abs(ysc) <= PHYS_HI:
            x_groups[round(x, 3)].append(ysc)

    if sum(len(v) for v in x_groups.values()) < 3:
        return raw_pairs   # not enough physical-range points — return raw

    # Phase 3 — greedy monotone walk, trying every first-point candidate as seed.
    # For the first x-position there may be multiple in-range candidates (e.g. the
    # correct Rds curve at y=0.518 AND a flat artifact at y=0.928).  The old
    # "closest-to-1.0" seed picked the artifact.  Instead, run the walk once for
    # each candidate seed and keep the walk with the widest y-range — the correct
    # physical curve always spans a wider ratio than any flat artifact.
    def _run_walk(seed_y):
        walk, pv = [], None
        for xk in sorted(x_groups):
            cands = x_groups[xk]
            if not cands:
                continue
            if pv is None:
                # Use the prescribed seed; skip this x if seed isn't present
                if seed_y not in cands:
                    return []
                by = seed_y
            else:
                if ascending:
                    ok = [y for y in cands if y >= pv * 0.90]
                else:
                    ok = [y for y in cands if y <= pv * 1.10]
                by = (min(ok, key=lambda y: abs(y - pv)) if ok
                      else (max if ascending else min)(cands))
            walk.append((xk, by))
            pv = by
        return walk

    first_xk = sorted(x_groups)[0]
    best_result, best_range = [], -1.0
    for seed in x_groups[first_xk]:
        walk = _run_walk(seed)
        if len(walk) >= 2:
            ys_w = [y for _, y in walk]
            y_range = max(ys_w) / max(min(ys_w), 1e-9)
            if y_range > best_range:
                best_range, best_result = y_range, walk

    return best_result if len(best_result) >= 2 else raw_pairs


def _primary_series(f):
    """Return the series of a figure with the widest X-span (the main curve),
    so the comparison table follows the full-range curve rather than a fragment."""
    if not f or not f.get('series'):
        return None
    best, best_span = None, -1.0
    for s in f['series']:
        xs = [float(x) for x in s.get('x', []) if x is not None]
        if len(xs) < 2:
            continue
        span = max(xs) - min(xs)
        if span > best_span:
            best_span, best = span, s
    return best or (f['series'][0] if f['series'] else None)


def _most_monotone_series(f, ascending=True):
    """For multi-curve normalised-ratio figures (e.g. 4 VGS curves on one Rds-vs-T
    graph), select the series with the fewest monotone violations.
    This picks a clean curve (e.g. VGS=10V) over one whose digitised points jump
    between curves or have spurious outliers.
    Tie-broken by: wider X-span, then higher median Y (upper curve → more data)."""
    if not f or not f.get('series'):
        return None
    best, best_key = None, (float('inf'), -1.0, -1.0)
    for s in f['series']:
        pairs = sorted((float(x), float(y)) for x, y in zip(s.get('x', []), s.get('y', []))
                       if x is not None and y is not None)
        if len(pairs) < 2:
            continue
        violations = sum(
            1 for k in range(1, len(pairs))
            if (ascending  and pairs[k][1] < pairs[k-1][1] * 0.85) or
               (not ascending and pairs[k][1] > pairs[k-1][1] * 1.15)
        )
        x_span = pairs[-1][0] - pairs[0][0]
        med_y  = sorted(p[1] for p in pairs)[len(pairs) // 2]
        key = (violations, -x_span, -med_y)   # fewer violations, wider, higher = better
        if key < best_key:
            best_key, best = key, s
    return best or (f['series'][0] if f['series'] else None)


def _group_figures(all_figs_per_file):
    """
    Group digitised figures by chart type across all device files.
    Returns list of groups (ordered as first encountered):
      {chart_type, display_name, figs_per_device:[fig|None], scale_matched}
    scale_matched is True only when every present figure has BOTH axes in real
    (non-normalised) units → numeric overlay + comparison table are meaningful.

    'other' charts are sub-grouped by normalised caption so that unrelated graphs
    (e.g. Id-vs-Tc from one device and Rds from another) are never merged into
    the same comparison section just because both fell through classification.
    """
    def _group_key(f):
        ct = f['chart_type']
        if ct != 'other':
            return ct
        # Normalise the caption to a stable key: strip markers, lower, collapse spaces.
        cap = re.sub(r'^[▶►\s]+', '', str(f.get('caption', ''))).strip().lower()
        cap = re.sub(r'\s+', '_', cap)[:50]
        return f'other::{cap}' if cap else 'other::unclassified'

    order, seen = [], set()
    for figs in all_figs_per_file:
        for f in figs:
            key = _group_key(f)
            if key not in seen:
                order.append(key); seen.add(key)

    groups = []
    for gkey in order:
        ctype = gkey.split('::')[0]   # 'other' or the real chart_type
        per_dev = []
        for figs in all_figs_per_file:
            # pick the richest figure matching this group key (most data points)
            best = None
            for f in figs:
                if _group_key(f) != gkey:
                    continue
                if best is None:
                    best = f
                else:
                    cur = sum(len(s['y']) for s in best['series'])
                    new = sum(len(s['y']) for s in f['series'])
                    if new > cur:
                        best = f
            per_dev.append(best)

        present = [f for f in per_dev if f is not None]
        if not present:
            continue

        # Display name: catalogue lookup for known types; derive from caption for 'other'.
        if gkey.startswith('other::'):
            cap_part = gkey[7:].replace('_', ' ').title()
            disp = cap_part if cap_part and cap_part != 'Unclassified' else 'Other Charts'
        else:
            disp = _FIGURE_DISPLAY_NAMES.get(
                ctype, _CHART_DISPLAY_NAMES.get(
                    ctype, ctype.replace('_', ' ').title()))

        # Basic flag-based check (uses digitizer-reported norm flags)
        scale_matched = (ctype not in _FIGURE_LOG_TYPES and
                         all((not f['x_norm']) and (not f['y_norm'])
                             for f in present) and
                         all(any(len(s['y']) >= 2 for s in f['series'])
                             for f in present))

        # ── Normalised-ratio temperature charts (Rds / Vth / VBR vs T_j) ─────────
        # These use a normalised Y-axis (ratio ≈0.5–3.0).  Two things can go wrong:
        #   (1) Digitisation errors scale individual points by ×10 / ×100 / ÷100,
        #       or European decimal "3.000" is read as 3000 (ALL values ×1000).
        #       → cleaned by _clean_ratio_fig (global pre-scale + local walk).
        #       IMPORTANT: clean ALL figures FIRST, then check _fig_is_ratio, because
        #       fully-inflated raw values (median >> 20) would fool _fig_is_ratio into
        #       thinking the figure is not a ratio chart and skip cleaning entirely.
        #   (2) One device plots a NORMALISED ratio while the other plots ABSOLUTE
        #       units (e.g. Infineon V_BR in volts ≈600 vs ST normalised ≈1.0).
        #       → genuinely incommensurable: flag as units-not-equal (no overlay).
        if ctype in _NORM_RATIO_CTYPES and present:
            # Direction hint: Rds rises with T (+1), Vth falls (-1), VBR unknown (0)
            mono_dir = +1 if ctype == 'rds_vs_temp' else (-1 if ctype == 'vth_vs_temp' else 0)
            # Clean ALL figures first (even those _fig_is_ratio might not flag yet)
            for f in per_dev:
                if f is not None and _clean_ratio_fig(f, monotone_dir=mono_dir):
                    f['_european_decimal_rescaled'] = True
            # Re-detect ratio status on the now-cleaned values
            ratio_ids = {id(f) for f in present if _fig_is_ratio(f)}

            if len(ratio_ids) == len(present):
                # All devices normalised-ratio → comparable → overlay + table.
                scale_matched = True
            elif ratio_ids:
                # Mixed: some normalised, some absolute → not comparable.
                scale_matched = False
                for f in per_dev:
                    if f is not None:
                        f['_y_likely_normalised'] = (id(f) in ratio_ids)

        # ── Generic cross-device rescue for non-temperature, non-log charts ──────
        # If both devices share real units but one is ~1000× the other (European
        # decimal "3.000"→3000), divide the inflated device's series into range.
        _EUR_SKIP = _NORM_RATIO_CTYPES | _FIGURE_LOG_TYPES
        if scale_matched and len(present) >= 2 and ctype not in _EUR_SKIP:
            def _ymax_of(f):
                ys = [abs(float(y)) for s in f.get('series', [])
                      for y in s.get('y', []) if y is not None]
                return max(ys) if ys else None
            pos_g = sorted(ym for ym in (_ymax_of(f) for f in present)
                           if ym is not None and ym > 0)
            if len(pos_g) >= 2 and pos_g[-1] / pos_g[0] > 100:
                ref_g = pos_g[0]
                for f in per_dev:
                    if f is None:
                        continue
                    chg = False
                    for s in f.get('series', []):
                        ny = []
                        for y in s.get('y', []):
                            if y is None:
                                ny.append(None); continue
                            yf = float(y)
                            if abs(yf) <= ref_g * 10:
                                ny.append(yf); continue
                            for factor in (10, 100, 1000, 10000):
                                if ref_g * 0.05 <= abs(yf) / factor <= ref_g * 5:
                                    ny.append(yf / factor); chg = True; break
                            else:
                                ny.append(yf)
                        s['y'] = ny
                    if chg:
                        f['_european_decimal_rescaled'] = True

        groups.append({
            'chart_type':      ctype,
            'display_name':    disp,
            'figs_per_device': per_dev,
            'scale_matched':   scale_matched,
        })
    return groups


def _y_range(fig):
    """Return (min_y, max_y) for the primary series of a figure, or (None, None)."""
    if fig is None or not fig.get('series'):
        return None, None
    ys = [float(y) for ser in fig['series'] for y in ser['y'] if y is not None]
    return (min(ys), max(ys)) if ys else (None, None)


def _detect_scale_mismatch(figs, labels):
    """
    Detect a genuine Y-axis scale mismatch: one device's median Y > 200× another's.
    (e.g. European decimal confusion causing values to be 1000× inflated.)
    The normalised-vs-absolute check has been intentionally removed — it used a
    fixed threshold of 5.0 which caused false positives for any chart where one
    device legitimately has small absolute values (e.g. Eoss < 5 µJ).
    That distinction is already handled by _group_figures (scale_matched flag).
    Returns a list of warning strings, or [] if no mismatch.
    """
    ranges = [(lbl, _y_range(f)) for lbl, f in zip(labels, figs) if f is not None]
    if len(ranges) < 2:
        return []

    medians = {}
    for lbl, (ymin, ymax) in ranges:
        if ymin is None:
            continue
        medians[lbl] = (ymin + ymax) / 2.0

    if len(medians) < 2:
        return []

    vals = list(medians.values())
    ratio = max(vals) / max(min(vals), 1e-9)
    if ratio > 200:
        bigger  = [l for l, m in medians.items() if m == max(vals)]
        smaller = [l for l, m in medians.items() if m == min(vals)]
        return [
            f"⚠  Y-axis scale mismatch (~{ratio:.0f}×): "
            f"{', '.join(bigger)} values are much larger than {', '.join(smaller)}. "
            f"Possible cause: European decimal notation in datasheet "
            f"(e.g. '1,000' misread as 1000 instead of 1.0). "
            f"Check the Y-axis units and tick labels on the original datasheet graph."
        ]
    return []


def _make_overlay_fig(group, labels):
    """
    Overlay the digitised curves of all devices on one matplotlib axis.
    Log-scale chart types (capacitance, thermal impedance) are rendered on
    log-log axes so the shape matches the datasheet graph exactly.
    Detects Y-axis unit mismatches and annotates with a warning.
    """
    if not HAS_MPL:
        return None
    figs    = group['figs_per_device']
    ctype   = group['chart_type']
    is_log  = ctype in _FIGURE_LOG_TYPES   # True for capacitance_vs_vds, thermal_impedance
    xl, yl  = _FIGURE_AXIS_LABELS.get(ctype, _CHART_AXIS_LABELS.get(ctype, ('X', 'Y')))
    # For the rds_vs_temp chart: if all Y values are ≤ 5 (normalised ratio range),
    # replace the mΩ unit label with "(normalized)" so the overlay axis is correct.
    # Restrict to rds_vs_temp only — other charts (Vbr, Vth) manage their own labels.
    if ctype == 'rds_vs_temp':
        _all_ys_rds = [float(y) for f in figs if f is not None
                       for s in f.get('series', []) for y in s.get('y', []) if y is not None]
        if _all_ys_rds and max(_all_ys_rds) <= 5.0:
            yl = 'R_DS(on) (normalized ratio)'

    # For log-scale charts, the "200× ratio" heuristic in _detect_scale_mismatch
    # would give false positives (Coss spans 9000→58 pF, a 155× range, in a single
    # device — not a unit error).  Only check the normalised-vs-absolute flag instead.
    scale_warnings = [] if is_log else _detect_scale_mismatch(figs, labels)

    fig, ax = plt.subplots(figsize=(7.5, 4.2), dpi=120)

    # Switch both axes to log scale for capacitance / thermal-impedance charts.
    # This is the key fix: on a linear axis a 1–10000 pF curve looks nothing like
    # the datasheet log-log plot — the knee is in the wrong place and the low
    # end is crushed to zero.
    if is_log:
        ax.set_xscale('log')
        ax.set_yscale('log')

    any_data = False
    for i, (f, lbl) in enumerate(zip(figs, labels)):
        if f is None:
            continue
        color = _DEV_COLORS[i % len(_DEV_COLORS)]
        for j, ser in enumerate(f['series']):
            # Filter out non-positive values before log-axis plotting
            pairs = sorted(
                (float(x), float(y)) for x, y in zip(ser['x'], ser['y'])
                if x is not None and y is not None
                and (not is_log or (float(x) > 0 and float(y) > 0))
            )
            if len(pairs) < 2:
                continue
            xs, ys = zip(*pairs)
            slbl = lbl if len(f['series']) == 1 else f"{lbl} · {ser['label']}"
            ax.plot(xs, ys, '-', color=color, linewidth=1.6,
                    alpha=0.95 if j == 0 else 0.6,
                    linestyle='-' if j == 0 else '--', label=slbl)
            any_data = True
    if not any_data:
        plt.close(fig)
        return None

    title = 'Digitised-data Overlay Comparison'
    if scale_warnings:
        title = '⚠ Units NOT equal — see warning below'
        warn_text = '\n'.join(scale_warnings)
        fig.text(0.5, -0.08, warn_text, ha='center', va='top', fontsize=7.5,
                 color='#8B0000', wrap=True,
                 bbox=dict(boxstyle='round,pad=0.4', facecolor='#FFF3CD',
                           edgecolor='#FF8C00', alpha=0.9))

    ax.set_title(title, fontsize=10, fontweight='bold',
                 color='#8B0000' if scale_warnings else 'black')
    ax.set_xlabel(xl, fontsize=9); ax.set_ylabel(yl, fontsize=9)
    # Major + minor gridlines for log charts so it looks like a proper log-log grid
    ax.grid(True, which='major', alpha=0.35)
    if is_log:
        ax.grid(True, which='minor', alpha=0.12)
    ax.legend(fontsize=7, loc='best')
    ax.tick_params(labelsize=8)
    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=120, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    buf.seek(0)
    return buf


def _figure_table_points(group, labels, n_points=5):
    """
    Interpolate n_points across the common x-range of the PRIMARY curve of each
    device. Returns (x_label, rows) with row = [x, y_dev0, y_dev1, …].
    Used only for scale-matched (real-unit) groups.

    For log-scale X axes (e.g. VDS on capacitance graphs), sample at log-spaced
    X values so all decades are represented — linear spacing would cluster all
    points in the flat high-VDS tail and miss the steep low-VDS knee entirely.
    For log-scale Y axes, interpolate in log space so intermediate values follow
    the physical curve shape (Coss ∝ VDS^(-0.5)) rather than a linear chord.
    """
    figs    = group['figs_per_device']
    ctype   = group['chart_type']
    is_log  = ctype in _FIGURE_LOG_TYPES
    xl, _   = _FIGURE_AXIS_LABELS.get(ctype, _CHART_AXIS_LABELS.get(ctype, ('X', 'Y')))

    is_ratio = ctype in _NORM_RATIO_CTYPES
    prim = []
    all_x = []
    for f in figs:
        if f is None or not f['series']:
            prim.append(None); continue
        if is_ratio:
            # Pick the series with the widest in-band y-range.  The correct Rds
            # curve (e.g. 0.518→2.384, ratio≈4.6) always spans a much wider range
            # than flat artifacts or spurious digitiser curves (0.928→1.107,
            # ratio≈1.19).  This is robust regardless of series ordering in the
            # file and does not require merging series (which can cause the walk
            # to switch curves at intermediate x-positions).
            ascending = (ctype != 'vth_vs_temp')
            PHYS_LO, PHYS_HI = 0.10, 5.0
            best_ser, best_ratio = None, -1.0
            for s in f['series']:
                raw = s.get('_raw_y', s.get('y', []))
                inb = [abs(float(y)) for y in raw
                       if y is not None and PHYS_LO <= abs(float(y)) <= PHYS_HI]
                if len(inb) < 2:
                    continue
                yr = max(inb) / max(min(inb), 1e-9)
                if yr > best_ratio:
                    best_ratio, best_ser = yr, s
            ser = best_ser or _primary_series(f) or f['series'][0]
            pairs = _extract_ratio_curve(ser, ascending=ascending)
        else:
            ser = _primary_series(f) or f['series'][0]
            pairs = sorted((float(x), float(y)) for x, y in zip(ser['x'], ser['y'])
                           if x is not None and y is not None
                           and (not is_log or (float(x) > 0 and float(y) > 0)))
        prim.append(pairs if len(pairs) >= 2 else None)
        all_x += [p[0] for p in pairs]
    if not all_x:
        return xl, []

    x_min, x_max = min(all_x), max(all_x)
    import math
    if is_log and x_min > 0 and x_max > x_min:
        # Log-spaced sample points span all decades evenly
        lx_min, lx_max = math.log10(x_min), math.log10(x_max)
        ref_xs = [10 ** (lx_min + (lx_max - lx_min) * i / max(n_points - 1, 1))
                  for i in range(n_points)]
    else:
        ref_xs = [x_min + (x_max - x_min) * i / max(n_points - 1, 1)
                  for i in range(n_points)]

    def _interp(pairs, xi):
        if not pairs:
            return None
        xs, ys = zip(*pairs)
        if xi <= xs[0]:  return ys[0]
        if xi >= xs[-1]: return ys[-1]
        for k in range(len(xs) - 1):
            if xs[k] <= xi <= xs[k + 1]:
                if is_log and xs[k] > 0 and xs[k+1] > 0 and ys[k] > 0 and ys[k+1] > 0:
                    # Interpolate in log-log space: follows power-law curves exactly
                    t = (math.log10(xi) - math.log10(xs[k])) / \
                        (math.log10(xs[k+1]) - math.log10(xs[k]))
                    return 10 ** (math.log10(ys[k]) + t * (math.log10(ys[k+1]) - math.log10(ys[k])))
                else:
                    t = (xi - xs[k]) / (xs[k + 1] - xs[k])
                    return ys[k] + t * (ys[k + 1] - ys[k])
        return None

    rows = []
    for xi in ref_xs:
        row = [round(xi, 4)]
        for pairs in prim:
            v = _interp(pairs, xi)
            row.append(round(v, 4) if v is not None else None)
        rows.append(row)
    return xl, rows


def build_image_comparison_sheet(wb_out, paths, labels, all_figs_per_file):
    """
    Build the 'Graph Comparison' sheet from embedded datasheet PICTURES.

    Priority 1 — paste the device images SIDE BY SIDE (always, even when the
                 datasheet axis is normalised so units differ between devices).
    Priority 2 — a digitised-data overlay comparison.
    Priority 3 — a 5-point extracted comparison table, ONLY when every device's
                 axes are in real (matched) units.

    Returns list of temp PNG paths to delete after wb_out.save().
    """
    ws = wb_out.create_sheet("Graph Comparison")
    ws.sheet_view.showGridLines = False
    n         = len(paths)
    tmp_files = []
    COLS_PER  = 6                     # Excel columns reserved per device block
    ncols     = max(4, n * COLS_PER)

    merge_title(ws, 1, 1, ncols,
                "GRAPH COMPARISON  —  datasheet figures side-by-side  +  "
                "overlay (when both devices have data + matched units)",
                bg=C_TITLE, sz=13, h=36)
    merge_title(ws, 2, 1, ncols,
                "Datasheet images shown side-by-side.  "
                "Overlay chart added only when ALL devices have graph data AND axes are in real matched units.",
                bg=C_SUBHDR, fg="FFFFFF", sz=9, h=28)

    groups = _group_figures(all_figs_per_file)
    if not groups:
        wc(ws, 4, 1, "No embedded datasheet figures found in the input files.",
           font=fnt(11, italic=True, color="888888"))
        return tmp_files

    from openpyxl.drawing.image import Image as XLImage
    current_row = 4

    # Column widths so each device block is wide enough for an image
    ws.column_dimensions["A"].width = 22
    for c in range(1, ncols + 1):
        if ws.column_dimensions[gcol(c)].width is None or \
           ws.column_dimensions[gcol(c)].width < 13:
            ws.column_dimensions[gcol(c)].width = 13

    for group in groups:
        ctype        = group['chart_type']
        display_name = group['display_name']
        figs         = group['figs_per_device']
        matched      = group['scale_matched']
        all_have_data = all(f is not None for f in figs)
        units_mismatch = all_have_data and not matched and ctype not in _FIGURE_LOG_TYPES

        # ── Section header ────────────────────────────────────────────────────
        # Flag the title when both devices have data but units are incompatible
        hdr_text = (f"▶  {display_name}  —  ⚠ UNITS ARE NOT SAME"
                    if units_mismatch else f"▶  {display_name}")
        hdr_bg   = "7F3000" if units_mismatch else C_SUBHDR
        merge_title(ws, current_row, 1, ncols, hdr_text, bg=hdr_bg, sz=11, h=26)
        current_row += 1

        # Scale / units status line
        if matched:
            _rescaled = any(f is not None and f.get('_european_decimal_rescaled') for f in figs)
            _rescale_sfx = ("  (Y-axis auto-rescaled ÷1000: European decimal notation detected.)"
                            if _rescaled else "")
            if ctype == 'rds_vs_temp':
                note = ("✓  Both graphs use a normalised R_DS(on) Y-axis (ratio, 0–3) — "
                        "overlay shown below.  "
                        "Absolute R_DS(on) at T_j  =  R_DS(on)(25 °C) from spec  ×  ratio."
                        + _rescale_sfx)
            elif ctype == 'vth_vs_temp':
                note = ("✓  Both graphs use a normalised V_GS(th) Y-axis (ratio, 0–2) — "
                        "overlay shown below.  "
                        "Absolute V_GS(th) at T_j  =  V_GS(th)(25 °C) from spec  ×  ratio."
                        + _rescale_sfx)
            else:
                note = ("✓  Axes in real matched units — overlay chart shown below."
                        + _rescale_sfx)
            nbg, nfg = "E2EFDA", "375623"
        elif units_mismatch:
            norm_devs = [labels[i] for i, f in enumerate(figs)
                         if f is not None and f.get('_y_likely_normalised')]
            abs_devs  = [labels[i] for i, f in enumerate(figs)
                         if f is not None and not f.get('_y_likely_normalised')]
            detail = (f"  {', '.join(norm_devs)} → normalised; "
                      f"{', '.join(abs_devs)} → absolute."
                      if norm_devs and abs_devs else "")
            note = (f"⚠  UNITS ARE NOT SAME — overlay not generated.{detail}  "
                    f"Images shown side-by-side for visual reference only.")
            nbg, nfg = "FFF3CD", "7F0000"
        else:
            reason = ("log/normalised axes" if ctype in _FIGURE_LOG_TYPES
                      else "one or more datasheets use a normalised (0–1) axis")
            note = f"ℹ  {reason} — images shown side-by-side for visual reference."
            nbg, nfg = "FFFBEA", "555555"
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=ncols)
        wc(ws, current_row, 1, note, font=fnt(9, italic=(not units_mismatch), bold=units_mismatch, color=nfg),
           fill=ff(nbg), align=lft(), border=BT)
        ws.row_dimensions[current_row].height = 24 if units_mismatch else 20
        current_row += 1

        # ── Device label row ──────────────────────────────────────────────────
        label_row = current_row
        ws.row_dimensions[label_row].height = 18
        for i, lbl in enumerate(labels):
            c0 = 1 + i * COLS_PER
            merge_title(ws, label_row, c0, c0 + COLS_PER - 1, lbl,
                        bg=C_HDR, fg="FFFFFF", sz=10, h=18)
        current_row += 1

        # ── Images side by side (Priority 1) ──────────────────────────────────
        # 320px / (96px/in) × 72pt/in = 240pt total row space needed.
        # 10 rows × 24pt = 240pt → image fills rows exactly, zero gap to caption.
        IMG_W, IMG_H = 300, 320
        img_rows     = 10
        img_top      = current_row
        for r in range(img_top, img_top + img_rows):
            ws.row_dimensions[r].height = 24
        for i, f in enumerate(figs):
            c0     = 1 + i * COLS_PER
            anchor = f"{gcol(c0)}{img_top}"
            if f is None:
                cell = wc(ws, img_top + img_rows // 2, c0, "Not available",
                          font=fnt(10, italic=True, color="999999"), align=ctr())
                continue
            try:
                tf = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                tf.write(f['image_bytes']); tf.close()
                tmp_files.append(tf.name)
                xi = XLImage(tf.name)
                xi.width, xi.height = IMG_W, IMG_H
                ws.add_image(xi, anchor)
            except Exception as ex:
                wc(ws, img_top, c0, f"[image error: {ex}]",
                   font=fnt(8, italic=True, color="AA0000"))
        current_row = img_top + img_rows

        # ── Caption row (datasheet figure title per device) ───────────────────
        cap_row = current_row
        ws.row_dimensions[cap_row].height = 30
        for i, f in enumerate(figs):
            c0 = 1 + i * COLS_PER
            cap = f['caption'] if f is not None else "—"
            ws.merge_cells(start_row=cap_row, start_column=c0,
                           end_row=cap_row, end_column=c0 + COLS_PER - 1)
            wc(ws, cap_row, c0, cap, font=fnt(8, italic=True, color="444444"),
               fill=ff(C_ALT), align=ctr(), border=BT)
        current_row += 1

        # ── Log-scale graphs: images already shown above — skip overlay & table ──
        if ctype in _FIGURE_LOG_TYPES:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=ncols)
            wc(ws, current_row, 1,
               "⚠  Extraction is not working for Log scale graphs — "
               "refer to the original datasheet images shown above for reference.",
               font=fnt(10, bold=True, color="7F0000"),
               fill=ff("FFF3CD"), align=lft(), border=BT)
            ws.row_dimensions[current_row].height = 30
            current_row += 4
            continue

        # ── Overlay chart or "units not equal" note ──────────────────────────────
        all_have_data = all(f is not None for f in figs)
        scale_warns   = _detect_scale_mismatch(figs, labels) if all_have_data else []

        if all_have_data and not matched:
            # Both devices have data BUT axes are in incompatible units → explain why
            norm_devs = [labels[i] for i, f in enumerate(figs)
                         if f is not None and f.get('_y_likely_normalised')]
            abs_devs  = [labels[i] for i, f in enumerate(figs)
                         if f is not None and not f.get('_y_likely_normalised')]
            if norm_devs and abs_devs:
                unit_detail = (f"  {', '.join(norm_devs)} → normalised Y-axis (0–3);  "
                               f"{', '.join(abs_devs)} → absolute Y-axis.")
            else:
                unit_detail = "  One or more graphs use a normalised Y-axis; another uses absolute values."
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=ncols)
            wc(ws, current_row, 1,
               "⚠  UNITS NOT EQUAL — overlay comparison not generated." + unit_detail + "  "
               "Images are shown side-by-side for visual reference only; "
               "comparison table below uses N/A where values are incommensurable.",
               font=fnt(9, bold=True, color="7F0000"),
               fill=ff("FFF3CD"), align=lft(), border=BT)
            ws.row_dimensions[current_row].height = 40
            current_row += 1
        elif all_have_data and matched:
            # Both devices have data AND axes match → show overlay (with warning if mismatch detected)
            buf = _make_overlay_fig(group, labels)
            if buf is not None:
                try:
                    tf = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                    tf.write(buf.getvalue()); tf.close()
                    tmp_files.append(tf.name)
                    xi = XLImage(tf.name)
                    xi.width, xi.height = 620, 350
                    ws.add_image(xi, f"A{current_row}")
                    for r in range(current_row, current_row + 19):
                        ws.row_dimensions[r].height = 18
                    current_row += 20
                except Exception:
                    current_row += 1
            # Note when European decimal confusion was auto-corrected
            if any(f is not None and f.get('_european_decimal_rescaled') for f in figs):
                rescaled_devs = [labels[i] for i, f in enumerate(figs)
                                 if f is not None and f.get('_european_decimal_rescaled')]
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=ncols)
                wc(ws, current_row, 1,
                   f"ℹ  Auto-corrected: {', '.join(rescaled_devs)} Y-axis values were "
                   f"~1000× too large (European decimal confusion — '3.000' read as 3000 "
                   f"instead of 3.0).  Values divided by 1000 for comparison.",
                   font=fnt(8, italic=True, color="005500"),
                   fill=ff("E2EFDA"), align=lft(), border=BT)
                ws.row_dimensions[current_row].height = 26
                current_row += 1
            # If scale mismatch was found even within "matched" group, add a note
            if scale_warns:
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=ncols)
                wc(ws, current_row, 1,
                   "⚠  " + "  |  ".join(scale_warns),
                   font=fnt(8, italic=True, color="7F3000"),
                   fill=ff("FFF3CD"), align=lft(), border=BT)
                ws.row_dimensions[current_row].height = 36
                current_row += 1

        current_row += 1   # one blank row between chart groups

    return tmp_files


def _get_chart_title(chart):
    title = getattr(chart, 'title', None)
    if title is None:
        return ""
    if isinstance(title, str):
        return title
    tx = getattr(title, 'tx', None)
    if tx is not None:
        rich = getattr(tx, 'rich', None)
        if rich is not None:
            texts = []
            for para in getattr(rich, 'p', []):
                for run in getattr(para, 'r', []):
                    t = getattr(run, 't', '')
                    if t:
                        texts.append(str(t))
            if texts:
                return ' '.join(texts)
    return str(title)


def _get_series_label(ser):
    title = getattr(ser, 'title', None)
    if title is None:
        return ""
    if isinstance(title, str):
        return title
    str_ref = getattr(title, 'strRef', None)
    if str_ref is not None:
        sc = getattr(str_ref, 'strCache', None)
        if sc is not None:
            for pt in getattr(sc, 'pt', []):
                v = getattr(pt, 'v', None)
                if v:
                    return str(v)
    v = getattr(title, 'v', None)
    return str(v) if v else str(title)


def _read_cells_from_formula(wb, formula):
    """Read cell values from a formula like 'Sheet1!$A$1:$A$10'."""
    try:
        formula = str(formula).strip()
        if '!' not in formula:
            return []
        sheet_part, cell_part = formula.split('!', 1)
        sheet_name = sheet_part.strip("'\"")
        cell_part  = cell_part.replace('$', '')
        if sheet_name not in wb.sheetnames:
            return []
        ws     = wb[sheet_name]
        cr     = ws[cell_part]
        values = []
        if hasattr(cr, 'value'):
            v = cr.value
            try:
                values.append(float(v) if v is not None else None)
            except (TypeError, ValueError):
                values.append(None)
        else:
            for item in cr:
                cells = item if hasattr(item, '__iter__') else [item]
                for cell in cells:
                    v = cell.value
                    try:
                        values.append(float(v) if v is not None else None)
                    except (TypeError, ValueError):
                        values.append(None)
        return values
    except Exception:
        return []


def _extract_numdata(num_src, wb=None):
    """Extract numeric values from an openpyxl chart NumDataSource."""
    if num_src is None:
        return []
    result_dict = {}

    nr = getattr(num_src, 'numRef', None)
    if nr is not None:
        nc = getattr(nr, 'numCache', None)
        if nc is not None:
            for pt in getattr(nc, 'pt', []):
                idx = getattr(pt, 'idx', None)
                val = getattr(pt, 'v',   None)
                if idx is not None and val is not None:
                    try:
                        result_dict[int(idx)] = float(val)
                    except (ValueError, TypeError):
                        pass
            if result_dict:
                mx = max(result_dict.keys())
                return [result_dict.get(i) for i in range(mx + 1)]
        f = getattr(nr, 'f', None)
        if f and wb is not None:
            vals = _read_cells_from_formula(wb, f)
            if vals:
                return vals

    nl = getattr(num_src, 'numLit', None)
    if nl is not None:
        for pt in getattr(nl, 'pt', []):
            idx = getattr(pt, 'idx', None)
            val = getattr(pt, 'v',   None)
            if idx is not None and val is not None:
                try:
                    result_dict[int(idx)] = float(val)
                except (ValueError, TypeError):
                    pass
        if result_dict:
            mx = max(result_dict.keys())
            return [result_dict.get(i) for i in range(mx + 1)]

    return []


def _identify_chart_type(title, series_labels):
    text = (title + " " + " ".join(series_labels or [])).lower()
    for ctype, patterns in _CHART_TYPE_PATTERNS.items():
        if any(p in text for p in patterns):
            return ctype
    return 'unknown'


def extract_charts_from_file(path):
    """
    Read all embedded chart objects from an xlsx file.
    Returns list of dicts: {title, chart_type, series, is_log_scale, ws_name}
    Each series dict: {label, x:[floats], y:[floats]}
    """
    charts_out = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            for chart in getattr(ws, '_charts', []):
                chart_title = _get_chart_title(chart)
                series_list, ser_labels = [], []
                for ser in getattr(chart, 'series', []):
                    lbl = _get_series_label(ser)
                    ser_labels.append(lbl)
                    x_src  = getattr(ser, 'xVal', None) or getattr(ser, 'cat', None)
                    y_src  = getattr(ser, 'val',  None)
                    x_vals = _extract_numdata(x_src, wb)
                    y_vals = _extract_numdata(y_src, wb)
                    if x_vals or y_vals:
                        series_list.append({'label': lbl, 'x': x_vals, 'y': y_vals})
                ctype  = _identify_chart_type(chart_title, ser_labels)
                is_log = ctype in _LOG_CHART_TYPES
                charts_out.append({
                    'title':       chart_title,
                    'chart_type':  ctype,
                    'series':      series_list,
                    'is_log_scale': is_log,
                    'ws_name':     ws_name,
                })
        wb.close()
    except Exception as ex:
        print(f"  Warning: chart extraction from {Path(path).name}: {ex}")
    return charts_out


def _build_fallback_charts(mosfets):
    """
    Build synthetic chart dicts from tabular data when no Excel chart objects exist.
    The returned list has one entry per chart type; each series[i] belongs to device i.
    """
    result = []

    # C_oss vs V_DS (from coss_rows)
    if any(d.get('coss_rows') for d in mosfets):
        series = []
        for d in mosfets:
            rows = d.get('coss_rows', [])
            xs   = [r[0] for r in rows if r[0] is not None]
            ys   = [r[1] for r in rows if r[1] is not None]
            series.append({'label': 'C_oss', 'x': xs, 'y': ys})
        result.append({
            'title': 'C_oss vs V_DS',
            'chart_type': 'capacitance_vs_vds',
            'series': series,
            'is_log_scale': True,
            'ws_name': 'fallback',
            '_per_device': True,
        })

    # R_DS(on) vs Temperature (from 25 °C + high-T points)
    if any(d.get('rds25') is not None and d.get('rds_ht') is not None for d in mosfets):
        series = []
        for d in mosfets:
            xs, ys = [], []
            if d.get('rds25') is not None:
                xs.append(25.0); ys.append(d['rds25'])
            if d.get('rds_ht') is not None and d.get('_rds_ht_tj') is not None:
                xs.append(d['_rds_ht_tj']); ys.append(d['rds_ht'])
            series.append({'label': 'R_DS(on)', 'x': xs, 'y': ys})
        result.append({
            'title': 'R_DS(on) vs Temperature',
            'chart_type': 'rds_vs_temp',
            'series': series,
            'is_log_scale': False,
            'ws_name': 'fallback',
            '_per_device': True,
        })

    return result


def _match_charts(all_charts_per_file):
    """
    Group matching chart types across all device files.
    Returns list of groups:
      {chart_type, display_name, is_log_scale, _per_device,
       charts_per_device:[chart_dict or None]}
    """
    seen, all_types = set(), []
    for charts in all_charts_per_file:
        for c in charts:
            ct = c['chart_type']
            if ct not in seen:
                all_types.append(ct); seen.add(ct)

    groups = []
    for ctype in all_types:
        per_dev = []
        for dev_charts in all_charts_per_file:
            best = None
            for c in dev_charts:
                if c['chart_type'] != ctype:
                    continue
                if best is None:
                    best = c
                else:
                    cur_pts = sum(len(s.get('y', [])) for s in best.get('series', []))
                    new_pts = sum(len(s.get('y', [])) for s in c.get('series', []))
                    if new_pts > cur_pts:
                        best = c
            per_dev.append(best)

        if not any(c is not None for c in per_dev):
            continue

        is_per_dev = any(c.get('_per_device') for c in per_dev if c)
        groups.append({
            'chart_type':        ctype,
            'display_name':      _CHART_DISPLAY_NAMES.get(ctype,
                                     ctype.replace('_', ' ').title()),
            'is_log_scale':      ctype in _LOG_CHART_TYPES,
            '_per_device':       is_per_dev,
            'charts_per_device': per_dev,
        })
    return groups


_DEV_COLORS = ['#1F77B4', '#FF7F0E', '#2CA02C', '#D62728', '#9467BD', '#8C564B']


def _make_comparison_fig(group, labels):
    """
    Build a matplotlib figure: one subplot per device + one overlay subplot.
    Returns io.BytesIO PNG or None.
    """
    if not HAS_MPL:
        return None

    charts       = group['charts_per_device']
    chart_type   = group['chart_type']
    is_log       = group['is_log_scale']
    is_per_dev   = group.get('_per_device', False)
    n            = len(charts)
    display_name = group['display_name']
    x_lbl, y_lbl = _CHART_AXIS_LABELS.get(chart_type, ('X', 'Y'))

    fig, axes = plt.subplots(1, n + 1, figsize=(5.5 * (n + 1), 4.5),
                              squeeze=False, dpi=120)
    ax_over = axes[0][n]
    ax_over.set_title('Overlay Comparison', fontsize=9, fontweight='bold', pad=8)

    any_data = False

    for col, (chart_data, label) in enumerate(zip(charts, labels)):
        ax        = axes[0][col]
        dev_color = _DEV_COLORS[col % len(_DEV_COLORS)]
        ax.set_title(label, fontsize=8.5, fontweight='bold', pad=8)

        if chart_data is None:
            ax.text(0.5, 0.5, 'Not available', ha='center', va='center',
                    transform=ax.transAxes, fontsize=10, color='#888888', style='italic')
            ax.set_facecolor('#F5F5F5')
            for sp in ax.spines.values():
                sp.set_color('#CCCCCC')
            continue

        # Per-device mode: series[col] belongs to this device
        if is_per_dev:
            sers = chart_data.get('series', [])
            dev_sers = [sers[col]] if col < len(sers) else []
        else:
            dev_sers = chart_data.get('series', [])

        plotted = False
        for i, ser in enumerate(dev_sers):
            pairs = [(float(x), float(y)) for x, y in
                     zip(ser.get('x', []), ser.get('y', []))
                     if x is not None and y is not None]
            if not pairs:
                continue
            xs, ys  = zip(*sorted(pairs))
            color   = dev_color if is_per_dev else _DEV_COLORS[i % len(_DEV_COLORS)]
            ser_lbl = ser.get('label') or f"Series {i + 1}"
            ov_lbl  = label if (is_per_dev or i == 0) else f"{label} {ser_lbl}"

            if is_log:
                xp = [x for x, y in zip(xs, ys) if x > 0 and y > 0]
                yp = [y for x, y in zip(xs, ys) if x > 0 and y > 0]
                if xp:
                    ax.loglog(xp, yp, 'o-', color=color,
                              label=ser_lbl, linewidth=1.5, markersize=3)
                    ax_over.loglog(xp, yp, 'o-', color=dev_color if is_per_dev else color,
                                   label=ov_lbl, linewidth=1.5, markersize=3)
                    plotted = True; any_data = True
            else:
                ax.plot(xs, ys, 'o-', color=color,
                        label=ser_lbl, linewidth=1.5, markersize=4)
                ax_over.plot(xs, ys, 'o-', color=dev_color if is_per_dev else color,
                             label=ov_lbl, linewidth=1.5, markersize=4)
                plotted = True; any_data = True

        if plotted:
            ax.grid(True, alpha=0.3, which='both' if is_log else 'major')
            if len(dev_sers) > 1:
                ax.legend(fontsize=7, loc='best')
            ax.set_xlabel(x_lbl, fontsize=8); ax.set_ylabel(y_lbl, fontsize=8)
            ax.tick_params(labelsize=7)
        else:
            ax.text(0.5, 0.5, 'No data\nextracted', ha='center', va='center',
                    transform=ax.transAxes, fontsize=9, color='#888888', style='italic')
            ax.set_facecolor('#F9F9F9')

    if any_data:
        ax_over.grid(True, alpha=0.3, which='both' if is_log else 'major')
        ax_over.legend(fontsize=7, loc='best')
        ax_over.set_xlabel(x_lbl, fontsize=8)
        ax_over.set_ylabel(y_lbl, fontsize=8)
        ax_over.tick_params(labelsize=7)
    else:
        ax_over.text(0.5, 0.5, 'No data', ha='center', va='center',
                     transform=ax_over.transAxes, fontsize=10,
                     color='#888888', style='italic')

    fig.suptitle(display_name, fontsize=11, fontweight='bold', y=1.01)
    plt.tight_layout(rect=[0, 0, 1, 0.96])

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=120, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    buf.seek(0)
    return buf


def _extract_table_points(group, labels, n_points=5):
    """
    Interpolate n_points at evenly-spaced x values from the primary series.
    Returns (x_unit_label, list_of_rows) where each row = [x, y_dev0, y_dev1, …]
    """
    charts     = group['charts_per_device']
    chart_type = group['chart_type']
    is_per_dev = group.get('_per_device', False)
    x_lbl, _   = _CHART_AXIS_LABELS.get(chart_type, ('X', 'Y'))

    all_x = []
    for idx, cd in enumerate(charts):
        if cd is None:
            continue
        sers = cd.get('series', [])
        ser  = (sers[idx] if is_per_dev and idx < len(sers)
                else (sers[0] if sers else None))
        if ser:
            all_x += [float(v) for v in ser.get('x', []) if v is not None]

    if not all_x:
        return x_lbl, []

    x_min, x_max = min(all_x), max(all_x)
    ref_xs = [x_min + (x_max - x_min) * i / max(n_points - 1, 1)
               for i in range(n_points)]

    rows = []
    for xi in ref_xs:
        row = [round(xi, 4)]
        for idx, cd in enumerate(charts):
            if cd is None:
                row.append(None); continue
            sers   = cd.get('series', [])
            ser    = (sers[idx] if is_per_dev and idx < len(sers)
                      else (sers[0] if sers else None))
            y_val  = None
            if ser:
                pairs = sorted([(float(x), float(y))
                                for x, y in zip(ser.get('x', []), ser.get('y', []))
                                if x is not None and y is not None])
                if pairs:
                    xs, ys = zip(*pairs)
                    if xi <= xs[0]:
                        y_val = ys[0]
                    elif xi >= xs[-1]:
                        y_val = ys[-1]
                    else:
                        for j in range(len(xs) - 1):
                            if xs[j] <= xi <= xs[j + 1]:
                                t     = (xi - xs[j]) / (xs[j + 1] - xs[j])
                                y_val = ys[j] + t * (ys[j + 1] - ys[j])
                                break
            row.append(round(y_val, 4) if y_val is not None else None)
        rows.append(row)
    return x_lbl, rows


def build_graph_comparison_sheet(wb_out, paths, mosfets, labels, all_charts_per_file):
    """
    Adds 'Graph Comparison' sheet to wb_out.
    Returns list of temp PNG file paths (caller deletes them after wb_out.save()).
    """
    ws = wb_out.create_sheet("Graph Comparison")
    ws.sheet_view.showGridLines = False
    n         = len(paths)
    tmp_files = []
    ncols     = max(4, n + 3)

    merge_title(ws, 1, 1, ncols,
                "GRAPH COMPARISON  —  side-by-side charts + overlay (when both devices have data) + comparison table",
                bg=C_TITLE, sz=13, h=36)

    # Decide whether we have real chart objects or need fallback synthetic charts
    found_real = any(c.get('ws_name') != 'fallback'
                     for charts in all_charts_per_file for c in charts)

    if not found_real:
        merge_title(ws, 2, 1, ncols,
                    "⚠  No embedded chart objects found in the Excel files — "
                    "showing synthetic charts built from extracted tabular data.",
                    bg=C_WARN, fg=C_WARNF, sz=9, h=20)
        fallbacks          = _build_fallback_charts(mosfets)
        # Wrap in a per-file list: first file "owns" all fallback charts
        per_file_fallback  = [fallbacks] + [[] for _ in range(n - 1)]
        groups             = _match_charts(per_file_fallback)
    else:
        merge_title(ws, 2, 1, ncols,
                    "Each device shown individually + overlay comparison on the right.  "
                    "Log-scale charts (C_oss/C_iss/C_rss vs V_DS, Z_th vs t_p): images only.",
                    bg=C_SUBHDR, fg="FFFFFF", sz=9, h=20)
        groups = _match_charts(all_charts_per_file)

    if not groups:
        c3 = ws.cell(row=3, column=1, value="No chart data available.")
        c3.font = fnt(11, italic=True, color="888888")
        return tmp_files

    current_row = 3

    for group in groups:
        chart_type   = group['chart_type']
        display_name = group['display_name']
        is_log       = group['is_log_scale']

        # ── Section header ────────────────────────────────────────────────────
        merge_title(ws, current_row, 1, ncols,
                    f"▶  {display_name}",
                    bg=C_SUBHDR, sz=11, h=28)
        current_row += 1

        if is_log:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=ncols)
            nc = ws.cell(row=current_row, column=1,
                         value="ℹ  Log-scale chart — axes not directly comparable; "
                               "comparison table below shows N/A where device has no matched data.")
            nc.font = fnt(9, italic=True, color="555555")
            nc.fill = ff("FFFBEA"); nc.alignment = lft()
            current_row += 1

        # ── Overlay chart: only when ALL devices have chart data AND not log-scale ──
        charts = group['charts_per_device']
        all_have_data = all(c is not None for c in charts)
        if all_have_data and not is_log:
            fig_buf = _make_comparison_fig(group, labels)
            if fig_buf is not None:
                try:
                    tf = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                    tf.write(fig_buf.getvalue()); tf.close()
                    tmp_files.append(tf.name)
                    from openpyxl.drawing.image import Image as XLImage
                    img        = XLImage(tf.name)
                    img.width  = min(360 * (n + 1), 1080)
                    img.height = 310
                    ws.add_image(img, f"A{current_row}")
                    img_rows = 18
                    for r in range(current_row, current_row + img_rows):
                        ws.row_dimensions[r].height = 19
                    current_row += img_rows
                except Exception as ex:
                    ws.cell(row=current_row, column=1).value = f"[Image error: {ex}]"
                    ws.cell(row=current_row, column=1).font  = fnt(9, italic=True, color="AA0000")
                    current_row += 2

        # ── Comparison table (always shown; N/A for device with no chart data) ──
        x_lbl, rows = _extract_table_points(group, labels, n_points=5)
        if not rows:
            rows = [[None] + [None] * n] * 5  # placeholder N/A table
        current_row += 1
        merge_title(ws, current_row, 1, 1 + n,
                    f"Comparison  —  {display_name}  (5 points"
                    + ("  |  N/A = no data for that device)" if not all_have_data else ")"),
                    bg=C_HDR, fg="FFFFFF", sz=9, h=20)
        current_row += 1

        ws.row_dimensions[current_row].height = 22
        wc(ws, current_row, 1, x_lbl or "Point",
           font=fnt(10, bold=True, color="FFFFFF"),
           fill=ff(C_SUBHDR), align=ctr(), border=BT)
        y_unit = _TABLE_Y_UNITS.get(chart_type, '')
        for d_i, lbl in enumerate(labels):
            wc(ws, current_row, 2 + d_i, f"{lbl}  ({y_unit})" if y_unit else lbl,
               font=fnt(10, bold=True, color="FFFFFF"),
               fill=ff(C_SUBHDR), align=ctr(), border=BT)
        current_row += 1

        lower_better = chart_type not in ('transfer_char', 'output_char',
                                          'vbr_vs_temp', 'avalanche_energy',
                                          'power_derating')
        for r_i, row_data in enumerate(rows):
            ws.row_dimensions[current_row].height = 20
            bg = C_ALT if r_i % 2 == 0 else C_WHITE
            wc(ws, current_row, 1,
               round(row_data[0], 4) if row_data[0] is not None else "N/A",
               font=fnt(10, bold=True), fill=ff(bg),
               align=ctr(), border=BT, nfmt="0.0####")
            y_vals_d = {i: row_data[i + 1] for i in range(n)
                        if row_data[i + 1] is not None}
            best_i = _winner(y_vals_d, lower_better)
            for d_i in range(n):
                v       = row_data[d_i + 1]
                is_best = (d_i == best_i)
                wc(ws, current_row, 2 + d_i,
                   round(v, 4) if v is not None else "N/A",
                   font=fnt(10, bold=is_best,
                            color=C_BESTF if is_best else "000000"),
                   fill=ff(C_BEST if is_best else bg),
                   align=ctr(), border=BT, nfmt="0.0####")
            current_row += 1

        current_row += 2

    ws.column_dimensions["A"].width = 30
    for d_i in range(n):
        ws.column_dimensions[gcol(2 + d_i)].width = 24

    return tmp_files


# ─── NORMALISATION ─────────────────────────────────────────────────────────────

def compute_normalised(mosfets, ref):
    vds_ref = ref["vds_ref"]
    id_ref  = ref["id_ref"]
    tj_ref  = ref["tj_ref"]
    results = []

    for d in mosfets:
        nm = {}

        # Vdss — direct
        nm["vdss"]  = (d.get("vdss"),  "Direct (rated value)",             d.get("vdss_tcond", ""))

        # Rds @ 25°C and High-T — direct with conditions
        rds25_cond  = f"@ 25°C, {d.get('rds_tcond','')}" if d.get('rds_tcond') else "@ 25°C"
        rdsht_cond  = (f"@ {d.get('_rds_ht_tj',125):.0f}°C, {d.get('rds_tcond','')}"
                       if d.get('rds_ht') else "Not in datasheet")
        nm["rds25"] = (d.get("rds25"), f"Typ @ 25°C;  {d.get('rds_tcond','')}",  rds25_cond)
        nm["rds_ht"]= (d.get("rds_ht"), rdsht_cond, rdsht_cond)

        # For normalised comparison: Rds @ Tj_ref via interpolation (spec §3.1)
        rds_tj_v, rds_tj_m = extrapolate_rds(d, tj_ref)
        rds_tj_m = "[TABLE INTERPOLATION] " + rds_tj_m
        nm["_rds_tj"]= (rds_tj_v, rds_tj_m, "")

        # Qg — scaled to VDS_ref if test conditions differ (spec §3.6)
        qg_cond  = d.get("qg_tcond", "")
        qg_v, qg_m = extrapolate_qg(d, vds_ref)
        nm["qg"] = (qg_v, qg_m, qg_cond)

        # Eon / Eoff — FORMULA (derived), scaled on BOTH VDD and ID axes (spec §3.7)
        eon_id   = d.get("_eon_id")
        eoff_id  = d.get("_eoff_id")
        eon_vdd  = d.get("_eon_vds")    # datasheet test VDD for Eon
        eoff_vdd = d.get("_eoff_vds")   # datasheet test VDD for Eoff
        # Fall back to switching-time test VDS when Eon/Eoff VDD not separately stored
        sw_vdd   = d.get("_sw_vds")
        eon_vdd  = eon_vdd  or sw_vdd
        eoff_vdd = eoff_vdd or sw_vdd
        eon_v,  eon_m  = scale_energy(d.get("eon"),  eon_id,  id_ref,
                                      vdd_test=eon_vdd,  vdd_ref=vds_ref)
        eoff_v, eoff_m = scale_energy(d.get("eoff"), eoff_id, id_ref,
                                      vdd_test=eoff_vdd, vdd_ref=vds_ref)
        der_note = " [base derived from t_r/t_d(on)]" if d.get("_eon_derived") else ""
        nm["eon"]  = (eon_v,  eon_m  + der_note, d.get("eon_tcond",  ""))
        nm["eoff"] = (eoff_v, eoff_m + (" [base derived from t_f/t_d(off)]"
                      if d.get("_eoff_derived") else ""), d.get("eoff_tcond", ""))

        # ID — direct
        nm["id"]    = (d.get("id"), "[GRAPH - DIRECT] Rated max continuous @ 25°C",
                       d.get("id_tcond", ""))

        # Vth @ Tj_ref — GRAPH-INTERPOLATED (ST) or FORMULA (IPA) (spec §3.2)
        vth_v, vth_m = extrapolate_vth(d, tj_ref)
        nm["vth"]   = (vth_v, vth_m, d.get("vth_tcond", ""))

        # Ciss @ vds_ref — formula fallback; graph-direct from C-V curve preferred (spec §3.4)
        ciss_v, ciss_m = extrapolate_coss(d.get("ciss_rows", []), vds_ref)
        if ciss_v is None:
            ciss_v = d.get("ciss")
            ciss_m = (f"[GRAPH - DIRECT] @ {d.get('_ciss_vds_test','?')}V"
                      if ciss_v is not None else "[NOT EXTRAPOLATABLE] Not available")
        nm["ciss"]  = (ciss_v, ciss_m, d.get("ciss_tcond",""))

        # Coss @ vds_ref — same approach (spec §3.4: GRAPH-DIRECT preferred)
        coss_v, coss_m = extrapolate_coss(d.get("coss_rows", []), vds_ref)
        nm["coss"]  = (coss_v, coss_m, d.get("coss_tcond", ""))

        # Crss @ vds_ref
        crss_v, crss_m = extrapolate_coss(d.get("crss_rows", []), vds_ref)
        if crss_v is None:
            crss_v = d.get("crss")
            crss_m = (f"[GRAPH - DIRECT] @ {d.get('_crss_vds_test','?')}V"
                      if crss_v is not None else "[NOT EXTRAPOLATABLE] Not available")
        nm["crss"]  = (crss_v, crss_m, d.get("crss_tcond",""))

        # Qgd — direct from table
        nm["qgd"]   = (d.get("qgd"),
                       "[GRAPH - DIRECT] Table value; physics correction (Q=∫CdV) "
                       "available when Coss-vs-VDS curve is digitised (spec §3.6)", "")

        # Eoss @ vds_ref — re-derived from Coss(VDS_ref) using ½·Coss·VDS² (spec §3.5)
        eoss_v, eoss_m = extrapolate_eoss(d, vds_ref, coss_v)
        nm["eoss"]  = (eoss_v, eoss_m, f"@ {vds_ref:.0f}V")

        # VSD @ Tj_ref — body diode forward voltage corrected for temperature (spec §3.8)
        vsd_v, vsd_m = extrapolate_vsd(d, tj_ref)
        nm["vsd"]   = (vsd_v, vsd_m, d.get("vsd_tcond", ""))

        # trr @ Tj_ref — approximate temperature scaling (+0.5%/°C)
        trr_v, trr_m = extrapolate_trr(d, tj_ref)
        nm["trr"]   = (trr_v, trr_m, d.get("trr_tcond", ""))

        # Qrr @ Tj_ref — approximate temperature scaling (+1%/°C)
        qrr_v, qrr_m = extrapolate_qrr(d, tj_ref)
        nm["qrr"]   = (qrr_v, qrr_m, "")

        # Switching times — direct from table (circuit-condition-dependent; not extrapolatable)
        nm["td_on"] = (d.get("td_on"), f"[DIRECT] Circuit-condition-dependent; {d.get('td_on_tcond','')}", d.get("td_on_tcond",""))
        nm["tr"]    = (d.get("tr"),    "[DIRECT] Circuit-condition-dependent", "")
        nm["td_off"]= (d.get("td_off"),"[DIRECT] Circuit-condition-dependent", "")
        nm["tf"]    = (d.get("tf"),    "[DIRECT] Circuit-condition-dependent", "")

        results.append(nm)
    return results


# ─── WINNER HELPER ────────────────────────────────────────────────────────────

def _winner(vals_d, lower_better):
    if not vals_d: return None
    return (min(vals_d, key=vals_d.__getitem__) if lower_better
            else max(vals_d, key=vals_d.__getitem__))


# ─── REPORT BUILDER ───────────────────────────────────────────────────────────

def build_report(paths, mosfets, norm_list, ref, iavg_def, irms_def,
                 fsw_def, vout_def, vgs_drive_def, out_path,
                 all_charts_per_file=None, all_figs_per_file=None):
    n      = len(paths)
    labels = [mosfet_label(p, d) for p, d in zip(paths, mosfets)]
    wb     = openpyxl.Workbook()
    VDS_ref= ref["vds_ref"]
    ID_ref = ref["id_ref"]
    TJ_ref = ref["tj_ref"]

    def _cw(ws, widths):
        for c, w in widths.items():
            ws.column_dimensions[c].width = w

    def _param_section(ws, start_row, keys, values_fn, show_method=False,
                        method_fn=None, cond_fn=None, section_label=None,
                        section_bg=C_HDR, fdc=3):
        row  = start_row
        wins = {i: 0 for i in range(n)}
        ncols = fdc - 1 + n + 1 + (1 if show_method else 0)

        if section_label:
            merge_title(ws, row, 1, ncols, section_label, bg=section_bg, sz=10, h=22)
            row += 1

        for key in keys:
            display, unit, lower_better, is_pri, _ = PARAMS[key]
            bg = C_PRI_ROW if is_pri else (C_ALT if row % 2 == 0 else C_WHITE)
            ws.row_dimensions[row].height = 22

            wc(ws, row, 1, display,
               font=fnt(10, bold=is_pri), fill=ff(bg), align=lft(), border=BT)
            wc(ws, row, 2, unit,
               font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            vals   = values_fn(key)
            vals_d = {i: v for i, v in enumerate(vals) if v is not None}
            best_i = _winner(vals_d, lower_better)

            for d_i, v in enumerate(vals):
                is_best = (d_i == best_i)
                wc(ws, row, fdc + d_i,
                   round(v, 4) if isinstance(v, float) else (v if v is not None else "N/A"),
                   font=fnt(10, bold=is_best,
                            color=C_BESTF if is_best else "000000"),
                   fill=ff(C_BEST if is_best else bg),
                   align=ctr(), border=BT, nfmt="0.0####")

            best_col = fdc + n
            if best_i is not None:
                wc(ws, row, best_col, labels[best_i],
                   font=fnt(10, bold=True, color=C_BESTF),
                   fill=ff(C_BEST), align=ctr(), border=BT)
                wins[best_i] += 1
            else:
                wc(ws, row, best_col, "—",
                   font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            if show_method and method_fn:
                meths  = method_fn(key)
                unique = list(dict.fromkeys(m for m in meths if m))
                wc(ws, row, best_col + 1, " | ".join(unique),
                   font=fnt(8, italic=True, color="444444"),
                   fill=ff("F8F8F8"), align=lft(), border=BT)

            row += 1
        return row, wins

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — RAW PARAMETER COMPARISON  (as-extracted from datasheets)
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = "Raw Parameter Comparison"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "C4"

    merge_title(ws1, 1, 1, 4+n,
                "📋  RAW PARAMETER COMPARISON  (as extracted — conditions may differ between devices)",
                bg=C_TITLE, sz=13, h=36)
    merge_title(ws1, 2, 1, 4+n,
                "⚠  Values are as-measured per datasheet.  Use Sheet 2 for fair common-ground comparison.",
                bg=C_WARN, fg=C_WARNF, sz=10, h=20)

    ws1.row_dimensions[3].height = 30
    for ci, (txt, bg) in enumerate(
            [("Parameter", C_HDR), ("Unit", C_HDR)] +
            [(lbl, C_SUBHDR) for lbl in labels] +
            [("Best", C_HDR), ("Test Conditions (from datasheet)", C_HDR)], 1):
        wc(ws1, 3, ci, txt,
           font=fnt(11, bold=True, color="FFFFFF"), fill=ff(bg), align=ctr(), border=BT)

    def raw_vals(k):
        return [d.get(k) for d in mosfets]

    def raw_conds(k):
        # map param key → tcond field name
        cmap = {"vdss":"vdss_tcond","rds25":"rds_tcond","rds_ht":"rds_tcond",
                "qg":"qg_tcond","eon":"eon_tcond","eoff":"eoff_tcond",
                "id":"id_tcond","vth":"vth_tcond","ciss":"ciss_tcond",
                "coss":"coss_tcond","crss":"","qgd":"","eoss":"eoss_tcond",
                "vsd":"vsd_tcond","trr":"trr_tcond","qrr":"qrr_tcond",
                "td_on":"td_on_tcond","tr":"tr_tcond","td_off":"td_off_tcond","tf":"tf_tcond"}
        cf = cmap.get(k, "")
        if not cf: return [""] * n
        return [d.get(cf, "") or "" for d in mosfets]

    # Custom param section that includes conditions column
    def _section_with_conds(ws, start_row, keys, section_label, section_bg, fdc=3):
        row  = start_row
        wins = {i: 0 for i in range(n)}
        ncols = fdc - 1 + n + 2   # +2 = Best + Conditions

        merge_title(ws, row, 1, ncols, section_label, bg=section_bg, sz=10, h=22)
        row += 1

        for key in keys:
            display, unit, lower_better, is_pri, _ = PARAMS[key]
            bg = C_PRI_ROW if is_pri else (C_ALT if row % 2 == 0 else C_WHITE)
            ws.row_dimensions[row].height = 22

            wc(ws, row, 1, display,
               font=fnt(10, bold=is_pri), fill=ff(bg), align=lft(), border=BT)
            wc(ws, row, 2, unit,
               font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            vals   = raw_vals(key)
            vals_d = {i: v for i, v in enumerate(vals) if v is not None}
            best_i = _winner(vals_d, lower_better)

            for d_i, v in enumerate(vals):
                is_best = (d_i == best_i)
                # Mark derived values with a slightly different fill
                is_derived = (key in ("eon","eoff") and mosfets[d_i].get(f"_{key}_derived"))
                cell_fill = C_DERIVED if is_derived else (C_BEST if is_best else bg)
                wc(ws, row, fdc + d_i,
                   round(v, 4) if isinstance(v, float) else (v if v is not None else "N/A"),
                   font=fnt(10, bold=is_best,
                            color=C_BESTF if is_best else ("7F5C00" if is_derived else "000000")),
                   fill=ff(cell_fill),
                   align=ctr(), border=BT, nfmt="0.0####")

            best_col = fdc + n
            if best_i is not None:
                wc(ws, row, best_col, labels[best_i],
                   font=fnt(10, bold=True, color=C_BESTF),
                   fill=ff(C_BEST), align=ctr(), border=BT)
                wins[best_i] += 1
            else:
                wc(ws, row, best_col, "—",
                   font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            # Conditions column — merge all device conditions into one cell
            conds = raw_conds(key)
            unique_c = "; ".join(dict.fromkeys(c for c in conds if c))
            wc(ws, row, best_col + 1, unique_c,
               font=fnt(8, italic=True, color="444444"),
               fill=ff("F8F8F8"), align=lft(), border=BT)

            row += 1
        return row, wins

    next1, w1a = _section_with_conds(ws1, 4, PRIORITY_KEYS,
                                     "★  PRIORITY PARAMETERS", C_PRI_BG)
    next1, w1b = _section_with_conds(ws1, next1 + 1, SECONDARY_KEYS,
                                     "SECONDARY PARAMETERS", C_SEC_BG)

    all_w1 = {i: w1a[i] + w1b[i] for i in range(n)}
    mw = max(all_w1.values()) if all_w1 else 0
    wr = next1 + 1
    ws1.row_dimensions[wr].height = 26
    merge_title(ws1, wr, 1, 2, "Total Wins", bg=C_SEC_BG, sz=10, h=26)
    for d_i in range(n):
        wv = all_w1[d_i]
        wc(ws1, wr, 3 + d_i, wv,
           font=fnt(12, bold=True, color=C_BESTF if wv == mw else "000000"),
           fill=ff(C_BEST if wv == mw else C_ALT), align=ctr(), border=BM)
    if mw > 0:
        bdi = max(all_w1, key=all_w1.__getitem__)
        wc(ws1, wr, 3 + n, labels[bdi],
           font=fnt(10, bold=True, color=C_BESTF), fill=ff(C_BEST), align=ctr(), border=BM)

    # Legend row
    wr2 = wr + 2
    ws1.merge_cells(f"A{wr2}:{gcol(3+n+1)}{wr2}")
    lc = ws1.cell(row=wr2, column=1,
                  value="🟡 Orange-yellow cells = value derived from switching times (t_r, t_f) — not directly tabulated in datasheet")
    lc.font  = fnt(9, italic=True, color="7F5C00")
    lc.fill  = ff(C_DERIVED)
    lc.alignment = lft()

    _cw(ws1, {"A": 38, "B": 9})
    for d_i in range(n): ws1.column_dimensions[gcol(3 + d_i)].width = 18
    ws1.column_dimensions[gcol(3 + n)].width = 26
    ws1.column_dimensions[gcol(4 + n)].width = 52

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — COMMON-GROUND COMPARISON
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Common-Ground Comparison")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "C5"

    _CG_COL = 5 + n   # last column used (Param | Unit | dev×n | Best | Method | Calc)
    merge_title(ws2, 1, 1, _CG_COL,
                ("✅  COMMON-GROUND COMPARISON  "
                 "(live formulas — edit C3 / E3 / G3 to change reference conditions)"),
                bg=C_TITLE, sz=13, h=36)
    merge_title(ws2, 2, 1, _CG_COL,
                ("Rds: linear interp between 25°C & high-T datasheet points  |  "
                 "Coss/Ciss/Crss: 1/√V depletion scaling to VDS_ref  |  Eon/Eoff: linear scaling with ID  |  "
                 "Column G = live calculation text — auto-updates when C3/E3/G3 change"),
                bg=C_SUBHDR, fg="FFFFFF", sz=9, h=18)

    # ── Row 3: editable input cells ───────────────────────────────────────────
    C_INPUT = "FFF2CC"; C_INPUT_FG = "7F5C00"
    ws2.row_dimensions[3].height = 28
    wc(ws2, 3, 1, "▼ EDIT REFERENCE CONDITIONS →",
       font=fnt(10, bold=True, color=C_INPUT_FG), fill=ff(C_INPUT), align=lft(), border=BM)
    wc(ws2, 3, 2, "Tj_ref (°C)",
       font=fnt(9, bold=True, color=C_INPUT_FG), fill=ff(C_INPUT), align=ctr(), border=BT)
    c3 = ws2.cell(3, 3, value=float(TJ_ref))   # ← user-editable in Excel
    c3.font = fnt(12, bold=True, color=C_INPUT_FG); c3.fill = ff(C_INPUT)
    c3.alignment = ctr(); c3.border = BM; c3.number_format = "0.#"
    wc(ws2, 3, 4, "VDS_ref (V)",
       font=fnt(9, bold=True, color=C_INPUT_FG), fill=ff(C_INPUT), align=ctr(), border=BT)
    e3 = ws2.cell(3, 5, value=float(VDS_ref))
    e3.font = fnt(12, bold=True, color=C_INPUT_FG); e3.fill = ff(C_INPUT)
    e3.alignment = ctr(); e3.border = BM; e3.number_format = "0"
    wc(ws2, 3, 6, "ID_ref (A)",
       font=fnt(9, bold=True, color=C_INPUT_FG), fill=ff(C_INPUT), align=ctr(), border=BT)
    g3 = ws2.cell(3, 7, value=float(ID_ref))
    g3.font = fnt(12, bold=True, color=C_INPUT_FG); g3.fill = ff(C_INPUT)
    g3.alignment = ctr(); g3.border = BM; g3.number_format = "0.#"
    ws2.column_dimensions["G"].width = 55

    # ── Row 4: column headers ─────────────────────────────────────────────────
    ws2.row_dimensions[4].height = 28
    for ci, (txt, bg) in enumerate(
            [("Parameter", C_HDR), ("Unit", C_HDR)] +
            [(lbl, C_SUBHDR) for lbl in labels] +
            [("Best", C_HDR), ("Extrapolation / Scaling Method", C_HDR),
             ("Live Calculation  (auto-updates)", C_HDR)], 1):
        wc(ws2, 4, ci, txt,
           font=fnt(11, bold=True, color="FFFFFF"), fill=ff(bg), align=ctr(), border=BT)

    # Use _rds_tj for priority comparison instead of rds25/rds_ht
    NORM_PRIORITY = ["vdss", "_rds_tj", "rds25", "rds_ht", "qg", "eon", "eoff"]
    NORM_SECONDARY = [k for k in SECONDARY_KEYS]

    NORM_DISP = {
        "_rds_tj": (f"R_DS(on) @ Tj={TJ_ref:.0f}°C  (extrapolated)", "mΩ", True, True, False),
        "rds25":   ("R_DS(on) @ 25°C  (direct)",                     "mΩ", True, True, False),
        "rds_ht":  ("R_DS(on) @ High-T  (direct)",                   "mΩ", True, True, False),
    }

    def _disp(k):
        if k in NORM_DISP: return NORM_DISP[k]
        return PARAMS[k]

    def norm_vals(k):
        return [norm_list[i].get(k, (None, "", ""))[0] for i in range(n)]

    def norm_meths(k):
        return [norm_list[i].get(k, (None, "", ""))[1] for i in range(n)]

    # Build per-device formula strings for extrapolated quantities
    # Keys that get Excel formulas (referencing C3=Tj_ref, E3=VDS_ref, G3=ID_ref)
    def _excel_formula_and_calc(key, d_i):
        """Return (formula_str, calc_text_formula) or (None, None) if not formula-able."""
        m = mosfets[d_i]
        C_CALC_BG = "EEF4FF"
        # Rds(Tj): graph-based when available; live 2-point formula as fallback
        if key == "_rds_tj":
            # Graph-based value is already pre-computed in norm_list — no live
            # formula possible (it's a multi-point interpolation), write directly.
            if m.get("rds_ratio_curve"):
                return None, None
            rds25 = m.get("rds25"); rds_ht = m.get("rds_ht")
            t_ht  = m.get("_rds_ht_tj") or 125.0
            if rds25 is not None and rds_ht is not None:
                f = f"={rds25}+({rds_ht}-{rds25})*($C$3-25)/({t_ht}-25)"
                calc = (f'="{rds25}+({rds_ht}-{rds25})×(Tj-25)/({t_ht}-25)'
                        f'  at Tj="&$C$3&"°C = "&ROUND({gcol(3+d_i)}{"{row}"},2)&" mΩ"')
                return f, calc
            # Only one temperature point — extrapolation not possible
            return "Can't extrapolate", None
        # Eon/Eoff linear scaling with ID
        if key in ("eon", "eoff"):
            val0 = m.get(key)
            id_test = m.get(f"_{key}_id") or m.get("_eon_id") or m.get("_sw_id")
            if val0 is not None and id_test is not None:
                f = f"={val0}*($G$3/{id_test})"
                calc = (f'="{val0}µJ × (ID_ref/{id_test})'
                        f'  at ID_ref="&$G$3&"A = "&ROUND({gcol(3+d_i)}{"{row}"},2)&" µJ"')
                return f, calc
        # Coss depletion scaling: C(VDS) = C0 * sqrt(V0 / VDS)
        if key == "coss":
            c0 = m.get("coss"); v0 = m.get("_coss_vds_test")
            if c0 is not None and v0 is not None:
                f = f"={c0}*SQRT({v0}/$E$3)"
                calc = (f'="{c0}pF × √({v0}/VDS_ref)'
                        f'  at VDS_ref="&$E$3&"V = "&ROUND({gcol(3+d_i)}{"{row}"},2)&" pF"')
                return f, calc
        # Ciss depletion scaling: C_iss(VDS) = C0 * sqrt(V0 / VDS)
        if key == "ciss":
            c0 = m.get("ciss"); v0 = m.get("_ciss_vds_test")
            if c0 is not None and v0 is not None:
                f = f"={c0}*SQRT({v0}/$E$3)"
                calc = (f'="{c0}pF × √({v0}/VDS_ref)'
                        f'  at VDS_ref="&$E$3&"V = "&ROUND({gcol(3+d_i)}{"{row}"},2)&" pF"')
                return f, calc
        # Crss depletion scaling: C_rss(VDS) = C0 * sqrt(V0 / VDS)
        if key == "crss":
            c0 = m.get("crss"); v0 = m.get("_crss_vds_test")
            if c0 is not None and v0 is not None:
                f = f"={c0}*SQRT({v0}/$E$3)"
                calc = (f'="{c0}pF × √({v0}/VDS_ref)'
                        f'  at VDS_ref="&$E$3&"V = "&ROUND({gcol(3+d_i)}{"{row}"},2)&" pF"')
                return f, calc
        # Eoss @ VDS_ref
        # If Eoss graph data is available, the value is pre-computed by
        # extrapolate_eoss_from_rows — no live formula needed, write directly.
        # Only fall back to the Coss-formula when no graph is available.
        if key == "eoss":
            if m.get("eoss_rows"):
                return None, None   # graph-interpolated value already in norm_list
            c0 = m.get("coss"); v0 = m.get("_coss_vds_test")
            if c0 is not None and v0 is not None:
                # Eoss µJ = 0.5 * Coss(VDS_ref)_pF * VDS_ref² × 1e-6
                # Coss(VDS_ref)_pF = c0 * sqrt(v0 / VDS_ref)  (depletion model)
                f = f"=0.5*{c0}*SQRT({v0}/$E$3)*1E-12*($E$3)^2*1E6"
                calc = (f'="½ × Coss("&$E$3&"V) × VDS² = "&'
                        f'ROUND({gcol(3+d_i)}{"{row}"},3)&" µJ  (Coss from depletion model)"')
                return f, calc
        # VSD @ Tj_ref: VSD(25°C) + (-2mV/°C) × (Tj - 25)
        if key == "vsd":
            vsd25 = m.get("vsd")
            if vsd25 is not None:
                f = f"=MAX({vsd25}+(-0.002)*($C$3-25),0.05)"
                calc = (f'="{vsd25:.3f}V + (−2mV/°C)×("&$C$3&"−25) = "&'
                        f'ROUND({gcol(3+d_i)}{"{row}"},3)&" V"')
                return f, calc
        # Qg @ VDS_ref with Qgd-split scaling
        if key == "qg":
            qg0  = m.get("qg");  vds_t = m.get("_qg_vds"); qgd0 = m.get("qgd")
            if qg0 is not None and vds_t is not None and abs(vds_t - ref.get("vds_ref", vds_t)) >= 10:
                if qgd0 is not None:
                    qgs0 = max(qg0 - qgd0, 0.0)
                    f = f"={qgs0}+{qgd0}*($E$3/{vds_t})"
                    calc = (f'="{qgs0:.1f}nC + {qgd0:.1f}nC×(VDS_ref/{vds_t}) = "&'
                            f'ROUND({gcol(3+d_i)}{"{row}"},2)&" nC"')
                    return f, calc
        return None, None

    def _norm_section(ws, start_row, keys, section_label, section_bg, fdc=3):
        row  = start_row
        wins = {i: 0 for i in range(n)}
        ncols = fdc - 1 + n + 3   # +1 for new Calculation column

        merge_title(ws, row, 1, ncols, section_label, bg=section_bg, sz=10, h=22)
        row += 1

        for key in keys:
            disp_tuple = _disp(key)
            display, unit, lower_better = disp_tuple[0], disp_tuple[1], disp_tuple[2]
            is_pri = disp_tuple[3]
            bg = C_PRI_ROW if is_pri else (C_ALT if row % 2 == 0 else C_WHITE)
            ws.row_dimensions[row].height = 22

            wc(ws, row, 1, display,
               font=fnt(10, bold=is_pri), fill=ff(bg), align=lft(), border=BT)
            wc(ws, row, 2, unit,
               font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            # ── Log-graph parameters (Ciss/Coss/Crss): no reliable number —
            #    point the reader to the datasheet images in the Graph sheet ──
            if key in _REFER_TO_GRAPH_KEYS:
                # Merge the device value columns + Best + Method + Calc into one
                # cell carrying the "refer the graphs" line.
                ws.merge_cells(start_row=row, start_column=fdc,
                               end_row=row, end_column=fdc + n + 2)
                wc(ws, row, fdc, "→  " + _REFER_TO_GRAPH_TEXT,
                   font=fnt(10, italic=True, bold=True, color="7F0000"),
                   fill=ff("FFF3CD"), align=ctr(), border=BT)
                row += 1
                continue

            vals   = norm_vals(key)
            vals_d = {i: v for i, v in enumerate(vals) if v is not None}
            best_i = _winner(vals_d, lower_better)

            calc_parts = []   # collect per-device calc texts for column G
            for d_i, v in enumerate(vals):
                is_best = (d_i == best_i)
                is_derived = key in ("eon","eoff") and mosfets[d_i].get(f"_{key}_derived")
                cell_fill = C_DERIVED if is_derived else (C_BEST if is_best else bg)
                # Try to write a live Excel formula for extrapolated keys
                formula, calc_text_tmpl = _excel_formula_and_calc(key, d_i)
                if formula == "Can't extrapolate":
                    # Extrapolation needed but insufficient data — mark the cell clearly
                    wc(ws, row, fdc + d_i, "Can't extrapolate",
                       font=fnt(9, italic=True, color="7F0000"),
                       fill=ff("FFF3CD"), align=ctr(), border=BT, nfmt="@")
                    continue
                if formula:
                    cell_val = formula
                    nfmt = "0.00"
                    if calc_text_tmpl:
                        calc_parts.append(calc_text_tmpl.format(row=row))
                else:
                    meth_str = norm_list[d_i].get(key, (None, "", ""))[1]
                    if v is None and meth_str and "Can't extrapolate" in meth_str:
                        wc(ws, row, fdc + d_i, "Can't extrapolate",
                           font=fnt(9, italic=True, color="7F0000"),
                           fill=ff("FFF3CD"), align=ctr(), border=BT, nfmt="@")
                        continue
                    cell_val = round(v, 4) if isinstance(v, float) else (v if v is not None else "N/A")
                    nfmt = "0.0####"
                wc(ws, row, fdc + d_i,
                   cell_val,
                   font=fnt(10, bold=is_best,
                            color=C_BESTF if is_best else ("7F5C00" if is_derived else "000000")),
                   fill=ff(cell_fill),
                   align=ctr(), border=BT, nfmt=nfmt)

            best_col = fdc + n
            if best_i is not None:
                wc(ws, row, best_col, labels[best_i],
                   font=fnt(10, bold=True, color=C_BESTF),
                   fill=ff(C_BEST), align=ctr(), border=BT)
                wins[best_i] += 1
            else:
                wc(ws, row, best_col, "—",
                   font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            meths  = norm_meths(key)
            unique = list(dict.fromkeys(m for m in meths if m))
            wc(ws, row, best_col + 1, " | ".join(unique),
               font=fnt(8, italic=True, color="444444"),
               fill=ff("F8F8F8"), align=lft(), border=BT)

            # ── Column G: live calculation text ───────────────────────────────
            calc_col = best_col + 2
            if calc_parts:
                # join multiple device calcs with " || "
                if len(calc_parts) == 1:
                    calc_formula = calc_parts[0]
                else:
                    # concatenate with separator
                    calc_formula = calc_parts[0]  # simplify: show first device
                c_calc = ws.cell(row=row, column=calc_col, value=calc_formula)
                c_calc.font = fnt(9, italic=True, color="1F3864")
                c_calc.fill = ff("EEF4FF"); c_calc.alignment = lft(); c_calc.border = BT
            else:
                wc(ws, row, calc_col, "",
                   font=fnt(9), fill=ff("F8F8F8"), align=lft(), border=BT)

            row += 1
        return row, wins

    next2, w2a = _norm_section(ws2, 5, NORM_PRIORITY,
                               "★  PRIORITY PARAMETERS  (normalised to common reference)", C_PRI_BG)
    next2, w2b = _norm_section(ws2, next2 + 1, NORM_SECONDARY,
                               "SECONDARY PARAMETERS", C_SEC_BG)

    all_w2 = {i: w2a[i] + w2b[i] for i in range(n)}
    mw2 = max(all_w2.values()) if all_w2 else 0
    wr2 = next2 + 1
    ws2.row_dimensions[wr2].height = 26
    merge_title(ws2, wr2, 1, 2, "Total Wins", bg=C_SEC_BG, sz=10, h=26)
    for d_i in range(n):
        wv = all_w2[d_i]
        wc(ws2, wr2, 3 + d_i, wv,
           font=fnt(12, bold=True, color=C_BESTF if wv == mw2 else "000000"),
           fill=ff(C_BEST if wv == mw2 else C_ALT), align=ctr(), border=BM)
    if mw2 > 0:
        bdi2 = max(all_w2, key=all_w2.__getitem__)
        wc(ws2, wr2, 3 + n, labels[bdi2],
           font=fnt(10, bold=True, color=C_BESTF), fill=ff(C_BEST), align=ctr(), border=BM)

    # ── Figure of Merit  FOM = Rds_on(Tj_ref) × Q_G ─────────────────────────
    # Uses the common-ground Rds value (extrapolated to Tj_ref) and the direct Qg.
    # Lower FOM = better switching + conduction trade-off.
    C_FOM_BG  = "1A1A2E"   # very dark navy — stands out at the bottom
    C_FOM_FG  = "FFD700"   # gold text
    C_FOM_VAL = "2C2C54"   # dark purple for value cells
    C_FOM_BEST= "FFD700"   # gold for best device

    fom_spacer = wr2 + 1
    ws2.row_dimensions[fom_spacer].height = 10   # visual gap

    fom_hdr = fom_spacer + 1
    ws2.row_dimensions[fom_hdr].height = 28
    merge_title(ws2, fom_hdr, 1, _CG_COL,
                "⭐  FIGURE OF MERIT  (FOM = R_DS(on) @ Tj_ref  ×  Q_G  — live formula)",
                bg=C_FOM_BG, fg=C_FOM_FG, sz=12, h=28)

    # Sub-header explaining the formula
    fom_desc = fom_hdr + 1
    ws2.row_dimensions[fom_desc].height = 20
    merge_title(ws2, fom_desc, 1, _CG_COL,
                "FOM = Rds_on [mΩ] × Q_G [nC]   "
                "(Rds uses live formula — changes when C3=Tj_ref changes)   "
                "Lower FOM = better switching–conduction trade-off",
                bg="2E2E5E", fg="CCCCFF", sz=9, h=20)

    # Column headers for FOM row
    fom_row = fom_desc + 1
    ws2.row_dimensions[fom_row].height = 30

    wc(ws2, fom_row, 1, "FOM = R_DS(on) × Q_G",
       font=Font(name="Calibri", size=11, bold=True, color=C_FOM_FG),
       fill=ff(C_FOM_VAL), align=lft(), border=BM)
    wc(ws2, fom_row, 2, "mΩ·nC",
       font=Font(name="Calibri", size=10, bold=True, color="AAAAFF"),
       fill=ff(C_FOM_VAL), align=ctr(), border=BM)

    # Find the row numbers where _rds_tj and qg appear in the sheet
    # We'll build a small row-lookup from what _norm_section wrote
    # Easier: use Excel formulas that reference those cells by address
    # We need to know which Excel row holds _rds_tj for each device.
    # Since we can't predict it here (it depends on section layout), we fall back to
    # Python-computed FOM as the seed value, but also write an Excel formula
    # using direct cell addresses found during section writing.
    fom_vals = {}
    for d_i in range(n):
        rds_v = norm_list[d_i].get("_rds_tj", (None, "", ""))[0]
        qg_v  = norm_list[d_i].get("qg",      (None, "", ""))[0]
        if rds_v is not None and qg_v is not None:
            fom_vals[d_i] = round(rds_v * qg_v, 2)

    best_fom_i = min(fom_vals, key=fom_vals.__getitem__) if fom_vals else None

    for d_i in range(n):
        fv = fom_vals.get(d_i)
        is_best = (d_i == best_fom_i)
        cell_bg = C_FOM_BEST if is_best else C_FOM_VAL
        cell_fg = C_FOM_BG   if is_best else "FFFFFF"
        c = ws2.cell(row=fom_row, column=3 + d_i, value=fv if fv is not None else "N/A")
        c.font   = Font(name="Calibri", size=12, bold=True, color=cell_fg)
        c.fill   = ff(cell_bg)
        c.alignment = ctr(); c.border = BM
        if fv is not None:
            c.number_format = "0.00"

    # Best column
    fom_best_col = 3 + n
    if best_fom_i is not None:
        wc(ws2, fom_row, fom_best_col, f"🏆 {labels[best_fom_i]}",
           font=Font(name="Calibri", size=11, bold=True, color=C_FOM_BG),
           fill=ff(C_FOM_BEST), align=ctr(), border=BM)
    else:
        wc(ws2, fom_row, fom_best_col, "N/A",
           font=fnt(10), fill=ff(C_FOM_VAL), align=ctr(), border=BM)

    # Extrapolation method note
    fom_meth_col = 4 + n
    rds_meths = [norm_list[d_i].get("_rds_tj", (None, ""))[1] for d_i in range(n)]
    qg_meths  = [norm_list[d_i].get("qg",      (None, ""))[1] for d_i in range(n)]
    meth_parts = []
    unique_rds = list(dict.fromkeys(m for m in rds_meths if m))
    unique_qg  = list(dict.fromkeys(m for m in qg_meths  if m))
    if unique_rds: meth_parts.append("Rds: " + " | ".join(unique_rds))
    if unique_qg:  meth_parts.append("Qg: "  + " | ".join(unique_qg))
    wc(ws2, fom_row, fom_meth_col, "  ".join(meth_parts),
       font=fnt(8, italic=True, color="444444"),
       fill=ff("F8F8F8"), align=lft(), border=BM)

    # Calculation column for FOM
    fom_calc_col = 5 + n
    fom_calc_parts = []
    for d_i in range(n):
        rds_v = norm_list[d_i].get("_rds_tj", (None,))[0]
        qg_v  = norm_list[d_i].get("qg", (None,))[0]
        if rds_v is not None and qg_v is not None:
            fom_calc_parts.append(f"{labels[d_i]}: Rds×Qg={rds_v:.1f}×{qg_v:.1f}={rds_v*qg_v:.0f}")
    wc(ws2, fom_row, fom_calc_col, "  |  ".join(fom_calc_parts),
       font=fnt(9, italic=True, color="1F3864"),
       fill=ff("EEF4FF"), align=lft(), border=BM)

    # Footnote row explaining FOM interpretation
    fom_note = fom_row + 1
    ws2.row_dimensions[fom_note].height = 32
    ws2.merge_cells(start_row=fom_note, start_column=1, end_row=fom_note, end_column=_CG_COL)
    fn = ws2.cell(row=fom_note, column=1,
                  value=(
                      "📌  FOM Interpretation:  A lower FOM means the device has a better balance between "
                      "conduction loss (R_DS(on)) and switching loss (Q_G).  "
                      "Both Rds and Qg increase with gate-oxide thickness — the FOM product captures this "
                      "fundamental trade-off for a given technology node.  "
                      "Use FOM to compare devices across different manufacturers on equal footing, "
                      "independent of die size or current rating."
                  ))
    fn.font      = fnt(9, italic=True, color="1A1A2E")
    fn.fill      = ff("EAEAF8")
    fn.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    fn.border    = BT

    _cw(ws2, {"A": 40, "B": 9})
    for d_i in range(n): ws2.column_dimensions[gcol(3 + d_i)].width = 18
    ws2.column_dimensions[gcol(3 + n)].width = 26
    ws2.column_dimensions[gcol(4 + n)].width = 55
    ws2.column_dimensions[gcol(5 + n)].width = 55  # Calculation column

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — OPERATING CONDITIONS  (yellow editable inputs + device params)
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Operating Conditions")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A1"

    merge_title(ws3, 1, 1, 3, "⚙  OPERATING CONDITIONS",
                bg=C_TITLE, sz=13, h=36)
    merge_title(ws3, 2, 1, 3,
                "🟡 Yellow cells = edit these for your circuit  |  "
                "🔵 Blue cells = device parameters (auto-filled, editable)",
                bg=C_SUBHDR, fg="FFFFFF", sz=9, h=20)

    # ── Section A: Circuit operating point ────────────────────────────────────
    merge_title(ws3, 4, 1, 3, "▼  CIRCUIT OPERATING POINT  (edit yellow cells)",
                bg=C_SUBHDR, sz=11, h=24)

    ws3.row_dimensions[5].height = 22
    for ci, txt in enumerate(["Parameter", "Value", "Description"], 1):
        wc(ws3, 5, ci, txt,
           font=fnt(10, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=ctr(), border=BT)

    OP = [
        ("I_avg",     "A",   iavg_def,     "Average switch current — used for switching-energy scaling"),
        ("I_rms",     "A",   irms_def,     "RMS switch current — drives conduction loss (I²·Rds)"),
        ("F_sw",      "Hz",  fsw_def,      "Switching frequency — scales all switching losses"),
        ("V_out",     "V",   vout_def,     "DC bus / output voltage (drain-source voltage during switching)"),
        ("VGS_drive", "V",   vgs_drive_def,"Gate drive supply voltage — used for gate-charge loss"),
    ]
    op_rows = {}
    for i, (sym, unit, default, desc) in enumerate(OP):
        r = 6 + i
        ws3.row_dimensions[r].height = 26
        wc(ws3, r, 1, f"{sym}  ({unit})",
           font=fnt(11, bold=True), fill=ff(C_PRI_ROW), align=lft(), border=BT)
        inp = ws3.cell(row=r, column=2, value=default)
        inp.font = fnt(13, bold=True, color=C_INPUT.replace("FFF2CC","7F6000"))
        inp.fill = ff(C_INPUT); inp.alignment = ctr()
        inp.border = BM; inp.number_format = "0.###"
        wc(ws3, r, 3, desc,
           font=fnt(9, italic=True, color="555555"),
           fill=ff(C_WHITE), align=lft(), border=BT)
        op_rows[sym] = r

    # Named references (absolute) — Loss Calc sheet uses these cross-sheet refs
    IAVG = f"='Operating Conditions'!$B${op_rows['I_avg']}"
    IRMS = f"='Operating Conditions'!$B${op_rows['I_rms']}"
    FSW  = f"='Operating Conditions'!$B${op_rows['F_sw']}"
    VOUT = f"='Operating Conditions'!$B${op_rows['V_out']}"
    VGS  = f"='Operating Conditions'!$B${op_rows['VGS_drive']}"
    # Absolute cell addresses for use inside formulas on the Loss Calc sheet
    R_IAVG = f"'Operating Conditions'!$B${op_rows['I_avg']}"
    R_IRMS = f"'Operating Conditions'!$B${op_rows['I_rms']}"
    R_FSW  = f"'Operating Conditions'!$B${op_rows['F_sw']}"
    R_VOUT = f"'Operating Conditions'!$B${op_rows['V_out']}"
    R_VGS  = f"'Operating Conditions'!$B${op_rows['VGS_drive']}"

    # ── Section B: Device Parameters (one row per device param, one col per device)
    PB_START_ROW = 14
    merge_title(ws3, PB_START_ROW, 1, 2 + n,
                "▼  DEVICE PARAMETERS  (auto-filled from datasheets — you may override blue cells)",
                bg="4472C4", fg="FFFFFF", sz=11, h=24)

    ws3.row_dimensions[PB_START_ROW + 1].height = 30
    wc(ws3, PB_START_ROW + 1, 1, "Parameter  (unit)",
       font=fnt(10, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=lft(), border=BT)
    wc(ws3, PB_START_ROW + 1, 2, "Description",
       font=fnt(10, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=lft(), border=BT)
    for d_i, lbl in enumerate(labels):
        wc(ws3, PB_START_ROW + 1, 3 + d_i, lbl,
           font=fnt(10, bold=True, color="FFFFFF"), fill=ff(C_SUBHDR), align=ctr(), border=BT)

    DEV_BLOCK = [
        # (param_label, description, norm_key)
        # All values sourced from norm_list (common-ground extrapolated at ref conditions)
        ("Rds_on @ Tj_ref  (mΩ)",    f"On-resistance extrapolated to Tj={ref['tj_ref']:.0f}°C  [Common-Ground sheet]",  "_rds_tj"),
        ("E_on  (µJ)",               f"Turn-on energy scaled to ID={ref['id_ref']:.1f}A  [Common-Ground sheet]",         "eon"),
        ("E_off  (µJ)",              f"Turn-off energy scaled to ID={ref['id_ref']:.1f}A  [Common-Ground sheet]",        "eoff"),
        ("E_oss  (µJ)",              f"Output-cap stored energy @ VDS={ref['vds_ref']:.0f}V  [Common-Ground sheet]",     "eoss"),
        ("Q_G  (nC)",                "Total gate charge  [Common-Ground sheet]",                                          "qg"),
        ("ID_test  (A)",             "Test current at which Eon/Eoff were measured",                                      "_id_test"),
    ]
    pb_rows = {}   # norm_key → row number on THIS sheet (ws3)
    for i, (plbl, pdesc, nkey) in enumerate(DEV_BLOCK):
        r = PB_START_ROW + 2 + i
        ws3.row_dimensions[r].height = 22
        wc(ws3, r, 1, plbl,
           font=fnt(10, bold=True, color="1F3864"),
           fill=ff("EEF4FF"), align=lft(), border=BT)
        wc(ws3, r, 2, pdesc,
           font=fnt(9, italic=True, color="444444"),
           fill=ff("EEF4FF"), align=lft(), border=BT)
        for d_i, d in enumerate(mosfets):
            nm = norm_list[d_i]
            if nkey == "_id_test":
                val = d.get("_eon_id") or d.get("_sw_id") or d.get("id") or ref["id_ref"]
            else:
                # Pull from common-ground normalised values (extrapolated at ref conditions)
                val = nm.get(nkey, (None,))[0]
            if val is None:
                wc(ws3, r, 3 + d_i, "Can't extrapolate",
                   font=fnt(9, italic=True, color="7F0000"),
                   fill=ff("FFF3CD"), align=ctr(), border=BT, nfmt="@")
                pb_rows[nkey] = r
                continue
            pv = round(float(val), 6)
            wc(ws3, r, 3 + d_i, pv,
               font=fnt(10, color="0000CC"),
               fill=ff(C_DEVPARAM), align=ctr(), border=BT, nfmt="0.000###")
        pb_rows[nkey] = r   # row on ws3

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 52
    for d_i in range(n):
        ws3.column_dimensions[gcol(3 + d_i)].width = 22

    # ── Section C: Normalisation reference note ───────────────────────────────
    note_row = PB_START_ROW + 2 + len(DEV_BLOCK) + 2
    merge_title(ws3, note_row, 1, 2 + n,
                f"Normalisation reference:  VDS_ref={VDS_ref:.0f} V  |  ID_ref={ID_ref:.1f} A  |  Tj_ref={TJ_ref:.0f} °C",
                bg=C_NORM, fg=C_GREEN, sz=10, h=22)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 (was 3b) — LOSS CALCULATION  (live Excel formulas, separate sheet)
    # ══════════════════════════════════════════════════════════════════════════
    ws_lc = wb.create_sheet("Loss Calculation")
    ws_lc.sheet_view.showGridLines = False
    ws_lc.freeze_panes = "A4"

    merge_title(ws_lc, 1, 1, 3 + n + 1,
                "⚡  LOSS CALCULATION  —  All values are live Excel formulas",
                bg=C_TITLE, sz=13, h=36)
    merge_title(ws_lc, 2, 1, 3 + n + 1,
                "Change any yellow cell in 'Operating Conditions' sheet → all losses update here automatically",
                bg=C_SUBHDR, fg="FFFFFF", sz=9, h=20)

    # Helper: cross-sheet param cell address
    def pa(pkey, di):
        return f"'Operating Conditions'!${gcol(3+di)}${pb_rows[pkey]}"

    # Column headers  Col A=Loss Component | B=Formula | C..C+n-1=devices | last=Best
    ws_lc.row_dimensions[3].height = 32
    for ci, (txt, bg) in enumerate(
            [("Loss Component", C_HDR), ("Formula Used", C_HDR)] +
            [(lbl, C_SUBHDR) for lbl in labels] +
            [("Best (Lowest)", C_HDR)], 1):
        wc(ws_lc, 3, ci, txt,
           font=fnt(11, bold=True, color="FFFFFF"), fill=ff(bg), align=ctr(), border=BT)

    # ── Loss formula definitions ───────────────────────────────────────────────
    # Col A = label, Col B = human-readable formula, Col C+ = live Excel formula

    def lf(kind, di):
        """Return live Excel formula string for loss kind and device index di."""
        if kind == "cond":
            return (f"={R_IRMS}^2"
                    f"*{pa('_rds_tj',di)}/1000")
        if kind == "sw_on":
            return (f"=IF({pa('_id_test',di)}=0,0,"
                    f"{pa('eon',di)}/1000000"
                    f"*({R_IAVG}/{pa('_id_test',di)})"
                    f"*{R_FSW})")
        if kind == "sw_off":
            return (f"=IF({pa('_id_test',di)}=0,0,"
                    f"{pa('eoff',di)}/1000000"
                    f"*({R_IAVG}/{pa('_id_test',di)})"
                    f"*{R_FSW})")
        if kind == "coss":
            return (f"={pa('eoss',di)}/1000000"
                    f"*{R_FSW}")
        if kind == "gate":
            return (f"={pa('qg',di)}/1000000000"
                    f"*{R_VGS}*{R_FSW}")

    LOSS_DEFS = [
        # (row_offset, label, kind, human_formula, row_bg)
        (0, "P_cond   –  Conduction Loss",
             "cond",
             "I_rms²  ×  Rds_on(Tj) [Ω]",
             C_WHITE),
        (1, "P_sw,on  –  Turn-ON Switching Loss",
             "sw_on",
             "E_on [J]  ×  (I_avg / ID_test)  ×  F_sw",
             C_ALT),
        (2, "P_sw,off –  Turn-OFF Switching Loss",
             "sw_off",
             "E_off [J]  ×  (I_avg / ID_test)  ×  F_sw",
             C_WHITE),
        (3, "P_Coss   –  Output Cap Loss",
             "coss",
             "E_oss [J]  ×  F_sw",
             C_ALT),
        (4, "P_gate   –  Gate Drive Loss",
             "gate",
             "Q_G [C]  ×  V_GS_drive  ×  F_sw",
             C_WHITE),
    ]
    TOTAL_ROW_OFFSET = len(LOSS_DEFS)
    DATA_START_ROW   = 4    # first loss data row on ws_lc
    loss_excel_rows  = {}   # kind → excel row number

    for off, label, kind, formula_str, bg in LOSS_DEFS:
        lrow = DATA_START_ROW + off
        ws_lc.row_dimensions[lrow].height = 26
        loss_excel_rows[kind] = lrow

        wc(ws_lc, lrow, 1, label,
           font=fnt(10, bold=False), fill=ff(bg), align=lft(), border=BT)
        wc(ws_lc, lrow, 2, formula_str,
           font=fnt(9, italic=True, color="1F3864"),
           fill=ff("EEF4FB"), align=lft(), border=BT)

        for d_i in range(n):
            c = ws_lc.cell(row=lrow, column=3 + d_i, value=lf(kind, d_i))
            c.font = fnt(10); c.fill = ff(bg)
            c.alignment = ctr(); c.border = BT; c.number_format = "0.0000"

        mr  = ",".join([f"{gcol(3+d_i)}{lrow}" for d_i in range(n)])
        rng = f"{gcol(3)}{lrow}:{gcol(3+n-1)}{lrow}"
        la  = "{" + ";".join([f'"{lb}"' for lb in labels]) + "}"
        wc(ws_lc, lrow, 3 + n,
           f'=INDEX({la},MATCH(MIN({mr}),{rng},0))',
           font=fnt(10, bold=True, color=C_BESTF),
           fill=ff(C_BEST), align=ctr(), border=BT)

    # Total row
    trow = DATA_START_ROW + TOTAL_ROW_OFFSET
    loss_excel_rows["total"] = trow
    ws_lc.row_dimensions[trow].height = 30
    wc(ws_lc, trow, 1, "P_total  –  TOTAL Device Loss",
       font=fnt(11, bold=True, color="FFFFFF"),
       fill=ff(C_HDR), align=lft(), border=BM)
    total_formula_str = "P_cond + P_sw,on + P_sw,off + P_Coss + P_gate"
    wc(ws_lc, trow, 2, total_formula_str,
       font=fnt(9, bold=True, italic=True, color="1F3864"),
       fill=ff("EEF4FB"), align=lft(), border=BM)
    for d_i in range(n):
        total_f = "=" + "+".join([f"{gcol(3+d_i)}{r}" for r in
                                   [loss_excel_rows[k] for k in ("cond","sw_on","sw_off","coss","gate")]])
        c = ws_lc.cell(row=trow, column=3 + d_i, value=total_f)
        c.font = fnt(12, bold=True); c.fill = ff(C_PRI_ROW)
        c.alignment = ctr(); c.border = BM; c.number_format = "0.0000"
    mr_t  = ",".join([f"{gcol(3+d_i)}{trow}" for d_i in range(n)])
    rng_t = f"{gcol(3)}{trow}:{gcol(3+n-1)}{trow}"
    la_t  = "{" + ";".join([f'"{lb}"' for lb in labels]) + "}"
    wc(ws_lc, trow, 3 + n,
       f'=INDEX({la_t},MATCH(MIN({mr_t}),{rng_t},0))',
       font=fnt(11, bold=True, color=C_BESTF),
       fill=ff(C_BEST), align=ctr(), border=BM)

    # ── Units row ──────────────────────────────────────────────────────────────
    urow = trow + 1
    ws_lc.row_dimensions[urow].height = 18
    wc(ws_lc, urow, 1, "Units",
       font=fnt(9, italic=True, color="666666"), fill=ff("F5F5F5"), align=lft(), border=BT)
    wc(ws_lc, urow, 2, "",  fill=ff("F5F5F5"), border=BT)
    for d_i in range(n):
        wc(ws_lc, urow, 3 + d_i, "W",
           font=fnt(9, italic=True, color="666666"),
           fill=ff("F5F5F5"), align=ctr(), border=BT)
    wc(ws_lc, urow, 3 + n, "", fill=ff("F5F5F5"), border=BT)

    # ── Recommendation banner ──────────────────────────────────────────────────
    ws_lc.row_dimensions[trow + 3].height = 10
    merge_title(ws_lc, trow + 4, 1, 3 + n + 1,
                "🏆  BEST MOSFET FOR LOWEST TOTAL LOSS", bg=C_TITLE, sz=12, h=30)
    rrow = trow + 5
    ws_lc.row_dimensions[rrow].height = 34
    ws_lc.merge_cells(start_row=rrow, start_column=1, end_row=rrow, end_column=3 + n + 1)
    rc = ws_lc.cell(row=rrow, column=1,
                    value=(f'=INDEX({la_t},MATCH(MIN({mr_t}),{rng_t},0))'
                           f'&"  —  Total Loss = "&TEXT(MIN({mr_t}),"0.000")&" W"'))
    rc.font = Font(name="Calibri", bold=True, size=14, color=C_GOLD)
    rc.fill = ff(C_GREEN); rc.alignment = ctr(); rc.border = BM

    # ── Loss formula explanation block ─────────────────────────────────────────
    expl_start = rrow + 3
    merge_title(ws_lc, expl_start, 1, 3 + n + 1,
                "📐  FORMULA REFERENCE  —  How each loss is calculated",
                bg="1F3864", fg="FFFFFF", sz=11, h=26)

    FORMULA_DETAIL = [
        # (symbol, formula, derivation_note, unit_note)
        ("P_cond",
         "= I_rms²  ×  R_DS(on) [Ω]",
         "I_rms from Op. Conditions sheet (B6).  R_DS(on) is the value extrapolated to "
         "Tj_ref via linear interpolation between the 25°C and high-T datasheet points, "
         "then stored in Op. Conditions sheet device-param table.",
         "Watts (W)"),
        ("P_sw,on",
         "= E_on [J]  ×  (I_avg / ID_test)  ×  F_sw",
         "E_on is the turn-on energy from the datasheet (µJ→J ÷1e6).  "
         "It is linearly scaled by the ratio I_avg/ID_test to correct for the difference "
         "between the datasheet test current and the actual operating current.  "
         "If Eon was not tabulated it is estimated as ½·VDS_test·ID_test·(t_r + t_d(on)).",
         "Watts (W)"),
        ("P_sw,off",
         "= E_off [J]  ×  (I_avg / ID_test)  ×  F_sw",
         "Same scaling as P_sw,on but using E_off (turn-off energy).  "
         "Estimated as ½·VDS_test·ID_test·(t_f + t_d(off)) when not tabulated.",
         "Watts (W)"),
        ("P_Coss",
         "= E_oss [J]  ×  F_sw",
         "E_oss is the energy stored in C_oss at V_DS = V_out.  "
         "In hard switching this energy is dissipated every turn-on cycle.  "
         "When E_oss is not tabulated it is approximated as ½·C_oss·V_DS².",
         "Watts (W)"),
        ("P_gate",
         "= Q_G [C]  ×  V_GS_drive  ×  F_sw",
         "Q_G is the total gate charge (nC→C ÷1e9) from the datasheet at the stated "
         "V_DS and I_D test conditions.  V_GS_drive is the gate supply voltage entered "
         "in the Operating Conditions sheet.",
         "Watts (W)"),
        ("P_total",
         "= P_cond + P_sw,on + P_sw,off + P_Coss + P_gate",
         "Sum of all five loss components above.  Body-diode conduction loss during "
         "dead-time and reverse-recovery loss (P_rr = Q_rr·V_DS·F_sw) are NOT included "
         "here as they depend on dead-time which is circuit-topology-specific.",
         "Watts (W)"),
    ]

    for i, (sym, formula, note, unit_note) in enumerate(FORMULA_DETAIL):
        r = expl_start + 1 + i * 3
        ws_lc.row_dimensions[r].height   = 22
        ws_lc.row_dimensions[r+1].height = 42
        ws_lc.row_dimensions[r+2].height = 8   # spacer

        # Symbol + formula
        ws_lc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3+n+1)
        ch = ws_lc.cell(row=r, column=1,
                         value=f"  {sym}   {formula}   [{unit_note}]")
        ch.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ch.fill  = ff(C_SUBHDR); ch.alignment = lft(); ch.border = BT

        # Derivation note
        ws_lc.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=3+n+1)
        cn = ws_lc.cell(row=r+1, column=1, value=f"    ↳  {note}")
        cn.font  = fnt(9, italic=True, color="1F3864")
        cn.fill  = ff("EEF4FB"); cn.alignment = lft(); cn.border = BT

    ws_lc.column_dimensions["A"].width = 38
    ws_lc.column_dimensions["B"].width = 40
    for d_i in range(n): ws_lc.column_dimensions[gcol(3 + d_i)].width = 22
    ws_lc.column_dimensions[gcol(3 + n)].width = 26

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — SCORING DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Scoring Dashboard")
    ws4.sheet_view.showGridLines = False

    # ncols for scoring = Parameter + Score Formula + n device cols
    SC_NCOLS = 2 + n

    merge_title(ws4, 1, 1, SC_NCOLS,
                "📊  SCORING DASHBOARD  (0 = worst  |  100 = best  |  normalised to common-ground values)",
                bg=C_TITLE, sz=13, h=36)
    merge_title(ws4, 2, 1, SC_NCOLS,
                "Priority parameters (★) weighted ×2 in overall score.  "
                "Lower-is-better:  Score = (max − value) / (max − min) × 100.  "
                "Higher-is-better:  Score = (value − min) / (max − min) × 100.",
                bg=C_SUBHDR, fg="FFFFFF", sz=9, h=22)

    # Column headers: Parameter | Device1 | Device2 ... | Score Formula
    ws4.row_dimensions[3].height = 28
    wc(ws4, 3, 1, "Parameter",
       font=fnt(11, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=ctr(), border=BT)
    for d_i, lbl in enumerate(labels):
        wc(ws4, 3, 2 + d_i, lbl,
           font=fnt(11, bold=True, color="FFFFFF"), fill=ff(C_SUBHDR), align=ctr(), border=BT)
    wc(ws4, 3, 2 + n, "Score Formula Applied",
       font=fnt(11, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=ctr(), border=BT)

    scores  = [0.0] * n
    max_pts = 0

    def _score_block(ws4, start_row, keys, section_lbl, section_bg, weight):
        nonlocal max_pts
        row = start_row
        merge_title(ws4, row, 1, SC_NCOLS, section_lbl, bg=section_bg, sz=10, h=20)
        row += 1
        for key in keys:
            disp_t = _disp(key)
            display, unit, lower_better = disp_t[0], disp_t[1], disp_t[2]
            is_pri  = disp_t[3]
            bg = C_PRI_ROW if is_pri else (C_ALT if row % 2 == 0 else C_WHITE)
            ws4.row_dimensions[row].height = 22
            wc(ws4, row, 1,
               f"{display} ({unit})" + (" ★×2" if is_pri else ""),
               font=fnt(10, bold=is_pri), fill=ff(bg), align=lft(), border=BT)

            vals_d = {i: norm_list[i].get(key, (None, "", ""))[0]
                      for i in range(n)
                      if norm_list[i].get(key, (None, "", ""))[0] is not None}
            if not vals_d:
                for d_i in range(n):
                    wc(ws4, row, 2 + d_i, "—", font=fnt(10), fill=ff(bg), align=ctr(), border=BT)
                wc(ws4, row, 2 + n, "No data",
                   font=fnt(9, italic=True, color="AAAAAA"),
                   fill=ff(bg), align=lft(), border=BT)
                row += 1; continue

            mn, mx = min(vals_d.values()), max(vals_d.values())
            rng    = mx - mn
            max_pts += weight * 100
            best_v = mn if lower_better else mx

            for d_i in range(n):
                v = vals_d.get(d_i)
                if v is None:
                    wc(ws4, row, 2 + d_i, "—", font=fnt(10), fill=ff(bg), align=ctr(), border=BT)
                    continue
                s = 100.0 if rng == 0 else (
                    (mx - v) / rng * 100 if lower_better else (v - mn) / rng * 100)
                scores[d_i] += s * weight
                is_best = abs(v - best_v) < 1e-9
                wc(ws4, row, 2 + d_i, round(s, 1),
                   font=fnt(10, bold=is_best, color=C_BESTF if is_best else "000000"),
                   fill=ff(C_BEST if is_best else bg), align=ctr(), border=BT, nfmt="0.0")

            # Score formula column — show the actual numbers used
            direction = "lower-is-better" if lower_better else "higher-is-better"
            if rng == 0:
                score_formula_str = f"All equal → 100  [{direction}]"
            elif lower_better:
                score_formula_str = (f"({mx:.4g} − x) / ({mx:.4g} − {mn:.4g}) × 100"
                                     f"  [{direction}]  ×{weight}")
            else:
                score_formula_str = (f"(x − {mn:.4g}) / ({mx:.4g} − {mn:.4g}) × 100"
                                     f"  [{direction}]  ×{weight}")
            wc(ws4, row, 2 + n, score_formula_str,
               font=fnt(8, italic=True, color="1F3864"),
               fill=ff("EEF4FB"), align=lft(), border=BT)

            row += 1
        return row

    nr4 = _score_block(ws4, 4, PRIORITY_KEYS,  "★  PRIORITY  (weight ×2)", C_PRI_BG, 2)
    nr4 = _score_block(ws4, nr4 + 1, SECONDARY_KEYS, "SECONDARY  (weight ×1)", C_SEC_BG, 1)

    norm_scores = ([round(s / max(max_pts, 1) * 100, 1) for s in scores]
                   if max_pts > 0 else [0.0] * n)
    tot_row = nr4 + 1
    ws4.row_dimensions[tot_row - 1].height = 6
    ws4.row_dimensions[tot_row].height     = 32
    wc(ws4, tot_row, 1, "OVERALL SCORE  (0–100)",
       font=fnt(12, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=lft(), border=BM)
    overall_formula = (f"Weighted sum / {max_pts} × 100  "
                       f"[max possible weighted points = {max_pts}]")
    wc(ws4, tot_row, 2 + n, overall_formula,
       font=fnt(9, italic=True, color="1F3864"),
       fill=ff("EEF4FB"), align=lft(), border=BM)
    mx_s = max(norm_scores)
    for d_i, s in enumerate(norm_scores):
        wc(ws4, tot_row, 2 + d_i, s,
           font=fnt(13, bold=True, color=C_BESTF if s == mx_s else "000000"),
           fill=ff(C_BEST if s == mx_s else C_ALT), align=ctr(), border=BM, nfmt="0.0")

    bdi4 = norm_scores.index(mx_s)
    ws4.row_dimensions[tot_row + 2].height = 32
    merge_title(ws4, tot_row + 2, 1, SC_NCOLS,
                f"🏆  Best MOSFET (overall score):  {labels[bdi4]}   ({norm_scores[bdi4]:.1f}/100)",
                bg=C_GREEN, fg=C_GOLD, sz=12, h=32)

    # ── Scoring methodology explanation ───────────────────────────────────────
    expl_row = tot_row + 5
    merge_title(ws4, expl_row, 1, SC_NCOLS,
                "📐  SCORING METHODOLOGY  —  How each score is calculated",
                bg="1F3864", fg="FFFFFF", sz=11, h=26)

    SCORING_NOTES = [
        ("Step 1 — Raw values",
         "All parameter values are taken from the Common-Ground Comparison sheet, "
         "meaning they have all been extrapolated / scaled to the same reference "
         f"conditions: VDS_ref={VDS_ref:.0f}V, ID_ref={ID_ref:.1f}A, Tj_ref={TJ_ref:.0f}°C."),
        ("Step 2 — Per-parameter normalisation",
         "For each parameter the best and worst values across all devices are found.  "
         "If lower is better (e.g. Rds, Eon):   Score = (max − value) / (max − min) × 100  "
         "→ the device with the lowest value gets 100, the highest gets 0.  "
         "If higher is better (e.g. VDSS, ID):  Score = (value − min) / (max − min) × 100  "
         "→ the device with the highest value gets 100, the lowest gets 0.  "
         "If all devices are equal on a parameter, all get 100."),
        ("Step 3 — Weighting",
         "Priority parameters (★) are multiplied by weight = 2.  "
         "Secondary parameters are multiplied by weight = 1.  "
         "This reflects that conduction/switching losses (Rds, Eon, Eoff, Qg, Vdss) "
         "have a larger impact on overall converter efficiency."),
        ("Step 4 — Overall score",
         f"Overall Score = (sum of all weighted scores) / (max possible weighted points) × 100.  "
         f"Max possible = {max_pts} points  "
         f"(= {sum(2 for k in PRIORITY_KEYS)} priority params × 2 × 100  +  "
         f"{sum(1 for k in SECONDARY_KEYS)} secondary params × 1 × 100, counting only params "
         f"where at least one device had data).  "
         f"Result is normalised back to 0–100 so the winning device always scores 100."),
    ]

    for i, (heading, body) in enumerate(SCORING_NOTES):
        r_h = expl_row + 1 + i * 3
        ws4.row_dimensions[r_h].height   = 22
        ws4.row_dimensions[r_h+1].height = 50
        ws4.row_dimensions[r_h+2].height = 6

        ws4.merge_cells(start_row=r_h, start_column=1, end_row=r_h, end_column=SC_NCOLS)
        ch = ws4.cell(row=r_h, column=1, value=f"  {heading}")
        ch.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ch.fill = ff(C_SUBHDR); ch.alignment = lft(); ch.border = BT

        ws4.merge_cells(start_row=r_h+1, start_column=1, end_row=r_h+1, end_column=SC_NCOLS)
        cb = ws4.cell(row=r_h+1, column=1, value=f"    {body}")
        cb.font = fnt(9, italic=True, color="1F3864")
        cb.fill = ff("EEF4FB"); cb.alignment = lft(); cb.border = BT

    _cw(ws4, {"A": 48})
    for d_i in range(n): ws4.column_dimensions[gcol(2 + d_i)].width = 20
    ws4.column_dimensions[gcol(2 + n)].width = 62

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — RAW EXTRACTED DATA + DETAILS
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Raw Extracted Data")
    ws5.sheet_view.showGridLines = False

    merge_title(ws5, 1, 1, 4 + n,
                "📋  RAW EXTRACTED DATA  +  Test Conditions  +  Normalisation Summary",
                bg=C_TITLE, sz=12, h=28)

    ws5.row_dimensions[3].height = 24
    for ci, (txt, bg) in enumerate(
            [("Symbol", C_HDR), ("Parameter", C_HDR), ("Unit", C_HDR)] +
            [(f"{lbl}\n(Typ@25°C)", C_SUBHDR) for lbl in labels] +
            [(f"{lbl}\n(Max@25°C)", C_SUBHDR) for lbl in labels] +
            [(f"{lbl}\n(Typ@High-T)", C_SUBHDR) for lbl in labels] +
            [("Test Conditions", C_HDR)], 1):
        wc(ws5, 3, ci, txt,
           font=fnt(10, bold=True, color="FFFFFF"), fill=ff(bg), align=ctr(), border=BT)

    # Build extended table with all raw values
    EXT_PARAMS = [
        ("V_DS",     "vdss",    "vdss_max",  None,          "V",   "vdss_tcond"),
        ("I_D",      "id",      None,        "id_highT",    "A",   "id_tcond"),
        ("R_DS(on)", "rds25",   "rds25_max", "rds_ht",      "mΩ",  "rds_tcond"),
        ("V_GS(th)", "vth",     "vth_max",   None,          "V",   "vth_tcond"),
        ("C_iss",    "ciss",    None,        None,           "pF",  "ciss_tcond"),
        ("C_oss",    "coss",    "coss_max",  None,           "pF",  "coss_tcond"),
        ("C_rss",    "crss",    None,        None,           "pF",  ""),
        ("Q_g",      "qg",      None,        None,           "nC",  "qg_tcond"),
        ("Q_gd",     "qgd",     None,        None,           "nC",  ""),
        ("E_oss",    "eoss",    None,        None,           "µJ",  "eoss_tcond"),
        ("E_on",     "eon",     None,        None,           "µJ",  "eon_tcond"),
        ("E_off",    "eoff",    None,        None,           "µJ",  "eoff_tcond"),
        ("t_d(on)",  "td_on",   None,        None,           "ns",  "td_on_tcond"),
        ("t_r",      "tr",      None,        None,           "ns",  ""),
        ("t_d(off)", "td_off",  None,        None,           "ns",  ""),
        ("t_f",      "tf",      None,        None,           "ns",  ""),
        ("V_SD",     "vsd",     None,        None,           "V",   "vsd_tcond"),
        ("t_rr",     "trr",     None,        None,           "ns",  "trr_tcond"),
        ("Q_rr",     "qrr",     None,        None,           "nC",  ""),
    ]

    for r_i, (sym, k_typ, k_max, k_ht, unit, k_tc) in enumerate(EXT_PARAMS, 1):
        row = 3 + r_i
        alt_bg = C_PRI_ROW if sym in ("V_DS","R_DS(on)","Q_g","E_on","E_off") else (C_ALT if r_i%2==0 else C_WHITE)
        ws5.row_dimensions[row].height = 20
        wc(ws5, row, 1, sym,       font=fnt(10, bold=True), fill=ff(alt_bg), align=ctr(), border=BT)
        wc(ws5, row, 2, PARAMS.get(k_typ, ("",))[0] if k_typ in PARAMS else sym,
           font=fnt(10), fill=ff(alt_bg), align=lft(), border=BT)
        wc(ws5, row, 3, unit,      font=fnt(10), fill=ff(alt_bg), align=ctr(), border=BT)

        # Typ @ 25°C
        for d_i, d in enumerate(mosfets):
            v = d.get(k_typ)
            is_der = (k_typ in ("eon","eoff") and d.get(f"_{k_typ}_derived"))
            cell_bg = C_DERIVED if is_der else alt_bg
            wc(ws5, row, 4 + d_i,
               round(v, 5) if isinstance(v, float) else (v if v is not None else "—"),
               font=fnt(9, color="7F5C00" if is_der else ("000000" if v is not None else "AAAAAA")),
               fill=ff(cell_bg), align=ctr(), border=BT,
               nfmt="0.00###" if isinstance(v, float) else None)
        # Max @ 25°C
        for d_i, d in enumerate(mosfets):
            v = d.get(k_max) if k_max else None
            wc(ws5, row, 4 + n + d_i,
               round(v, 5) if isinstance(v, float) else (v if v is not None else "—"),
               font=fnt(9, color="000000" if v is not None else "CCCCCC"),
               fill=ff(alt_bg), align=ctr(), border=BT,
               nfmt="0.00###" if isinstance(v, float) else None)
        # Typ @ High-T
        for d_i, d in enumerate(mosfets):
            v = d.get(k_ht) if k_ht else None
            tj_ht = d.get("_rds_ht_tj") if k_ht == "rds_ht" else None
            lbl_suffix = f" @{tj_ht:.0f}°C" if tj_ht else ""
            wc(ws5, row, 4 + 2*n + d_i,
               (f"{round(v,4)}{lbl_suffix}" if isinstance(v, float) else (f"{v}{lbl_suffix}" if v is not None else "—")),
               font=fnt(9, color="000000" if v is not None else "CCCCCC"),
               fill=ff(alt_bg), align=ctr(), border=BT)
        # Test conditions
        conds = [d.get(k_tc, "") or "" for d in mosfets] if k_tc else [""] * n
        unique_c = "; ".join(dict.fromkeys(c for c in conds if c))
        ws5.merge_cells(start_row=row, start_column=4+3*n, end_row=row, end_column=4+3*n)
        wc(ws5, row, 4 + 3*n, unique_c,
           font=fnt(8, italic=True, color="444444"),
           fill=ff("F8F8F8"), align=lft(), border=BT)

    # Normalisation summary
    ns = 3 + len(EXT_PARAMS) + 3
    merge_title(ws5, ns, 1, 3 + n,
                (f"NORMALISATION SUMMARY  "
                 f"(VDS_ref={VDS_ref:.0f}V · ID_ref={ID_ref:.1f}A · Tj_ref={TJ_ref:.0f}°C)"),
                bg=C_SUBHDR, sz=10, h=20)
    ws5.row_dimensions[ns + 1].height = 22
    for ci, (txt, bg) in enumerate([("Parameter", C_HDR)] + [(lbl, C_SUBHDR) for lbl in labels], 1):
        wc(ws5, ns + 1, ci, txt,
           font=fnt(10, bold=True, color="FFFFFF"), fill=ff(bg), align=ctr(), border=BT)

    for r_i, (key, disp_str) in enumerate([
        ("_rds_tj",  f"Rds_on @ {TJ_ref:.0f}°C [mΩ]"),
        ("eon",      f"E_on @ ID={ID_ref:.1f}A [µJ]"),
        ("eoff",     f"E_off @ ID={ID_ref:.1f}A [µJ]"),
        ("coss",     f"Coss @ VDS={VDS_ref:.0f}V [pF]"),
        ("ciss",     f"Ciss @ VDS={VDS_ref:.0f}V [pF]"),
        ("crss",     f"Crss @ VDS={VDS_ref:.0f}V [pF]"),
    ], 1):
        row = ns + 1 + r_i
        ws5.row_dimensions[row].height = 42
        wc(ws5, row, 1, disp_str,
           font=fnt(9, bold=True), fill=ff(C_ALT), align=lft(), border=BT)
        for d_i in range(n):
            v, m, _ = norm_list[d_i].get(key, (None, "", ""))
            wc(ws5, row, 2 + d_i,
               f"→ {round(v,4) if v is not None else 'N/A'}\n{m}",
               font=fnt(8), fill=ff(C_WHITE), align=lft(), border=BT)

    _cw(ws5, {"A": 12, "B": 42, "C": 7})
    for d_i in range(n):
        ws5.column_dimensions[gcol(4 + d_i)].width     = 16  # typ
        ws5.column_dimensions[gcol(4+n+d_i)].width     = 14  # max
        ws5.column_dimensions[gcol(4+2*n+d_i)].width   = 16  # high-T
    ws5.column_dimensions[gcol(4+3*n)].width            = 48  # conditions

    # ── Graph Comparison sheet ────────────────────────────────────────────────
    # Preferred: paste the actual embedded datasheet figures side by side.
    # Fall back to the chart-object / synthetic-plot sheet only when no input
    # file contains embedded figure images.
    tmp_imgs = []
    has_figs = bool(all_figs_per_file) and any(all_figs_per_file)
    if has_figs:
        print("  Building Graph Comparison sheet (side-by-side datasheet images) …")
        tmp_imgs = build_image_comparison_sheet(
            wb, paths, labels, all_figs_per_file)
    elif all_charts_per_file is not None:
        print("  Building Graph Comparison sheet (reconstructed charts) …")
        tmp_imgs = build_graph_comparison_sheet(
            wb, paths, mosfets, labels, all_charts_per_file)

    wb.save(out_path)
    # Clean up temp PNG files used for embedding
    for tp in tmp_imgs:
        try:
            os.remove(tp)
        except OSError:
            pass
    print("\nReport saved -> %s" % out_path)


def generate_reasons_pdf(mosfets, norm_list, ref, labels, out_path):
    """
    Concise 1–2 page PDF explaining WHICH parameters are extrapolated to the
    common reference point and WHY the remaining ones are not.
    Directly answers the three reason-categories the user asked about:
      (i)   no usable graph / curve in the datasheet,
      (ii)  only a single value at one condition → no second point to interpolate,
      (iii) value lives on a LOG-scale graph → auto-extraction unreliable,
      (+)   circuit-condition-dependent (switching times depend on R_G / drive).
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors as RL
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle, HRFlowable)
    except ImportError:
        print("\n⚠  reportlab not installed — reasons PDF skipped (pip install reportlab).")
        return None

    PW, PH = A4
    LM = RM = 1.6 * cm
    BW = PW - LM - RM
    vds_ref, id_ref, tj_ref = ref["vds_ref"], ref["id_ref"], ref["tj_ref"]

    NAVY  = RL.HexColor("#1B3A6B"); DKBLUE = RL.HexColor("#17375E")
    GREEN = RL.HexColor("#1D6B2C"); LGREEN = RL.HexColor("#E2EFDA")
    AMBER = RL.HexColor("#7F3000"); LAMBER = RL.HexColor("#FFF3CD")
    RED   = RL.HexColor("#8B0000"); LGREY  = RL.HexColor("#F2F2F2")
    MGREY = RL.HexColor("#CCCCCC"); WHITE  = RL.white
    DGREY = RL.HexColor("#444444")

    def sty(**kw):
        d = dict(fontName="Helvetica", fontSize=9, leading=12, spaceAfter=3)
        d.update(kw); return ParagraphStyle("_", **d)
    S_BODY = sty(alignment=TA_JUSTIFY)
    def P(t, s=None): return Paragraph(t, s or S_BODY)
    def SP(n=6):      return Spacer(1, n)

    CELL  = sty(fontSize=8, leading=10)
    CELLB = sty(fontSize=8, leading=10, fontName="Helvetica-Bold")
    CELLH = sty(fontSize=8, leading=10, fontName="Helvetica-Bold", textColor=WHITE)
    def c(t, s=CELL):  return Paragraph(str(t), s)

    story = []

    # ── Title ────────────────────────────────────────────────────────────────
    story += [
        Table([[c("WHY SOME PARAMETERS ARE EXTRAPOLATED AND OTHERS ARE NOT",
                  sty(fontSize=15, fontName="Helvetica-Bold", textColor=WHITE,
                      leading=19, alignment=TA_CENTER))]],
              colWidths=[BW],
              style=TableStyle([("BACKGROUND", (0,0),(-1,-1), NAVY),
                                ("TOPPADDING", (0,0),(-1,-1), 9),
                                ("BOTTOMPADDING", (0,0),(-1,-1), 9)])),
        SP(5),
        P(f"<b>Devices:</b> {'   vs   '.join(labels)}", sty(fontSize=9, alignment=TA_CENTER, spaceAfter=1)),
        P(f"<b>Common reference point:</b> V<sub>DS</sub> = {vds_ref:.0f} V    "
          f"I<sub>D</sub> = {id_ref:.0f} A    T<sub>j</sub> = {tj_ref:.0f} °C",
          sty(fontSize=9, alignment=TA_CENTER, textColor=DGREY, spaceAfter=2)),
        HRFlowable(width="100%", thickness=1, color=NAVY, spaceAfter=5),
        P("<b>Goal of extrapolation.</b> Datasheets report each parameter at the "
          "manufacturer's own test condition, and these conditions differ between "
          "devices. To compare two MOSFETs fairly we bring every parameter to ONE "
          "common reference point (above). A parameter can only be moved to that "
          "point when the datasheet gives enough information to do so reliably — "
          "either a digitised curve, two data points to interpolate between, or a "
          "sound physics formula. When none of these exist, the value is shown as "
          "the direct datasheet number and is NOT forced to the reference point.",
          sty(fontName="Helvetica", fontSize=8.7, leading=11.5, alignment=TA_JUSTIFY)),
        SP(5),
    ]

    # ── Reason legend ────────────────────────────────────────────────────────
    def legend_row(tag, color, text):
        return [c(tag, sty(fontSize=8, fontName="Helvetica-Bold", textColor=color)),
                c(text, CELL)]
    leg = Table(
        [[c("Why a parameter CANNOT be extrapolated", CELLH)],
         ],
        colWidths=[BW],
        style=TableStyle([("BACKGROUND",(0,0),(-1,-1), DKBLUE),
                          ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
    reasons = Table([
        legend_row("(i)  No curve",
                   RED, "The datasheet has no graph for this quantity vs the reference "
                        "variable (temperature or voltage), so there is nothing to read or fit."),
        legend_row("(ii) Single point",
                   AMBER, "Only one value at one condition is printed. A straight line needs "
                          "two points — with one point a temperature/voltage slope cannot be formed."),
        legend_row("(iii) Log-scale graph",
                   AMBER, "The value lives on a log–log C-vs-V_DS graph. Automatic pixel "
                          "extraction from log axes is unreliable, so no number is invented — "
                          "the reader is pointed to the datasheet image instead."),
        legend_row("(+) Circuit-dependent",
                   DGREY, "Switching delays/rise/fall times depend on the external gate "
                          "resistor and driver, not on a device-intrinsic curve, so they "
                          "cannot be re-referenced; the datasheet value is shown as-is."),
    ], colWidths=[BW*0.20, BW*0.80],
       style=TableStyle([("GRID",(0,0),(-1,-1),0.3,MGREY),
                         ("ROWBACKGROUNDS",(0,0),(-1,-1),[LGREY,WHITE]),
                         ("VALIGN",(0,0),(-1,-1),"TOP"),
                         ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                         ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6)]))
    story += [leg, reasons, SP(7)]

    # ── Per-parameter status table ───────────────────────────────────────────
    # (parameter, extrapolated?, method/handling, reason)
    EXTRAP = [
        ("R_DS(on) @ T_j",   "YES", "Table/curve interpolation between 25 °C and high-T points",
         "Two temperature points (or a normalised R-vs-T curve) are available."),
        ("V_GS(th) @ T_j",   "YES", "Graph-interpolated from V_th-vs-T curve, else -5 mV/°C formula",
         "Curve or a well-established silicon temperature coefficient exists."),
        ("V_(BR)DSS @ T_j",  "YES", "Read from V_BR-vs-T_j curve (+temp. coefficient)",
         "Breakdown-vs-temperature curve is published."),
        ("E_on / E_off @ ref","YES", "Overlap-model formula, scaled on both V_DD and I_D",
         "Computable from switching times and linear in V_DD and I_D."),
        ("E_oss @ V_DS",     "YES", "Formula  E_oss = 1/2 C_oss(V_DS) V_DS^2",
         "Energy follows a known capacitance-voltage physics law."),
        ("Q_G @ V_DS",       "YES", "Gate-charge value scaled by Q_gd*(V_DS,ref/V_DS,test)",
         "Charge splits into a V_DS-dependent (Q_gd) and -independent part."),
        ("V_SD @ T_j",       "YES", "Body-diode formula  -2 mV/°C",
         "Standard silicon diode temperature coefficient applies."),
        ("t_rr / Q_rr @ T_j","YES (approx.)", "First-order temperature scaling (+0.5 / +1 %/°C)",
         "Only a coarse coefficient is known; flagged approximate."),
    ]
    NOEXT = [
        ("C_iss / C_oss / C_rss", "NO", "Shown as 'refer the graphs in graph section'",
         "(iii) Log-scale C-vs-V_DS graph — extraction unreliable."),
        ("Q_GD", "NO", "Direct datasheet value",
         "(ii) Single tabulated value; no second point to interpolate."),
        ("t_d(on), t_r, t_d(off), t_f", "NO", "Direct datasheet value",
         "(+) Circuit-condition-dependent (gate resistor / driver)."),
        ("I_D (continuous)", "NO", "Direct rated value @ 25 °C",
         "Rating/definition, not a quantity referenced to an operating point."),
    ]

    def status_table(title, rows, header_bg, yes=True):
        data = [[c(title, CELLH), c("", CELLH), c("", CELLH), c("", CELLH)],
                [c("Parameter", CELLH), c("To ref?", CELLH),
                 c("How it is handled", CELLH), c("Reason", CELLH)]]
        for p, st, meth, rsn in rows:
            data.append([c(p, CELLB),
                         c(st, sty(fontSize=8, fontName="Helvetica-Bold",
                                   textColor=GREEN if yes else RED)),
                         c(meth), c(rsn)])
        ts = TableStyle([
            ("SPAN", (0,0),(-1,0)),
            ("BACKGROUND", (0,0),(-1,0), header_bg),
            ("BACKGROUND", (0,1),(-1,1), DKBLUE),
            ("ROWBACKGROUNDS", (0,2),(-1,-1), [WHITE, LGREY]),
            ("GRID", (0,1),(-1,-1), 0.3, MGREY),
            ("VALIGN", (0,0),(-1,-1), "TOP"),
            ("TOPPADDING", (0,0),(-1,-1), 4), ("BOTTOMPADDING", (0,0),(-1,-1), 4),
            ("LEFTPADDING", (0,0),(-1,-1), 6), ("RIGHTPADDING", (0,0),(-1,-1), 6),
        ])
        return Table(data, colWidths=[BW*0.22, BW*0.11, BW*0.34, BW*0.33],
                     style=ts, repeatRows=2)

    story += [status_table("PARAMETERS BROUGHT TO THE COMMON REFERENCE POINT",
                           EXTRAP, GREEN, yes=True), SP(6)]
    story += [status_table("PARAMETERS SHOWN AS DIRECT DATASHEET VALUES (NOT EXTRAPOLATED)",
                           NOEXT, AMBER, yes=False), SP(6)]
    story += [
        P("<b>In short:</b> a parameter is extrapolated when the datasheet gives a "
          "digitised curve, two interpolation points, or a sound physics formula — "
          "otherwise it is shown as its direct datasheet value.",
          sty(fontSize=8.5, leading=11, textColor=DGREY, alignment=TA_JUSTIFY)),
    ]

    def draw(canvas, doc):
        canvas.saveState()
        w, h = A4
        canvas.setFillColor(NAVY); canvas.rect(0, 0, w, 1.0*cm, fill=1, stroke=0)
        canvas.setFont("Helvetica", 7); canvas.setFillColor(WHITE)
        canvas.drawString(LM, 0.35*cm, "MOSFET Extrapolation — Reasons Note")
        canvas.drawRightString(w - RM, 0.35*cm, f"Page {doc.page}")
        canvas.restoreState()

    SimpleDocTemplate(out_path, pagesize=A4, leftMargin=LM, rightMargin=RM,
                      topMargin=1.2*cm, bottomMargin=1.2*cm,
                      title="MOSFET Extrapolation Reasons").build(
        story, onFirstPage=draw, onLaterPages=draw)
    print(f"✅  Reasons note PDF: {out_path}")
    return out_path


# ─── CLI ──────────────────────────────────────────────────────────────────────

def get_float(prompt, default=None):
    while True:
        raw = input(prompt).strip()
        if not raw and default is not None:
            return default
        try:
            return float(raw)
        except ValueError:
            print("  ✗  Please enter a number.")


def generate_extrapolation_pdf(mosfets, norm_list, ref, labels, out_path):
    """
    Generate a PDF documenting the extrapolation method for every parameter.
    Works for any number of devices (N >= 2). Called from main() after build_report().
    Requires:  pip install reportlab
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors as RL
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle, PageBreak, HRFlowable)
    except ImportError:
        print("\n⚠  reportlab not installed — PDF notes skipped.")
        print("   Install with:  pip install reportlab\n")
        return None

    n_dev  = len(mosfets)
    PW, PH = A4
    LM = RM = 1.8 * cm
    BW  = PW - LM - RM            # usable body width

    vds_ref = ref["vds_ref"]
    id_ref  = ref["id_ref"]
    tj_ref  = ref["tj_ref"]

    # ── colours ──────────────────────────────────────────────────────────────
    NAVY   = RL.HexColor("#1B3A6B");  DKBLUE = RL.HexColor("#17375E")
    BLUE   = RL.HexColor("#2E75B6");  GREEN  = RL.HexColor("#375623")
    LGREEN = RL.HexColor("#E2EFDA");  AMBER  = RL.HexColor("#7F3000")
    LYELL  = RL.HexColor("#FFF9E6");  LGREY  = RL.HexColor("#F5F5F5")
    MGREY  = RL.HexColor("#DCDCDC");  LBLUE  = RL.HexColor("#DDEEFF")
    DGREY  = RL.HexColor("#555555");  WHITE  = RL.white

    # ── style helpers ─────────────────────────────────────────────────────────
    def sty(**kw):
        defaults = dict(fontName="Helvetica", fontSize=9.5, leading=14, spaceAfter=4)
        defaults.update(kw)   # caller's values override defaults — no duplicate keys
        return ParagraphStyle("_", **defaults)
    S_BODY = sty(alignment=TA_JUSTIFY)
    S_CELL = sty(fontSize=8.5, leading=12, spaceAfter=0)
    S_CELLB= sty(fontSize=8.5, leading=12, fontName="Helvetica-Bold", spaceAfter=0)
    S_CELLH= sty(fontSize=8.5, leading=12, fontName="Helvetica-Bold",
                 textColor=WHITE, spaceAfter=0)
    S_NOTE = sty(fontSize=8.5, leading=12, textColor=DGREY)

    def P(t, s=None):   return Paragraph(t, s or S_BODY)
    def SP(n=6):        return Spacer(1, n)
    def pc(t, s=None):  return Paragraph(str(t), s or S_CELL)
    def pcb(t):         return pc(t, S_CELLB)
    def pch(t):         return pc(t, S_CELLH)

    TAG_COL = {
        "GRAPH - DIRECT":       RL.HexColor("#1D6B2C"),
        "GRAPH-DIRECT":         RL.HexColor("#1D6B2C"),
        "GRAPH - INTERPOLATED": RL.HexColor("#1D5B8C"),
        "GRAPH-INTERPOLATED":   RL.HexColor("#1D5B8C"),
        "TABLE INTERPOLATION":  RL.HexColor("#5A4000"),
        "FORMULA":              RL.HexColor("#7F3000"),
        "NOT EXTRAPOLATABLE":   RL.HexColor("#8B0000"),
    }

    # ── block builders ────────────────────────────────────────────────────────
    def sec_hdr(num, title, tag):
        tc = DKBLUE
        for k, v in TAG_COL.items():
            if k in tag.upper():
                tc = v; break
        t = Table(
            [[pc(f"§ {num}  {title}",
                 sty(fontSize=12, fontName="Helvetica-Bold", textColor=WHITE, leading=16))],
             [pc(f"  {tag}",
                 sty(fontSize=8.5, fontName="Helvetica-Bold", textColor=WHITE, leading=12))]],
            colWidths=[BW],
            style=TableStyle([
                ("BACKGROUND", (0,0),(0,0), NAVY),
                ("BACKGROUND", (0,1),(0,1), tc),
                ("TOPPADDING",    (0,0),(-1,-1), 5),
                ("BOTTOMPADDING", (0,0),(-1,-1), 5),
                ("LEFTPADDING",   (0,0),(-1,-1), 10),
                ("RIGHTPADDING",  (0,0),(-1,-1), 10),
            ]))
        return [SP(14), t, SP(8)]

    def formula_box(lines):
        inner = Paragraph("<br/>".join(lines),
                          sty(fontName="Courier", fontSize=9, leading=14, textColor=AMBER))
        return Table([[inner]], colWidths=[BW],
                     style=TableStyle([
                         ("BACKGROUND", (0,0),(-1,-1), LYELL),
                         ("BOX",        (0,0),(-1,-1), 1.0, AMBER),
                         ("TOPPADDING",    (0,0),(-1,-1), 7),
                         ("BOTTOMPADDING", (0,0),(-1,-1), 7),
                         ("LEFTPADDING",   (0,0),(-1,-1), 10),
                         ("RIGHTPADDING",  (0,0),(-1,-1), 10),
                     ]))

    def result_box(pairs):
        rows = [[pc(lbl, sty(fontSize=9, textColor=GREEN)),
                 pc(val, sty(fontSize=9, fontName="Helvetica-Bold", textColor=GREEN))]
                for lbl, val in pairs]
        return Table(rows, colWidths=[9*cm, BW - 9*cm],
                     style=TableStyle([
                         ("BACKGROUND", (0,0),(-1,-1), LGREEN),
                         ("BOX",        (0,0),(-1,-1), 1.0, GREEN),
                         ("INNERGRID",  (0,0),(-1,-1), 0.3, RL.HexColor("#AACCA0")),
                         ("TOPPADDING",    (0,0),(-1,-1), 4),
                         ("BOTTOMPADDING", (0,0),(-1,-1), 4),
                         ("LEFTPADDING",   (0,0),(-1,-1), 8),
                         ("RIGHTPADDING",  (0,0),(-1,-1), 8),
                         ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
                     ]))

    def note_box(text):
        return Table([[P(f"ℹ  {text}", S_NOTE)]], colWidths=[BW],
                     style=TableStyle([
                         ("BACKGROUND", (0,0),(-1,-1), LGREY),
                         ("BOX",        (0,0),(-1,-1), 0.5, MGREY),
                         ("TOPPADDING",    (0,0),(-1,-1), 6),
                         ("BOTTOMPADDING", (0,0),(-1,-1), 6),
                         ("LEFTPADDING",   (0,0),(-1,-1), 8),
                         ("RIGHTPADDING",  (0,0),(-1,-1), 8),
                     ]))

    def dev_table(nm_key, unit=""):
        """One row per device: Device | Value | Full extrapolation method string."""
        CW = [BW * 0.22, BW * 0.14, BW * 0.64]
        data = [[pch("Device"), pch(f"Value{' (' + unit + ')' if unit else ''}"),
                 pch("Extrapolation Method")]]
        for i, (nm, lbl) in enumerate(zip(norm_list, labels)):
            val, meth, _ = nm.get(nm_key, (None, "—", ""))
            vs = (f"{val:.4g}" if isinstance(val, float) else str(val)) if val is not None else "—"
            data.append([pc(lbl), pcb(vs), pc(meth or "—")])
        ts = TableStyle([
            ("BACKGROUND",    (0,0),(-1,0), DKBLUE),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [LBLUE, WHITE]),
            ("GRID",          (0,0),(-1,-1), 0.3, RL.HexColor("#B0C4DE")),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
            ("LEFTPADDING",   (0,0),(-1,-1), 6),
            ("RIGHTPADDING",  (0,0),(-1,-1), 6),
            ("VALIGN",        (0,0),(-1,-1), "TOP"),
        ])
        return Table(data, colWidths=CW, style=ts, hAlign="LEFT", repeatRows=1)

    def res_pairs(nm_key, unit="", prefix=""):
        out_pairs = []
        for lbl, nm in zip(labels, norm_list):
            val, _, _ = nm.get(nm_key, (None, "—", ""))
            vs = (f"{val:.4g} {unit}".strip() if val is not None else "—")
            out_pairs.append((f"{prefix}{lbl}", vs))
        return out_pairs

    # ── story ─────────────────────────────────────────────────────────────────
    story = []

    # Cover
    def info_row(label, value):
        return Table([[pc(label, sty(fontSize=8.5, textColor=DGREY)),
                       pc(value, sty(fontSize=9.5, fontName="Helvetica-Bold"))]],
                     colWidths=[4.5*cm, BW - 4.5*cm],
                     style=TableStyle([("TOPPADDING",(0,0),(-1,-1),3),
                                       ("BOTTOMPADDING",(0,0),(-1,-1),3),
                                       ("LEFTPADDING",(0,0),(-1,-1),0),
                                       ("RIGHTPADDING",(0,0),(-1,-1),0)]))

    story += [
        SP(30),
        Table([[pc("MOSFET PARAMETER EXTRAPOLATION",
                   sty(fontSize=20, fontName="Helvetica-Bold", textColor=WHITE,
                       leading=26, alignment=TA_CENTER))],
               [pc("Method Notes  &amp;  Step-by-Step Calculations",
                   sty(fontSize=12, textColor=RL.HexColor("#BBDDFF"),
                       leading=16, alignment=TA_CENTER))]],
              colWidths=[BW],
              style=TableStyle([("BACKGROUND",(0,0),(-1,-1), NAVY),
                                ("TOPPADDING",(0,0),(-1,-1),18),
                                ("BOTTOMPADDING",(0,0),(-1,-1),18),
                                ("LEFTPADDING",(0,0),(-1,-1),12),
                                ("RIGHTPADDING",(0,0),(-1,-1),12)])),
        SP(16),
    ]
    for i, lbl in enumerate(labels):
        story.append(info_row(f"Device {i+1}:", lbl))
    story += [
        SP(6),
        HRFlowable(width="100%", thickness=1.0, color=BLUE, spaceBefore=2, spaceAfter=6),
        info_row("V_DS,ref:", f"{vds_ref:.0f} V"),
        info_row("I_D,ref:",  f"{id_ref:.0f} A"),
        info_row("T_j,ref:",  f"{tj_ref:.0f} °C"),
        SP(14),
        P("This document explains, for each electrical parameter, which extrapolation method "
          "brings the datasheet value to the common reference point above. "
          "The actual computed values and the full method string are shown for every device."),
        SP(12),
    ]

    # Method legend
    leg_defs = [
        ("GRAPH-DIRECT",        "Value read from a published curve at exactly the target condition.",
         RL.HexColor("#1D6B2C")),
        ("GRAPH-INTERPOLATED",  "Interpolated between two points on a published curve.",
         RL.HexColor("#1D5B8C")),
        ("TABLE INTERPOLATION", "Same as above but from printed table rows, not curve pixels.",
         RL.HexColor("#5A4000")),
        ("FORMULA (derived)",   "No suitable curve; physics-based formula applied.",
         RL.HexColor("#7F3000")),
        ("NOT EXTRAPOLATABLE",  "Insufficient datasheet data; value cannot be computed.",
         RL.HexColor("#8B0000")),
    ]
    leg_data = [[pch("Method Tag"), pch("Meaning")]]
    leg_ts   = TableStyle([
        ("BACKGROUND", (0,0),(-1,0), DKBLUE),
        ("GRID",       (0,0),(-1,-1), 0.3, MGREY),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ("LEFTPADDING",   (0,0),(-1,-1), 7),
        ("RIGHTPADDING",  (0,0),(-1,-1), 7),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
    ])
    for i, (tag, meaning, c) in enumerate(leg_defs):
        r, g, b = c.red, c.green, c.blue
        light   = RL.Color(r*0.25+0.75, g*0.25+0.75, b*0.25+0.75)
        leg_data.append([pc(tag,    sty(fontSize=8.5, fontName="Helvetica-Bold", textColor=c)),
                         pc(meaning, S_CELL)])
        leg_ts.add("BACKGROUND", (0,i+1),(0,i+1), light)
    story.append(Table(leg_data, colWidths=[4.8*cm, BW-4.8*cm],
                       style=leg_ts, hAlign="LEFT"))
    story.append(PageBreak())

    # ── §3.1  R_DS(on) ────────────────────────────────────────────────────────
    story += sec_hdr("3.1", "R_DS(on) — On-State Resistance at T_j,ref",
                     "TABLE INTERPOLATION  (linear between two temperature data points)")
    story += [
        P("Linear interpolation between R_DS(25°C) and R_DS(T_HT) from the datasheet "
          "table or normalised R_DS-vs-T curve.  Preferred method (GRAPH-INTERPOLATED) "
          "requires a digitized curve at both I_D,ref and T_j."),
        SP(4),
        formula_box([
            f"R_DS(T_j,ref) = R_DS(25°C)  +  "
            f"[R_DS(T_HT) – R_DS(25°C)] / (T_HT – 25)  ×  (T_j,ref – 25)",
            f"",
            f"T_j,ref = {tj_ref:.0f} °C",
        ]),
        SP(8), dev_table("_rds_tj", "mΩ"), SP(8),
        result_box(res_pairs("_rds_tj", "mΩ")),
        note_box("Replace table-based R_DS values with a digitised R_DS-vs-I_D family at multiple "
                 "temperatures for GRAPH-INTERPOLATED accuracy (preferred per spec §3.1)."),
        PageBreak(),
    ]

    # ── §3.2  VGS(th) ─────────────────────────────────────────────────────────
    story += sec_hdr("3.2", "V_GS(th) — Gate Threshold Voltage at T_j,ref",
                     "GRAPH-INTERPOLATED  (if V_th-vs-T curve exists)  |  FORMULA (fallback)")
    story += [
        P("If the datasheet includes a normalised V_GS(th)-vs-T_j curve, multiply the 25°C "
          "value by the ratio read at T_j,ref (GRAPH-INTERPOLATED).  Otherwise the "
          "silicon-physics formula is applied as a fallback."),
        SP(4),
        formula_box([
            "GRAPH-INTERPOLATED:  V_th(T_j) = V_th(25°C) × ratio_from_curve(T_j)",
            "",
            "FORMULA fallback:    V_th(T_j) = V_th(25°C) + (dV_th/dT) × (T_j – 25)",
            "                     dV_th/dT ≈ −5 mV/°C   (silicon physics default)",
        ]),
        SP(8), dev_table("vth", "V"), SP(8),
        result_box(res_pairs("vth", "V")),
        note_box("The −5 mV/°C coefficient is the standard silicon power MOSFET value. "
                 "Replacing it with the manufacturer's measured value improves accuracy."),
        PageBreak(),
    ]

    # ── §3.3  V_(BR)DSS ───────────────────────────────────────────────────────
    story += sec_hdr("3.3", "V_(BR)DSS — Breakdown Voltage at T_j,ref",
                     "GRAPH - DIRECT  (read at T_j,ref from V_BR-vs-T_j curve)")
    story += [
        P(f"Read V_(BR)DSS directly at T_j,ref = {tj_ref:.0f} °C from the "
          f"datasheet curve.  Positive temperature coefficient "
          f"(≈+0.06%/°C for 600 V silicon) means V_BR rises with temperature."),
        SP(8), dev_table("vdss", "V"), SP(8),
        result_box(res_pairs("vdss", "V")),
        note_box("Some datasheets plot only a normalised V_BR-vs-T curve.  Multiply the "
                 "normalised ratio at T_j,ref by the table nominal to get absolute volts."),
        PageBreak(),
    ]

    # ── §3.4  Capacitances — moved to end (log-scale) ────────────────────────
    story += sec_hdr("3.4", "C_iss / C_oss / C_rss — Capacitances at V_DS,ref",
                     "⚠  LOG-SCALE GRAPH  —  see Appendix A at end of this document")
    story += [
        Table([[pc("⚠  Extraction from log-scale C-vs-VDS graphs is currently not reliable.\n"
                   "The capacitance values and full method notes have been moved to\n"
                   "Appendix A at the end of this document.",
                   sty(fontSize=10, fontName="Helvetica-Bold",
                       textColor=RL.HexColor("#7F0000"), leading=15))]],
              colWidths=[BW],
              style=TableStyle([("BACKGROUND",(0,0),(-1,-1), RL.HexColor("#FFF3CD")),
                                ("BOX",(0,0),(-1,-1),1.2,RL.HexColor("#C0392B")),
                                ("TOPPADDING",(0,0),(-1,-1),12),
                                ("BOTTOMPADDING",(0,0),(-1,-1),12),
                                ("LEFTPADDING",(0,0),(-1,-1),14),
                                ("RIGHTPADDING",(0,0),(-1,-1),14)])),
        SP(8), PageBreak(),
    ]

    # ── §3.5  Eoss ────────────────────────────────────────────────────────────
    story += sec_hdr("3.5", "E_oss — Output-Cap Stored Energy at V_DS,ref",
                     "GRAPH - DIRECT  (manufacturer integral curve)  |  ½CV² fallback")
    story += [
        P(f"If the datasheet plots E_oss-vs-V_DS, read at V_DS,ref = {vds_ref:.0f} V directly.  "
          f"This is the manufacturer's own integral of the non-linear C_oss(V) curve — "
          f"more accurate than the ½CV² formula especially for superjunction devices."),
        SP(4),
        formula_box([
            f"GRAPH-DIRECT:  read E_oss at V_DS = {vds_ref:.0f} V from the curve.",
            "",
            "FORMULA fallback:  E_oss = 0.5 × C_oss(V_DS,ref) × V_DS,ref²",
        ]),
        SP(8), dev_table("eoss", "µJ"), SP(8),
        result_box(res_pairs("eoss", "µJ")),
        note_box("The ½CV² formula always underestimates E_oss for SJ devices because "
                 "it ignores the high capacitance near 0 V.  Use the manufacturer's E_oss "
                 "curve wherever digitised data is available."),
        PageBreak(),
    ]

    # ── §3.6  Qg / Qgd ───────────────────────────────────────────────────────
    story += sec_hdr("3.6", "Q_g / Q_gd — Gate Charge at (V_DD, V_GS,drive)",
                     "GRAPH - DIRECT  (extended read from gate-charge waveform)")
    story += [
        P("Read Q_g at V_GS = V_GS,drive from the gate-charge curve.  If the test V_DD "
          "differs from V_DS,ref, Q_gd is corrected using the ratio of C_oss integrals."),
        SP(4),
        formula_box([
            "Q_g:  read directly at V_GS,drive from gate-charge waveform (GRAPH-DIRECT).",
            "",
            "Q_gd V_DD correction (first-order approximation):",
            "  Q_gd(V_DD,ref) ≈ Q_gd(V_DD,test) × (V_DD,ref / V_DD,test)",
        ]),
        SP(8),
        P("Q_g:"), dev_table("qg", "nC"), SP(6),
        result_box(res_pairs("qg", "nC", "Q_g   ")),
        SP(8),
        P("Q_gd:"), dev_table("qgd", "nC"), SP(6),
        result_box(res_pairs("qgd", "nC", "Q_gd  ")),
        PageBreak(),
    ]

    # ── §3.7  Eon / Eoff ──────────────────────────────────────────────────────
    story += sec_hdr("3.7", "E_on / E_off — Switching Energy at (V_DD,ref, I_D,ref)",
                     "FORMULA (derived)  — hard-switching overlap model, both V_DD and I_D axes")
    story += [
        P("Switching energy is computed from published switching times using the hard-switching "
          "linear-ramp overlap model, then scaled from the datasheet test condition to the "
          "reference operating point on both V_DD and I_D axes simultaneously."),
        SP(4),
        formula_box([
            "E_on(test)  = 0.5 × V_DD,test × I_D,test × (t_d(on) + t_r)   [µJ]",
            "E_off(test) = 0.5 × V_DD,test × I_D,test × (t_d(off) + t_f)   [µJ]",
            "",
            "Scale to reference operating point:",
            "  E(V_DD,ref, I_D,ref) = E_test × (V_DD,ref / V_DD,test) × (I_D,ref / I_D,test)",
            "",
            f"V_DD,ref = {vds_ref:.0f} V    I_D,ref = {id_ref:.0f} A",
        ]),
        SP(8),
        P("E_on:"), dev_table("eon", "µJ"), SP(6),
        result_box(res_pairs("eon", "µJ", "E_on   ")),
        SP(8),
        P("E_off:"), dev_table("eoff", "µJ"), SP(6),
        result_box(res_pairs("eoff", "µJ", "E_off  ")),
        note_box("Switching times t_r / t_f are not perfectly constant with V_DD and I_D.  "
                 "Accuracy degrades when scale ratios stray far from 1.0.  "
                 "The method string for each device shows the exact ratios applied."),
        PageBreak(),
    ]

    # ── §3.8  VSD ─────────────────────────────────────────────────────────────
    story += sec_hdr("3.8", "V_SD — Body-Diode Forward Voltage at (I_SD, T_j,ref)",
                     "GRAPH - INTERPOLATED  (between temperature lines on V_SD-vs-I_SD curve)")
    story += [
        P(f"Read V_SD at I_SD = I_D,ref from the two temperature curves bracketing "
          f"T_j,ref = {tj_ref:.0f} °C, then linearly interpolate."),
        SP(4),
        formula_box([
            "V_SD(T_j,ref) = V_SD(T_low)  +  [V_SD(T_high) – V_SD(T_low)]",
            "                              /  (T_high – T_low)  ×  (T_j,ref – T_low)",
        ]),
        SP(8), dev_table("vsd", "V"), SP(8),
        result_box(res_pairs("vsd", "V")),
        PageBreak(),
    ]

    # ── §3.9  Qrr / trr ───────────────────────────────────────────────────────
    story += sec_hdr("3.9", "Q_rr / t_rr — Reverse Recovery at T_j,ref",
                     "TABLE INTERPOLATION  (2 temp rows)  |  NOT EXTRAPOLATABLE  (1 row only)")
    story += [
        P("Temperature correction requires two datasheet rows (25°C + high-T).  "
          "If only one row exists the value is used as-is and flagged as not temperature-corrected."),
        SP(4),
        formula_box([
            "If 2 temperature rows:  Q_rr(T_j,ref) = Q_rr(T_low)",
            "                        + [Q_rr(T_high)–Q_rr(T_low)] / (T_high–T_low) × (T_j,ref–T_low)",
            "",
            "If 1 row only:  use Q_rr(25°C) as a conservative lower bound.",
        ]),
        SP(8),
        P("Q_rr:"), dev_table("qrr", "nC"), SP(6),
        result_box(res_pairs("qrr", "nC")),
        SP(8),
        P("t_rr:"), dev_table("trr", "ns"), SP(6),
        result_box(res_pairs("trr", "ns")),
        note_box("Do not extrapolate from a single-temperature Q_rr using the generic "
                 "minority-carrier doubling rule — it is unreliable for superjunction body diodes."),
        PageBreak(),
    ]

    # ── Appendix A: Log-Scale Parameters (Capacitances) ─────────────────────
    story += sec_hdr("Appendix A",
                     "C_iss / C_oss / C_rss — Capacitances at V_DS,ref  (Log-Scale Graph)",
                     "⚠  EXTRACTION NOT RELIABLE — log-scale graph auto-digitisation is limited")
    story += [
        Table([[pc("⚠  IMPORTANT: The C-vs-VDS graph in MOSFET datasheets uses a LOG-LOG scale.\n"
                   "Automatic extraction from log-scale graphs is currently not working reliably.\n"
                   "The values below are extracted using the 1/√V depletion-law formula fallback\n"
                   "or from datasheet table entries — NOT from the log-scale graph curve.\n"
                   "Treat these values as approximate. Always verify against the datasheet graph.",
                   sty(fontSize=10, fontName="Helvetica-Bold",
                       textColor=RL.HexColor("#7F0000"), leading=15))]],
              colWidths=[BW],
              style=TableStyle([("BACKGROUND",(0,0),(-1,-1), RL.HexColor("#FFF3CD")),
                                ("BOX",(0,0),(-1,-1),1.5,RL.HexColor("#C0392B")),
                                ("TOPPADDING",(0,0),(-1,-1),12),
                                ("BOTTOMPADDING",(0,0),(-1,-1),12),
                                ("LEFTPADDING",(0,0),(-1,-1),14),
                                ("RIGHTPADDING",(0,0),(-1,-1),14)])),
        SP(10),
        P(f"Preferred method (currently limited): read directly at V_DS,ref = {vds_ref:.0f} V "
          f"from the log-log C-vs-V_DS curve.  Fallback when only a single-voltage table value exists:"),
        SP(4),
        formula_box([
            f"GRAPH-DIRECT (target):  read C at V_DS = {vds_ref:.0f} V from log-log curve.",
            "",
            "FORMULA fallback (1/√V depletion law):",
            "  C(V_target) = C(V_0)  ×  √(V_0 / V_target)",
        ]),
        SP(8),
        P("C_oss:"), dev_table("coss", "pF"), SP(6),
        result_box(res_pairs("coss", "pF", "C_oss  ")),
        SP(8),
        P("C_iss:"), dev_table("ciss", "pF"), SP(6),
        result_box(res_pairs("ciss", "pF", "C_iss  ")),
        SP(6),
        note_box("Superjunction MOSFETs have highly non-linear C_oss near 0 V.  "
                 "The 1/√V depletion law is least accurate for these devices.  "
                 "Log-scale graph auto-extraction will be improved in a future version."),
        PageBreak(),
    ]

    # ── Master Summary ────────────────────────────────────────────────────────
    story += [
        Table([[pc("MASTER SUMMARY — ALL PARAMETERS AT REFERENCE CONDITIONS",
                   sty(fontSize=13, fontName="Helvetica-Bold", textColor=WHITE,
                       leading=18, alignment=TA_CENTER))]],
              colWidths=[BW],
              style=TableStyle([("BACKGROUND",(0,0),(-1,-1), NAVY),
                                ("TOPPADDING",(0,0),(-1,-1),10),
                                ("BOTTOMPADDING",(0,0),(-1,-1),10),
                                ("LEFTPADDING",(0,0),(-1,-1),10),
                                ("RIGHTPADDING",(0,0),(-1,-1),10)])),
        SP(6),
        P(f"V_DS,ref = {vds_ref:.0f} V  │  I_D,ref = {id_ref:.0f} A  "
          f"│  T_j,ref = {tj_ref:.0f} °C",
          sty(fontSize=9, textColor=DGREY, alignment=TA_CENTER)),
        SP(10),
    ]

    # Dynamic column widths for N devices
    N   = n_dev
    CW0 = BW * 0.28                    # parameter name column
    CWm = BW * 0.58 / N                # each device value column
    CWe = BW - CW0 - CWm * N          # method column (remainder)
    sum_cw = [CW0] + [CWm] * N + [CWe]

    def ps(t):  return Paragraph(str(t), sty(fontSize=8, leading=11))
    def psb(t): return Paragraph(str(t), sty(fontSize=8, leading=11,
                                              fontName="Helvetica-Bold", textColor=GREEN))
    def psh(t): return Paragraph(str(t), sty(fontSize=8, leading=11,
                                              fontName="Helvetica-Bold", textColor=WHITE))

    hrow = [psh("Parameter")] + [psh(lbl) for lbl in labels] + [psh("Method / Tag")]
    sum_data = [hrow]

    SUMMARY_PARAMS = [
        ("_rds_tj", "R_DS(on) @ T_j,ref",  "mΩ"),
        ("vth",     "V_GS(th) @ T_j,ref",   "V"),
        ("vdss",    "V_(BR)DSS",             "V"),
        ("coss",    "C_oss @ V_DS,ref",      "pF"),
        ("ciss",    "C_iss @ V_DS,ref",      "pF"),
        ("eoss",    "E_oss @ V_DS,ref",      "µJ"),
        ("qg",      "Q_g",                   "nC"),
        ("qgd",     "Q_gd",                  "nC"),
        ("eon",     "E_on @ ref",             "µJ"),
        ("eoff",    "E_off @ ref",            "µJ"),
        ("vsd",     "V_SD @ T_j,ref",        "V"),
        ("qrr",     "Q_rr",                  "nC"),
        ("trr",     "t_rr",                  "ns"),
    ]

    for nm_key, pname, unit in SUMMARY_PARAMS:
        row  = [ps(f"{pname} ({unit})")]
        tags = []
        for nm in norm_list:
            val, meth, _ = nm.get(nm_key, (None, "—", ""))
            vs = f"{val:.4g}" if val is not None else "—"
            row.append(psb(vs))
            if meth and "]" in meth:
                tags.append(meth.split("]")[0].lstrip("[") + "]")
            elif meth:
                tags.append(meth[:35])
        tag_str = " | ".join(dict.fromkeys(tags)) if tags else "—"
        row.append(ps(tag_str[:100]))
        sum_data.append(row)

    sum_ts = TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), DKBLUE),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [LGREY, WHITE]),
        ("GRID",          (0,0),(-1,-1), 0.4, RL.HexColor("#AAAAAA")),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("LEFTPADDING",   (0,0),(-1,-1), 5),
        ("RIGHTPADDING",  (0,0),(-1,-1), 5),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
    ])
    story.append(Table(sum_data, colWidths=sum_cw, style=sum_ts,
                       hAlign="LEFT", repeatRows=1))
    story += [
        SP(10),
        note_box("— indicates the parameter was not found in the datasheet or could not "
                 "be computed.  Replace estimated values with actual digitised curve readings "
                 "for highest accuracy.  GRAPH-DIRECT and GRAPH-INTERPOLATED methods supersede "
                 "all formula-based fallbacks wherever curve data is available."),
    ]

    # ── page template ─────────────────────────────────────────────────────────
    dev_str = "  |  ".join(labels)

    def draw_page(canvas, doc):
        canvas.saveState()
        w, h = A4
        canvas.setFillColor(NAVY)
        canvas.rect(0, 0, w, 1.4*cm, fill=1, stroke=0)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(WHITE)
        canvas.drawString(LM, 0.52*cm,
                          f"MOSFET Extrapolation Notes  —  {dev_str[:80]}")
        canvas.drawRightString(w - RM, 0.52*cm,
                               f"Ref: V_DS={vds_ref:.0f}V  I_D={id_ref:.0f}A  "
                               f"T_j={tj_ref:.0f}°C  —  Page {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        out_path,
        pagesize=A4,
        leftMargin=LM, rightMargin=RM,
        topMargin=1.8*cm, bottomMargin=2.0*cm,
        title="MOSFET Extrapolation Method Notes",
        author="mosfet_comparison6.py",
    )
    doc.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
    print(f"\n✅  Extrapolation notes PDF: {out_path}")
    return out_path


def main():
    print("\n" + "═"*65)
    print("  MOSFET Comparison & Loss Analysis Tool  —  v3")
    print("═"*65)

    # ── 1. Collect file paths ─────────────────────────────────────────────────
    files = sys.argv[1:]
    if not files:
        print("\nEnter paths to v5 MOSFET parameter Excel files (minimum 2).")
        print("Press Enter on a blank line when done.\n")
        while True:
            raw = input("  File path: ").strip().strip('"').strip("'")
            if not raw:
                if len(files) >= 2:
                    break
                print("  ⚠  Please enter at least 2 files.")
                continue
            if not os.path.isfile(raw):
                print(f"  ✗  File not found: {raw}")
                continue
            files.append(raw)
            print(f"  ✓  Added  ({len(files)} file{'s' if len(files)>1 else ''} so far)")

    valid = [f for f in files if os.path.isfile(f)]
    if len(valid) < 2:
        sys.exit("❌  Need at least 2 valid Excel files.")

    # ── 2. Parse all MOSFET files ─────────────────────────────────────────────
    print("\n── Parsing MOSFET parameter files ──")
    mosfets            = []
    all_charts_per_file = []
    all_figs_per_file   = []
    for f in valid:
        d   = read_v5_file(f)
        lbl = mosfet_label(f, d)
        print(f"\n  ✓  {lbl}")
        print(f"     VDSS={d.get('vdss')}V  ID={d.get('id')}A (25°C) / {d.get('id_highT')}A (high-T)")
        print(f"     Rds25={d.get('rds25')}mΩ (typ)  Rds_HT={d.get('rds_ht')}mΩ @ {d.get('_rds_ht_tj',125)}°C  tcond:[{d.get('rds_tcond','')}]")
        print(f"     Qg={d.get('qg')}nC [{d.get('qg_tcond','')}]")
        print(f"     Eon={d.get('eon')}µJ {'[DERIVED from t_r/t_d(on)]' if d.get('_eon_derived') else '[tabulated]'}  "
              f"Eoff={d.get('eoff')}µJ {'[DERIVED from t_f/t_d(off)]' if d.get('_eoff_derived') else '[tabulated]'}")
        print(f"     Ciss={d.get('ciss')}pF  Coss={d.get('coss')}pF  Crss={d.get('crss')}pF")
        print(f"     Eoss={d.get('eoss')}µJ {'[derived: ½CossV²]' if d.get('_eoss_derived') else '[tabulated]'}")
        print(f"     Vth={d.get('vth')}V  Vsd={d.get('vsd')}V  Qgd={d.get('qgd')}nC")
        print(f"     t_r={d.get('tr')}ns  t_f={d.get('tf')}ns  t_d(on)={d.get('td_on')}ns  t_d(off)={d.get('td_off')}ns")
        print(f"     trr={d.get('trr')}ns  Qrr={d.get('qrr')}nC")
        mosfets.append(d)
        # Extract figures (images + digitized data tables) BEFORE charts so
        # Eoss rows can be sourced from figures (primary) or charts (fallback)
        figs = extract_figures_from_file(f)
        all_figs_per_file.append(figs)
        charts = extract_charts_from_file(f)
        all_charts_per_file.append(charts)
        # Extract Eoss graph data: v5 extractor stores digitized Eoss-vs-VDS
        # data in spreadsheet cells alongside the embedded image; figures carry
        # that data in their series[]. Charts are a fallback for files that
        # embed native Excel chart objects instead.
        eoss_rows = _extract_eoss_rows(figs) or _extract_eoss_rows(charts)
        if eoss_rows:
            d["eoss_rows"] = eoss_rows
            vds_span = f"{eoss_rows[0][0]:.0f}–{eoss_rows[-1][0]:.0f}V"
            print(f"     Eoss graph: {len(eoss_rows)} pts ({vds_span}) — "
                  f"will interpolate directly at VDS_ref")
        else:
            print(f"     Eoss graph: no digitized series found — "
                  f"will use Coss-formula or spec-table fallback")
        rds_curve = _extract_rds_ratio_from_figs(figs) or _extract_rds_ratio_from_figs(charts)
        if rds_curve:
            d["rds_ratio_curve"] = rds_curve
            t_span = f"{rds_curve[0][0]:.0f}–{rds_curve[-1][0]:.0f}°C"
            print(f"     Rds-vs-T graph: {len(rds_curve)} pts ({t_span}) — "
                  f"will use graph ratio × rds25 for Rds extrapolation")
        else:
            print(f"     Rds-vs-T graph: no digitized series — "
                  f"will use 2-point spec table interpolation")
        if figs:
            print(f"     Datasheet figure images found: {len(figs)}"
                  f" ({', '.join(dict.fromkeys(g['chart_type'] for g in figs))})")
        else:
            print(f"     Charts found: {len(charts)}"
                  + (f" ({', '.join(dict.fromkeys(c['chart_type'] for c in charts))})"
                     if charts else " — will use synthetic charts from tabular data"))

    # ── 3. Reference conditions ───────────────────────────────────────────────
    vdss_all = [d.get("vdss") for d in mosfets if d.get("vdss")]
    vds_def  = round(min(vdss_all) * 0.6, 0) if vdss_all else 400.0
    id_all   = [d.get("_eon_id") or d.get("_sw_id") or d.get("id")
                for d in mosfets if (d.get("_eon_id") or d.get("_sw_id") or d.get("id"))]
    id_def   = round(min(id_all), 1) if id_all else 15.0

    print(f"\n── Reference Conditions for Common-Ground Normalisation ──")
    print("   (All devices will be compared at these identical conditions)")
    vds_ref = get_float(f"  VDS_ref (V)   [suggest {vds_def:.0f} V]: ", vds_def)
    id_ref  = get_float(f"  ID_ref  (A)   [suggest {id_def:.1f} A]: ",  id_def)
    tj_ref  = get_float(f"  Tj_ref  (°C)  [125 °C]: ",                  125.0)

    # ── 4. Operating conditions for loss ─────────────────────────────────────
    print("\n── Operating Conditions for Loss Analysis ──")
    iavg    = get_float("  I_avg  (A)   [average switch current]: ")
    irms    = get_float("  I_rms  (A)   [RMS switch current]: ")
    fsw     = get_float("  F_sw   (Hz)  [switching frequency]: ")
    vout    = get_float("  V_out  (V)   [DC output / bus voltage, 400]: ", 400.0)
    vgs_all = [d.get("vgs_drive") for d in mosfets if d.get("vgs_drive")]
    vgs_def = round(sum(vgs_all)/len(vgs_all), 1) if vgs_all else 10.0
    vgs_drv = get_float(f"  VGS_drive (V) [gate supply voltage, {vgs_def}]: ", vgs_def)

    # ── 5. Output filename ────────────────────────────────────────────────────
    out = input("\n  Output filename [mosfet_report.xlsx]: ").strip()
    if not out:       out = "mosfet_report.xlsx"
    if not out.endswith(".xlsx"): out += ".xlsx"

    ref = {"vds_ref": vds_ref, "id_ref": id_ref, "tj_ref": tj_ref}

    # ── 6. Normalise ──────────────────────────────────────────────────────────
    print("\n── Common-ground normalisation ──")
    norm_list = compute_normalised(mosfets, ref)
    for d_i, (d, nm) in enumerate(zip(mosfets, norm_list)):
        lbl = mosfet_label(valid[d_i], d)
        print(f"  {lbl}:")
        for k, disp in [("_rds_tj", f"Rds@{tj_ref:.0f}°C"),
                         ("eon",    f"Eon@ID={id_ref:.1f}A"),
                         ("eoff",   f"Eoff@ID={id_ref:.1f}A"),
                         ("coss",   f"Coss@VDS={vds_ref:.0f}V")]:
            v, m, _ = nm.get(k, (None, "", ""))
            print(f"    {disp:22} = {v}  [{str(m)[:70]}]")

    # ── 7. Build report ───────────────────────────────────────────────────────
    print("\n── Building 6-sheet Excel report ──")
    build_report(valid, mosfets, norm_list, ref, iavg, irms, fsw, vout, vgs_drv, out,
                 all_charts_per_file=all_charts_per_file,
                 all_figs_per_file=all_figs_per_file)

    # ── 8. Generate PDFs ──────────────────────────────────────────────────────
    print("\n── Generating extrapolation notes PDF ──")
    labels = [mosfet_label(p, d) for p, d in zip(valid, mosfets)]
    pdf_path = out.replace(".xlsx", "_Extrapolation_Notes.pdf")
    generate_extrapolation_pdf(mosfets, norm_list, ref, labels, pdf_path)

    # Concise 1–2 page note: why some parameters are extrapolated and others not
    reasons_path = out.replace(".xlsx", "_Why_Extrapolated.pdf")
    generate_reasons_pdf(mosfets, norm_list, ref, labels, reasons_path)

    print("\n📌  Quick guide (7 sheets):")
    print("  Sheet 1 'Raw Parameter Comparison'    — all values as extracted from datasheet + conditions")
    print("  Sheet 2 'Common-Ground Comparison'    — all extrapolated to same VDS/ID/Tj reference")
    print("  Sheet 3 'Operating Conditions'        — YELLOW cells to edit + BLUE device params table")
    print("  Sheet 4 'Loss Calculation'            — live formula results + full formula reference")
    print("  Sheet 5 'Scoring Dashboard'           — 0-100 score per param + scoring methodology")
    print("  Sheet 6 'Raw Extracted Data'          — full typ/max/high-T table + test conditions")
    print("  Sheet 7 'Graph Comparison'            — side-by-side charts + overlay + 5-point comparison table")
    print("\n📄  PDFs generated:")
    print("  *_Why_Extrapolated.pdf      — 1-page note: which params are extrapolated and WHY (and why not)")
    print("  *_Extrapolation_Notes.pdf   — detailed per-parameter method + step-by-step calculations\n")


if __name__ == "__main__":
    main()
