"""
MOSFET Comparison & Loss Analysis Tool  —  v3
==============================================
Reads the Universal Design Reference v5 extractor Excel format.

Column layout of each input file (v5 format):
  Col A (0): Symbol          e.g. "V_DS", "R_DS(on)", "E_on" ...
  Col B (1): Parameter name  (ignored — we use Symbol)
  Col C (2): Unit            e.g. "mΩ", "μJ", "pF"
  Col D (3): Typ @ 25 °C     ← PRIMARY value used
  Col E (4): Max @ 25 °C
  Col F (5): Typ @ High-T    ← used for Rds @ 125 °C
  Col G (6): Max @ High-T
  Col H (7): Test Conditions ← parsed for VDS/ID/Tj context
  Col I (8): Loss Relevance  (ignored)
  Col J (9): Status          "Found ✓" / "Not Found"

Row 2 holds "Part: XXX | Manufacturer: YYY | ..."
Row 4 is the header row.
Data starts at row 5 (0-indexed row 4).

PARAMETER MAPPINGS (Symbol → internal key):
  V_DS          → vdss          (V, rated)
  I_D           → id            (A, @ 25°C col D; high-T col F)
  R_DS(on)      → rds25 (col D typ@25°C), rds_max25 (col E max@25°C),
                  rds_highT (col F typ@highT)
  V_GS(th)      → vth           (V, typ col D, max col E)
  C_iss         → ciss          (pF)
  C_oss         → coss          (pF, typ col D; max col E)
  C_rss         → crss          (pF)
  Q_g           → qg            (nC, with test conditions in col H)
  Q_gd          → qgd           (nC)
  t_d(on)       → td_on         (ns, with test conditions)
  t_r           → tr            (ns)
  t_d(off)      → td_off        (ns)
  t_f           → tf            (ns)
  E_oss         → eoss          (μJ)
  V_SD          → vsd           (V)
  E_on          → eon           (μJ, with test conditions → _eon_itest, _eon_vds)
  E_off         → eoff          (μJ, with test conditions → _eoff_itest)
  t_rr          → trr           (ns)
  Q_rr / E_rr   → qrr           (nC)

HIGH-TEMP Rds handling:
  If col F (Typ @ High-T) is not "—" and the test condition in col H
  mentions a temperature (e.g. "Tj = 150 °C") we store that as rds_highT
  with its temperature in _rds_highT_tj.
  If no explicit temperature is found we default to 125 °C.

NORMALISATION (common-ground comparison):
  Rds_on(Tj_ref)  → linear interpolation/extrapolation between the two
                    temperature points (25 °C and the high-T point).
  Coss @ VDS_ref  → depletion-cap model  C = C0/√(1+VDS/Vbi)
  Eon/Eoff @ ID_ref → linear scaling  E_scaled = E_ref × (ID_ref / ID_test)
  Computed switching energy from timing (when Eon/Eoff not tabulated):
    E_on  ≈ ½ × VDS_test × ID_test × (t_r + t_d_on)   [ns×A×V → µJ /1000]
    E_off ≈ ½ × VDS_test × ID_test × (t_f + t_d_off)

LOSS MODEL:
  P_cond   = Irms² × Rds_on(Tj) [Ω]
  P_sw_on  = Eon[J] × (Iavg / ID_test) × Fsw
  P_sw_off = Eoff[J] × (Iavg / ID_test) × Fsw
  P_coss   = Eoss[J] × Fsw
  P_gate   = Qg[C] × VGS_drive × Fsw
  P_body_diode = Vsd × Iavg × deadtime_ratio  (informational, not summed)
  P_total  = P_cond + P_sw_on + P_sw_off + P_coss + P_gate

All loss values in Excel are LIVE FORMULAS.
Yellow cells = user-editable operating conditions.
Blue cells   = device parameters (pre-filled, fully editable).
"""

import sys, os, re, warnings, io, tempfile
from pathlib import Path

import numpy as np
try:
    from scipy.optimize import curve_fit
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ─── colours ──────────────────────────────────────────────────────────────────
C_TITLE   = "13315C";  C_SUBHDR  = "2E75B6";  C_HDR     = "1F3864"
C_PRI_BG  = "1A3C5E";  C_SEC_BG  = "2E75B6"
C_BEST    = "E2EFDA";  C_BESTF   = "375623"
C_WARN    = "FCE4D6";  C_WARNF   = "843C0C"
C_ALT     = "F2F8FF";  C_WHITE   = "FFFFFF"
C_INPUT   = "FFF2CC";  C_GOLD    = "FFD700"
C_GREEN   = "375623";  C_NORM    = "E8F5E9"
C_PRI_ROW = "E3EEF8";  C_DEVPARAM= "D6EEFF"
C_DERIVED = "FFF0CC"   # orange-yellow for values derived from timing

_t = Side(style="thin",   color="AAAAAA")
_m = Side(style="medium", color="444444")
BT = Border(left=_t, right=_t, top=_t, bottom=_t)
BM = Border(left=_m, right=_m, top=_m, bottom=_m)

def ff(h):  return PatternFill("solid", fgColor=h)
def ctr():  return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft():  return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def fnt(sz=10, bold=False, color="000000", italic=False):
    return Font(name="Calibri", size=sz, bold=bold, color=color, italic=italic)

def wc(ws, r, c, v, font=None, fill=None, align=None, border=None, nfmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font          = font
    if fill:   cell.fill          = fill
    if align:  cell.alignment     = align
    if border: cell.border        = border
    if nfmt:   cell.number_format = nfmt
    return cell

def merge_title(ws, r, c1, c2, text, bg=C_HDR, fg="FFFFFF", sz=12, h=30):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(row=r, column=c1, value=text)
    c.font = Font(name="Calibri", bold=True, size=sz, color=fg)
    c.fill = ff(bg); c.alignment = ctr(); c.border = BM
    ws.row_dimensions[r].height = h

def gcol(n): return get_column_letter(n)

# ─── PARAMETER CATALOGUE ──────────────────────────────────────────────────────
# key: (display_name, unit, lower_is_better, is_priority, show_conditions)
PARAMS = {
    # ── PRIORITY ──────────────────────────────────────────────────────────────
    "vdss":   ("V_DS – Breakdown Voltage",           "V",    False, True,  False),
    "rds25":  ("R_DS(on) @ 25 °C  (typ)",            "mΩ",   True,  True,  True),
    "rds_ht": ("R_DS(on) @ High-T  (typ)",           "mΩ",   True,  True,  True),
    "qg":     ("Q_G – Total Gate Charge",            "nC",   True,  True,  True),
    "eon":    ("E_on – Turn-ON Switching Energy",    "µJ",   True,  True,  True),
    "eoff":   ("E_off – Turn-OFF Switching Energy",  "µJ",   True,  True,  True),
    # ── SECONDARY ─────────────────────────────────────────────────────────────
    "id":     ("I_D – Continuous Drain Current",     "A",    False, False, False),
    "vth":    ("V_GS(th) – Gate Threshold",          "V",    False, False, True),
    "ciss":   ("C_iss – Input Capacitance",          "pF",   True,  False, True),
    "coss":   ("C_oss – Output Capacitance",         "pF",   True,  False, True),
    "crss":   ("C_rss – Reverse Transfer Cap.",      "pF",   True,  False, False),
    "qgd":    ("Q_GD – Miller Gate-Drain Charge",    "nC",   True,  False, False),
    "eoss":   ("E_oss – Output Cap Stored Energy",   "µJ",   True,  False, False),
    "vsd":    ("V_SD – Body Diode Fwd Voltage",      "V",    True,  False, True),
    "trr":    ("t_rr – Reverse Recovery Time",       "ns",   True,  False, True),
    "qrr":    ("Q_rr – Rev. Recovery Charge",        "nC",   True,  False, False),
    "td_on":  ("t_d(on) – Turn-On Delay",            "ns",   True,  False, True),
    "tr":     ("t_r – Rise Time",                    "ns",   True,  False, False),
    "td_off": ("t_d(off) – Turn-Off Delay",          "ns",   True,  False, False),
    "tf":     ("t_f – Fall Time",                    "ns",   True,  False, False),
}

PRIORITY_KEYS  = [k for k, v in PARAMS.items() if v[3]]
SECONDARY_KEYS = [k for k, v in PARAMS.items() if not v[3]]
ALL_KEYS       = PRIORITY_KEYS + SECONDARY_KEYS

# ─── v5 EXCEL READER ──────────────────────────────────────────────────────────

def _val(raw):
    """Convert a cell value to float, returning None for blanks/dashes."""
    if raw is None: return None
    s = str(raw).strip()
    if s in ("", "—", "-", "N/A", "Not Found"): return None
    m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
    return float(m.group()) if m else None

def _parse_tcond(tcond):
    """
    Parse test condition string.
    Returns dict with any of: vds, id, vgs, rg, tj
    Examples:
      "VGS=10V, ID=15A"         → {vgs:10, id:15}
      "VDD=400V, ID=30A, RG=4.7Ω, VGS=10V"  → {vds:400, id:30, rg:4.7, vgs:10}
      "Tj=150°C"                → {tj:150}
    """
    if not tcond: return {}
    t = str(tcond)
    out = {}
    # VDS / VDD / VDrain
    m = re.search(r"V(?:DS|DD|D)\s*=\s*([\d.]+)", t, re.I)
    if m: out["vds"] = float(m.group(1))
    # ID / IDS
    m = re.search(r"I_?D\s*=\s*([\d.]+)\s*A", t, re.I)
    if m: out["id"] = float(m.group(1))
    # VGS
    m = re.search(r"V_?GS\s*=\s*([\d.]+)\s*V", t, re.I)
    if m: out["vgs"] = float(m.group(1))
    # RG (gate resistance)
    m = re.search(r"R_?G\s*=\s*([\d.]+)", t, re.I)
    if m: out["rg"] = float(m.group(1))
    # Tj
    m = re.search(r"T_?j\s*=\s*([\d.]+)\s*°?[Cc]", t, re.I)
    if m: out["tj"] = float(m.group(1))
    return out


def read_v5_file(path):
    """
    Read a v5-format MOSFET parameter Excel file.
    Returns a rich dict with all extracted values plus metadata.
    """
    wb  = openpyxl.load_workbook(path, data_only=True)
    ws  = wb.active
    d   = {
        "part": Path(path).stem, "mfr": "",
        "coss_rows": [],   # (vds, coss_pF) for extrapolation
    }

    # ── Row 2: Part / Manufacturer ──────────────────────────────────────────
    hdr = str(ws.cell(row=2, column=1).value or "")
    m = re.search(r"Part:\s*([^|]+)", hdr)
    if m: d["part"] = m.group(1).strip()
    m = re.search(r"Manufacturer:\s*([^|]+)", hdr)
    if m: d["mfr"]  = m.group(1).strip()

    # ── Data rows (row 5 onward) ─────────────────────────────────────────────
    for row in ws.iter_rows(min_row=5, values_only=True):
        sym    = str(row[0] or "").strip()  if row[0] else ""
        unit   = str(row[2] or "").strip()  if len(row) > 2 else ""
        v25    = _val(row[3])               if len(row) > 3 else None  # Typ @ 25°C
        vmax25 = _val(row[4])               if len(row) > 4 else None  # Max @ 25°C
        v_ht   = _val(row[5])               if len(row) > 5 else None  # Typ @ High-T
        vmax_ht= _val(row[6])               if len(row) > 6 else None
        tcond  = str(row[7] or "").strip()  if len(row) > 7 else ""
        status = str(row[9] or "").strip()  if len(row) > 9 else ""

        if sym.startswith("▶") or sym.lower() in ("symbol", "legend:", ""):
            continue

        tc = _parse_tcond(tcond)

        # ── V_DS ───────────────────────────────────────────────────────────
        if sym == "V_DS":
            d["vdss"]      = v25
            d["vdss_max"]  = vmax25
            d["vdss_tcond"]= tcond

        # ── I_D ────────────────────────────────────────────────────────────
        elif sym == "I_D":
            d["id"]       = v25           # @ 25°C
            d["id_highT"] = v_ht          # @ high-T (e.g. 100°C)
            d["id_tcond"] = tcond

        # ── R_DS(on) ────────────────────────────────────────────────────────
        elif sym == "R_DS(on)":
            d["rds25"]      = v25
            d["rds25_max"]  = vmax25
            d["rds_ht"]     = v_ht          # High-T typical (if present)
            d["rds_ht_max"] = vmax_ht
            d["rds_tcond"]  = tcond
            # Parse temperature of high-T column from tcond or default 125°C
            d["_rds_ht_tj"] = tc.get("tj", 125.0) if v_ht is not None else None
            d["_rds_id_test"] = tc.get("id")   # current at which Rds was measured
            d["_rds_vgs_test"]= tc.get("vgs")

        # ── V_GS(th) ────────────────────────────────────────────────────────
        elif sym == "V_GS(th)":
            d["vth"]     = v25
            d["vth_max"] = vmax25
            d["vth_tcond"]= tcond

        # ── C_iss ────────────────────────────────────────────────────────────
        elif sym == "C_iss":
            d["ciss"]      = v25
            d["ciss_tcond"]= tcond
            d["_ciss_vds"] = tc.get("vds")

        # ── C_oss ────────────────────────────────────────────────────────────
        elif sym == "C_oss":
            d["coss"]      = v25
            d["coss_max"]  = vmax25
            d["coss_tcond"]= tcond
            vds_c = tc.get("vds")
            # Try to extract VDS from tcond or fall back to unit context
            if vds_c is None:
                # sometimes condition is just "—" – try to infer from VDSS
                # we'll add to coss_rows when we know VDSS (done after loop)
                pass
            if vds_c is not None and v25 is not None:
                d["coss_rows"].append((vds_c, v25))
            d["_coss_vds_test"] = vds_c

        # ── C_rss ────────────────────────────────────────────────────────────
        elif sym == "C_rss":
            d["crss"]      = v25
            d["crss_tcond"]= tcond

        # ── Q_g ──────────────────────────────────────────────────────────────
        elif sym == "Q_g":
            d["qg"]       = v25
            d["qg_tcond"] = tcond
            d["_qg_vds"]  = tc.get("vds")
            d["_qg_id"]   = tc.get("id")
            d["_qg_vgs"]  = tc.get("vgs")

        # ── Q_gd ─────────────────────────────────────────────────────────────
        elif sym == "Q_gd":
            d["qgd"]       = v25
            d["qgd_tcond"] = tcond

        # ── t_d(on) ──────────────────────────────────────────────────────────
        elif sym == "t_d(on)":
            d["td_on"]       = v25
            d["td_on_tcond"] = tcond
            d["_sw_vds"]     = tc.get("vds")   # switching test VDS (shared)
            d["_sw_id"]      = tc.get("id")    # switching test ID
            d["_sw_vgs"]     = tc.get("vgs")
            d["_sw_rg"]      = tc.get("rg")

        # ── t_r ──────────────────────────────────────────────────────────────
        elif sym == "t_r":
            d["tr"]       = v25
            d["tr_tcond"] = tcond

        # ── t_d(off) ─────────────────────────────────────────────────────────
        elif sym == "t_d(off)":
            d["td_off"]       = v25
            d["td_off_tcond"] = tcond

        # ── t_f ──────────────────────────────────────────────────────────────
        elif sym == "t_f":
            d["tf"]       = v25
            d["tf_tcond"] = tcond

        # ── E_oss ────────────────────────────────────────────────────────────
        elif sym == "E_oss":
            d["eoss"]       = v25
            d["eoss_tcond"] = tcond

        # ── V_SD ─────────────────────────────────────────────────────────────
        elif sym == "V_SD":
            d["vsd"]       = v25
            d["vsd_tcond"] = tcond

        # ── E_on ─────────────────────────────────────────────────────────────
        elif sym == "E_on":
            d["eon"]          = v25
            d["eon_tcond"]    = tcond
            d["_eon_vds"]     = tc.get("vds")
            d["_eon_id"]      = tc.get("id")

        # ── E_off ────────────────────────────────────────────────────────────
        elif sym == "E_off":
            d["eoff"]         = v25
            d["eoff_tcond"]   = tcond
            d["_eoff_vds"]    = tc.get("vds")
            d["_eoff_id"]     = tc.get("id")

        # ── t_rr ─────────────────────────────────────────────────────────────
        elif sym == "t_rr":
            d["trr"]       = v25
            d["trr_tcond"] = tcond

        # ── Q_rr / E_rr ──────────────────────────────────────────────────────
        elif re.match(r"Q_rr", sym):
            d["qrr"]       = v25
            d["qrr_tcond"] = tcond

    # ── Post-processing ───────────────────────────────────────────────────────

    # If Coss was found but no test VDS was extracted, add a fallback entry
    # using Coss_test_VDS = VDSS × 0.5 as a rough guess (or 400V if unknown)
    if d.get("coss") is not None and not d.get("coss_rows"):
        fallback_vds = d.get("vdss", 650.0) * 0.5 if d.get("vdss") else 400.0
        d["coss_rows"].append((fallback_vds, d["coss"]))
        d["_coss_vds_test"] = fallback_vds

    # Derive Eoss from Coss when not tabulated
    if d.get("eoss") is None and d.get("coss") is not None:
        vds_t = d.get("_coss_vds_test") or (d.get("vdss") or 650.0) * 0.5
        d["eoss"] = round(0.5 * d["coss"] * 1e-12 * vds_t**2 * 1e6, 4)
        d["_eoss_derived"] = True

    # Derive Eon/Eoff from switching times when not tabulated
    if d.get("eon") is None:
        vds_t = d.get("_sw_vds") or d.get("_eon_vds") or d.get("_qg_vds") or (d.get("vdss") or 400.0)*0.6
        id_t  = d.get("_sw_id")  or d.get("_eon_id")  or d.get("id") or 15.0
        tr    = d.get("tr")
        tdon  = d.get("td_on")
        if tr is not None and tdon is not None:
            # E = ½ × V × I × t_switch [V × A × ns → µJ = /1000]
            d["eon"]  = round(0.5 * vds_t * id_t * (tr + tdon) / 1000.0, 4)
            d["_eon_derived"]  = True
            d["_eon_vds"]      = vds_t
            d["_eon_id"]       = id_t
        elif tr is not None:
            d["eon"]  = round(0.5 * vds_t * id_t * tr / 1000.0, 4)
            d["_eon_derived"]  = True
            d["_eon_vds"]      = vds_t
            d["_eon_id"]       = id_t

    if d.get("eoff") is None:
        vds_t = d.get("_sw_vds")  or d.get("_eoff_vds") or d.get("_qg_vds") or (d.get("vdss") or 400.0)*0.6
        id_t  = d.get("_sw_id")   or d.get("_eoff_id")  or d.get("id") or 15.0
        tf    = d.get("tf")
        tdoff = d.get("td_off")
        if tf is not None and tdoff is not None:
            d["eoff"] = round(0.5 * vds_t * id_t * (tf + tdoff) / 1000.0, 4)
            d["_eoff_derived"] = True
            d["_eoff_vds"]     = vds_t
            d["_eoff_id"]      = id_t
        elif tf is not None:
            d["eoff"] = round(0.5 * vds_t * id_t * tf / 1000.0, 4)
            d["_eoff_derived"] = True
            d["_eoff_vds"]     = vds_t
            d["_eoff_id"]      = id_t

    # Default gate drive voltage (from Qg test condition if available)
    d.setdefault("vgs_drive", d.get("_qg_vgs") or d.get("_sw_vgs") or 10.0)
    d.setdefault("tjmax", 175.0)

    wb.close()
    return d


def mosfet_label(path, d):
    part = (d.get("part") or "").strip()
    mfr  = (d.get("mfr")  or "").strip()
    if part and part.lower() not in ("unknown", ""):
        return f"{part} ({mfr})" if mfr else part
    return Path(path).stem


# ─── PHYSICS MODELS ───────────────────────────────────────────────────────────

def extrapolate_rds(d, tj_ref):
    """
    Return (rds_at_tj, method_string).
    Uses linear interpolation/extrapolation between:
      - rds25 @ 25 °C
      - rds_ht @ _rds_ht_tj (default 125 °C)
    """
    r25  = d.get("rds25")
    r_ht = d.get("rds_ht")
    t_ht = d.get("_rds_ht_tj") or 125.0

    if r25 is None and r_ht is None:
        return None, "No Rds data available"

    if r_ht is None:
        # Only 25 °C point — just return it with a note
        return (round(r25, 4),
                f"Only 25°C point available; Rds={r25:.2f}mΩ (no temp correction)")

    if r25 is None:
        return (round(r_ht, 4),
                f"Only high-T ({t_ht:.0f}°C) point; Rds={r_ht:.2f}mΩ (no temp correction)")

    if abs(tj_ref - 25.0) < 0.5:
        return round(r25, 4), f"Direct @ 25°C: {r25:.2f} mΩ"

    if abs(tj_ref - t_ht) < 0.5:
        return round(r_ht, 4), f"Direct @ {t_ht:.0f}°C: {r_ht:.2f} mΩ"

    slope = (r_ht - r25) / (t_ht - 25.0)
    val   = r25 + slope * (tj_ref - 25.0)
    return (round(val, 4),
            f"Linear interp: {r25:.2f}mΩ@25°C → {r_ht:.2f}mΩ@{t_ht:.0f}°C → {val:.2f}mΩ@{tj_ref:.0f}°C")


def extrapolate_coss(coss_rows, vds_target):
    """
    Extrapolate C_oss to vds_target using depletion-cap model C = C0/√(1+VDS/Vbi).
    coss_rows: list of (vds, coss_pF) tuples.
    """
    rows = [(v, c) for v, c in coss_rows if v is not None and c is not None and v >= 0]
    if not rows:
        return None, "No Coss data"

    rows_s = sorted(rows, key=lambda x: x[0])
    VDs = np.array([r[0] for r in rows_s], dtype=float)
    Cs  = np.array([r[1] for r in rows_s], dtype=float)

    # Single-point: use 1/√V scaling (depletion approximation)
    if len(rows) == 1:
        v0, c0 = rows_s[0]
        if abs(v0 - vds_target) < 1.0:
            return round(float(c0), 2), f"Direct: Coss={c0:.0f}pF @ VDS={v0:.0f}V"
        val = c0 * np.sqrt(max(v0, 1.0) / max(vds_target, 1.0))
        return (round(float(val), 2),
                f"1/√V scaling: Coss={c0:.0f}pF@{v0:.0f}V → {val:.1f}pF@{vds_target:.0f}V")

    # Multi-point: fit C0/√(1+VDS/Vbi)
    def _dep(VD, C0, Vbi):
        return C0 / np.sqrt(1.0 + np.maximum(VD, 0) / max(abs(Vbi), 0.01))

    if HAS_SCIPY:
        try:
            Vg = max(VDs[0], 1.0)
            p0 = [Cs[0] * np.sqrt(1 + VDs[0] / Vg), Vg]
            popt, _ = curve_fit(_dep, VDs, Cs, p0=p0,
                                bounds=([0, 0.01], [1e8, 1e5]), maxfev=8000)
            val = _dep(vds_target, *popt)
            if 0 < val < Cs.max() * 20:
                pts = "; ".join(f"{c:.0f}pF@{v:.0f}V" for v, c in rows_s)
                return (round(float(val), 2),
                        f"Depletion fit C0/√(1+V/Vbi): C0={popt[0]:.1f}pF, Vbi={popt[1]:.2f}V [{pts}]")
        except Exception:
            pass

    # Power-law fallback
    try:
        b, a = np.polyfit(np.log(VDs + 1), np.log(Cs), 1)
        val  = np.exp(a) * (vds_target + 1) ** b
        if val > 0:
            return round(float(val), 2), f"Power-law fit Coss∝(VDS+1)^{b:.3f}"
    except Exception:
        pass

    # sqrt fallback
    v0, c0 = min(rows_s, key=lambda r: abs(r[0] - vds_target))
    val = c0 * np.sqrt(max(v0, 1) / max(vds_target, 1))
    return round(float(val), 2), f"Sqrt fallback: Coss={c0:.0f}pF@{v0:.0f}V"


def scale_energy(e_ref, id_test, id_target):
    """Linear scaling of Eon/Eoff with drain current."""
    if e_ref is None:
        return None, "Not available"
    if id_test is None or id_test == 0:
        return round(e_ref, 4), f"Direct (test current unknown; no scaling)"
    if abs(id_target - id_test) < 0.5:
        return round(e_ref, 4), f"Direct @ ID={id_test:.1f}A (matches ID_ref)"
    val = e_ref * (id_target / id_test)
    return round(val, 4), f"Linear scale: {e_ref:.3f}µJ × ({id_target:.1f}/{id_test:.1f}A)"


# ─── GRAPH EXTRACTION & COMPARISON ───────────────────────────────────────────

# Chart types rendered on log scales ─ images only, no extraction table
_LOG_CHART_TYPES = {'capacitance_vs_vds', 'thermal_impedance'}

_CHART_TYPE_PATTERNS = {
    'capacitance_vs_vds': ['capacitance', 'coss', 'ciss', 'crss', 'c_oss', 'c_iss',
                            'c_rss', 'output cap', 'input cap', 'reverse transfer cap'],
    'thermal_impedance':  ['thermal', 'zth', 'z_th', 'impedance', 'junction-to-case',
                            'transient thermal'],
    'transfer_char':      ['transfer char', 'id vs vgs', 'id-vgs', 'id vs. vgs'],
    'output_char':        ['output char', 'id vs vds', 'id-vds', 'drain current vs drain'],
    'body_diode':         ['body diode', 'vsd', 'forward voltage', 'diode forward',
                            'source drain'],
    'soa':                ['safe operating', 'soa', 'safe area'],
    'gate_charge':        ['gate charge', 'vgs vs qg', 'gate voltage vs charge'],
    'rds_vs_temp':        ['rds vs', 'on-resistance vs temp', 'rds(on) vs',
                            'normalised rds'],
}

_CHART_DISPLAY_NAMES = {
    'capacitance_vs_vds': 'Capacitance vs V_DS  (C_oss / C_iss / C_rss)',
    'thermal_impedance':  'Thermal Impedance  Z_th vs t_p',
    'transfer_char':      'Transfer Characteristics  I_D vs V_GS',
    'output_char':        'Output Characteristics  I_D vs V_DS',
    'body_diode':         'Body Diode Forward Voltage  V_SD vs I_SD',
    'soa':                'Safe Operating Area',
    'gate_charge':        'Gate Charge  V_GS vs Q_G',
    'rds_vs_temp':        'R_DS(on) vs Temperature',
    'unknown':            'Other Charts',
}

_CHART_AXIS_LABELS = {
    'capacitance_vs_vds': ('V_DS (V)',   'Capacitance (pF)'),
    'thermal_impedance':  ('t_p (s)',    'Z_th (°C/W)'),
    'transfer_char':      ('V_GS (V)',   'I_D (A)'),
    'output_char':        ('V_DS (V)',   'I_D (A)'),
    'body_diode':         ('I_SD (A)',   'V_SD (V)'),
    'gate_charge':        ('Q_G (nC)',   'V_GS (V)'),
    'rds_vs_temp':        ('T_j (°C)',   'R_DS(on) (mΩ)'),
}

_TABLE_Y_UNITS = {
    'capacitance_vs_vds': 'pF',
    'transfer_char':      'A',
    'output_char':        'A',
    'body_diode':         'V',
    'gate_charge':        'V',
    'rds_vs_temp':        'mΩ',
}


def _get_chart_title(chart):
    title = getattr(chart, 'title', None)
    if title is None:
        return ""
    if isinstance(title, str):
        return title
    tx = getattr(title, 'tx', None)
    if tx is not None:
        rich = getattr(tx, 'rich', None)
        if rich is not None:
            texts = []
            for para in getattr(rich, 'p', []):
                for run in getattr(para, 'r', []):
                    t = getattr(run, 't', '')
                    if t:
                        texts.append(str(t))
            if texts:
                return ' '.join(texts)
    return str(title)


def _get_series_label(ser):
    title = getattr(ser, 'title', None)
    if title is None:
        return ""
    if isinstance(title, str):
        return title
    str_ref = getattr(title, 'strRef', None)
    if str_ref is not None:
        sc = getattr(str_ref, 'strCache', None)
        if sc is not None:
            for pt in getattr(sc, 'pt', []):
                v = getattr(pt, 'v', None)
                if v:
                    return str(v)
    v = getattr(title, 'v', None)
    return str(v) if v else str(title)


def _read_cells_from_formula(wb, formula):
    """Read cell values from a formula like 'Sheet1!$A$1:$A$10'."""
    try:
        formula = str(formula).strip()
        if '!' not in formula:
            return []
        sheet_part, cell_part = formula.split('!', 1)
        sheet_name = sheet_part.strip("'\"")
        cell_part  = cell_part.replace('$', '')
        if sheet_name not in wb.sheetnames:
            return []
        ws     = wb[sheet_name]
        cr     = ws[cell_part]
        values = []
        if hasattr(cr, 'value'):
            v = cr.value
            try:
                values.append(float(v) if v is not None else None)
            except (TypeError, ValueError):
                values.append(None)
        else:
            for item in cr:
                cells = item if hasattr(item, '__iter__') else [item]
                for cell in cells:
                    v = cell.value
                    try:
                        values.append(float(v) if v is not None else None)
                    except (TypeError, ValueError):
                        values.append(None)
        return values
    except Exception:
        return []


def _extract_numdata(num_src, wb=None):
    """Extract numeric values from an openpyxl chart NumDataSource."""
    if num_src is None:
        return []
    result_dict = {}

    nr = getattr(num_src, 'numRef', None)
    if nr is not None:
        nc = getattr(nr, 'numCache', None)
        if nc is not None:
            for pt in getattr(nc, 'pt', []):
                idx = getattr(pt, 'idx', None)
                val = getattr(pt, 'v',   None)
                if idx is not None and val is not None:
                    try:
                        result_dict[int(idx)] = float(val)
                    except (ValueError, TypeError):
                        pass
            if result_dict:
                mx = max(result_dict.keys())
                return [result_dict.get(i) for i in range(mx + 1)]
        f = getattr(nr, 'f', None)
        if f and wb is not None:
            vals = _read_cells_from_formula(wb, f)
            if vals:
                return vals

    nl = getattr(num_src, 'numLit', None)
    if nl is not None:
        for pt in getattr(nl, 'pt', []):
            idx = getattr(pt, 'idx', None)
            val = getattr(pt, 'v',   None)
            if idx is not None and val is not None:
                try:
                    result_dict[int(idx)] = float(val)
                except (ValueError, TypeError):
                    pass
        if result_dict:
            mx = max(result_dict.keys())
            return [result_dict.get(i) for i in range(mx + 1)]

    return []


def _identify_chart_type(title, series_labels):
    text = (title + " " + " ".join(series_labels or [])).lower()
    for ctype, patterns in _CHART_TYPE_PATTERNS.items():
        if any(p in text for p in patterns):
            return ctype
    return 'unknown'


def extract_charts_from_file(path):
    """
    Read all embedded chart objects from an xlsx file.
    Returns list of dicts: {title, chart_type, series, is_log_scale, ws_name}
    Each series dict: {label, x:[floats], y:[floats]}
    """
    charts_out = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            for chart in getattr(ws, '_charts', []):
                chart_title = _get_chart_title(chart)
                series_list, ser_labels = [], []
                for ser in getattr(chart, 'series', []):
                    lbl = _get_series_label(ser)
                    ser_labels.append(lbl)
                    x_src  = getattr(ser, 'xVal', None) or getattr(ser, 'cat', None)
                    y_src  = getattr(ser, 'val',  None)
                    x_vals = _extract_numdata(x_src, wb)
                    y_vals = _extract_numdata(y_src, wb)
                    if x_vals or y_vals:
                        series_list.append({'label': lbl, 'x': x_vals, 'y': y_vals})
                ctype  = _identify_chart_type(chart_title, ser_labels)
                is_log = ctype in _LOG_CHART_TYPES
                charts_out.append({
                    'title':       chart_title,
                    'chart_type':  ctype,
                    'series':      series_list,
                    'is_log_scale': is_log,
                    'ws_name':     ws_name,
                })
        wb.close()
    except Exception as ex:
        print(f"  Warning: chart extraction from {Path(path).name}: {ex}")
    return charts_out


def _build_fallback_charts(mosfets):
    """
    Build synthetic chart dicts from tabular data when no Excel chart objects exist.
    The returned list has one entry per chart type; each series[i] belongs to device i.
    """
    result = []

    # C_oss vs V_DS (from coss_rows)
    if any(d.get('coss_rows') for d in mosfets):
        series = []
        for d in mosfets:
            rows = d.get('coss_rows', [])
            xs   = [r[0] for r in rows if r[0] is not None]
            ys   = [r[1] for r in rows if r[1] is not None]
            series.append({'label': 'C_oss', 'x': xs, 'y': ys})
        result.append({
            'title': 'C_oss vs V_DS',
            'chart_type': 'capacitance_vs_vds',
            'series': series,
            'is_log_scale': True,
            'ws_name': 'fallback',
            '_per_device': True,
        })

    # R_DS(on) vs Temperature (from 25 °C + high-T points)
    if any(d.get('rds25') is not None and d.get('rds_ht') is not None for d in mosfets):
        series = []
        for d in mosfets:
            xs, ys = [], []
            if d.get('rds25') is not None:
                xs.append(25.0); ys.append(d['rds25'])
            if d.get('rds_ht') is not None and d.get('_rds_ht_tj') is not None:
                xs.append(d['_rds_ht_tj']); ys.append(d['rds_ht'])
            series.append({'label': 'R_DS(on)', 'x': xs, 'y': ys})
        result.append({
            'title': 'R_DS(on) vs Temperature',
            'chart_type': 'rds_vs_temp',
            'series': series,
            'is_log_scale': False,
            'ws_name': 'fallback',
            '_per_device': True,
        })

    return result


def _match_charts(all_charts_per_file):
    """
    Group matching chart types across all device files.
    Returns list of groups:
      {chart_type, display_name, is_log_scale, _per_device,
       charts_per_device:[chart_dict or None]}
    """
    seen, all_types = set(), []
    for charts in all_charts_per_file:
        for c in charts:
            ct = c['chart_type']
            if ct not in seen:
                all_types.append(ct); seen.add(ct)

    groups = []
    for ctype in all_types:
        per_dev = []
        for dev_charts in all_charts_per_file:
            best = None
            for c in dev_charts:
                if c['chart_type'] != ctype:
                    continue
                if best is None:
                    best = c
                else:
                    cur_pts = sum(len(s.get('y', [])) for s in best.get('series', []))
                    new_pts = sum(len(s.get('y', [])) for s in c.get('series', []))
                    if new_pts > cur_pts:
                        best = c
            per_dev.append(best)

        if not any(c is not None for c in per_dev):
            continue

        is_per_dev = any(c.get('_per_device') for c in per_dev if c)
        groups.append({
            'chart_type':        ctype,
            'display_name':      _CHART_DISPLAY_NAMES.get(ctype,
                                     ctype.replace('_', ' ').title()),
            'is_log_scale':      ctype in _LOG_CHART_TYPES,
            '_per_device':       is_per_dev,
            'charts_per_device': per_dev,
        })
    return groups


_DEV_COLORS = ['#1F77B4', '#FF7F0E', '#2CA02C', '#D62728', '#9467BD', '#8C564B']


def _make_comparison_fig(group, labels):
    """
    Build a matplotlib figure: one subplot per device + one overlay subplot.
    Returns io.BytesIO PNG or None.
    """
    if not HAS_MPL:
        return None

    charts       = group['charts_per_device']
    chart_type   = group['chart_type']
    is_log       = group['is_log_scale']
    is_per_dev   = group.get('_per_device', False)
    n            = len(charts)
    display_name = group['display_name']
    x_lbl, y_lbl = _CHART_AXIS_LABELS.get(chart_type, ('X', 'Y'))

    fig, axes = plt.subplots(1, n + 1, figsize=(5.5 * (n + 1), 4.5),
                              squeeze=False, dpi=120)
    ax_over = axes[0][n]
    ax_over.set_title('Overlay Comparison', fontsize=9, fontweight='bold', pad=8)

    any_data = False

    for col, (chart_data, label) in enumerate(zip(charts, labels)):
        ax        = axes[0][col]
        dev_color = _DEV_COLORS[col % len(_DEV_COLORS)]
        ax.set_title(label, fontsize=8.5, fontweight='bold', pad=8)

        if chart_data is None:
            ax.text(0.5, 0.5, 'Not available', ha='center', va='center',
                    transform=ax.transAxes, fontsize=10, color='#888888', style='italic')
            ax.set_facecolor('#F5F5F5')
            for sp in ax.spines.values():
                sp.set_color('#CCCCCC')
            continue

        # Per-device mode: series[col] belongs to this device
        if is_per_dev:
            sers = chart_data.get('series', [])
            dev_sers = [sers[col]] if col < len(sers) else []
        else:
            dev_sers = chart_data.get('series', [])

        plotted = False
        for i, ser in enumerate(dev_sers):
            pairs = [(float(x), float(y)) for x, y in
                     zip(ser.get('x', []), ser.get('y', []))
                     if x is not None and y is not None]
            if not pairs:
                continue
            xs, ys  = zip(*sorted(pairs))
            color   = dev_color if is_per_dev else _DEV_COLORS[i % len(_DEV_COLORS)]
            ser_lbl = ser.get('label') or f"Series {i + 1}"
            ov_lbl  = label if (is_per_dev or i == 0) else f"{label} {ser_lbl}"

            if is_log:
                xp = [x for x, y in zip(xs, ys) if x > 0 and y > 0]
                yp = [y for x, y in zip(xs, ys) if x > 0 and y > 0]
                if xp:
                    ax.loglog(xp, yp, 'o-', color=color,
                              label=ser_lbl, linewidth=1.5, markersize=3)
                    ax_over.loglog(xp, yp, 'o-', color=dev_color if is_per_dev else color,
                                   label=ov_lbl, linewidth=1.5, markersize=3)
                    plotted = True; any_data = True
            else:
                ax.plot(xs, ys, 'o-', color=color,
                        label=ser_lbl, linewidth=1.5, markersize=4)
                ax_over.plot(xs, ys, 'o-', color=dev_color if is_per_dev else color,
                             label=ov_lbl, linewidth=1.5, markersize=4)
                plotted = True; any_data = True

        if plotted:
            ax.grid(True, alpha=0.3, which='both' if is_log else 'major')
            if len(dev_sers) > 1:
                ax.legend(fontsize=7, loc='best')
            ax.set_xlabel(x_lbl, fontsize=8); ax.set_ylabel(y_lbl, fontsize=8)
            ax.tick_params(labelsize=7)
        else:
            ax.text(0.5, 0.5, 'No data\nextracted', ha='center', va='center',
                    transform=ax.transAxes, fontsize=9, color='#888888', style='italic')
            ax.set_facecolor('#F9F9F9')

    if any_data:
        ax_over.grid(True, alpha=0.3, which='both' if is_log else 'major')
        ax_over.legend(fontsize=7, loc='best')
        ax_over.set_xlabel(x_lbl, fontsize=8)
        ax_over.set_ylabel(y_lbl, fontsize=8)
        ax_over.tick_params(labelsize=7)
    else:
        ax_over.text(0.5, 0.5, 'No data', ha='center', va='center',
                     transform=ax_over.transAxes, fontsize=10,
                     color='#888888', style='italic')

    fig.suptitle(display_name, fontsize=11, fontweight='bold', y=1.01)
    plt.tight_layout(rect=[0, 0, 1, 0.96])

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=120, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    buf.seek(0)
    return buf


def _extract_table_points(group, labels, n_points=5):
    """
    Interpolate n_points at evenly-spaced x values from the primary series.
    Returns (x_unit_label, list_of_rows) where each row = [x, y_dev0, y_dev1, …]
    """
    charts     = group['charts_per_device']
    chart_type = group['chart_type']
    is_per_dev = group.get('_per_device', False)
    x_lbl, _   = _CHART_AXIS_LABELS.get(chart_type, ('X', 'Y'))

    all_x = []
    for idx, cd in enumerate(charts):
        if cd is None:
            continue
        sers = cd.get('series', [])
        ser  = (sers[idx] if is_per_dev and idx < len(sers)
                else (sers[0] if sers else None))
        if ser:
            all_x += [float(v) for v in ser.get('x', []) if v is not None]

    if not all_x:
        return x_lbl, []

    x_min, x_max = min(all_x), max(all_x)
    ref_xs = [x_min + (x_max - x_min) * i / max(n_points - 1, 1)
               for i in range(n_points)]

    rows = []
    for xi in ref_xs:
        row = [round(xi, 4)]
        for idx, cd in enumerate(charts):
            if cd is None:
                row.append(None); continue
            sers   = cd.get('series', [])
            ser    = (sers[idx] if is_per_dev and idx < len(sers)
                      else (sers[0] if sers else None))
            y_val  = None
            if ser:
                pairs = sorted([(float(x), float(y))
                                for x, y in zip(ser.get('x', []), ser.get('y', []))
                                if x is not None and y is not None])
                if pairs:
                    xs, ys = zip(*pairs)
                    if xi <= xs[0]:
                        y_val = ys[0]
                    elif xi >= xs[-1]:
                        y_val = ys[-1]
                    else:
                        for j in range(len(xs) - 1):
                            if xs[j] <= xi <= xs[j + 1]:
                                t     = (xi - xs[j]) / (xs[j + 1] - xs[j])
                                y_val = ys[j] + t * (ys[j + 1] - ys[j])
                                break
            row.append(round(y_val, 4) if y_val is not None else None)
        rows.append(row)
    return x_lbl, rows


def build_graph_comparison_sheet(wb_out, paths, mosfets, labels, all_charts_per_file):
    """
    Adds 'Graph Comparison' sheet to wb_out.
    Returns list of temp PNG file paths (caller deletes them after wb_out.save()).
    """
    ws = wb_out.create_sheet("Graph Comparison")
    ws.sheet_view.showGridLines = False
    n         = len(paths)
    tmp_files = []
    ncols     = max(4, n + 3)

    merge_title(ws, 1, 1, ncols,
                "GRAPH COMPARISON  —  side-by-side charts + overlay + extracted comparison table",
                bg=C_TITLE, sz=13, h=36)

    # Decide whether we have real chart objects or need fallback synthetic charts
    found_real = any(c.get('ws_name') != 'fallback'
                     for charts in all_charts_per_file for c in charts)

    if not found_real:
        merge_title(ws, 2, 1, ncols,
                    "⚠  No embedded chart objects found in the Excel files — "
                    "showing synthetic charts built from extracted tabular data.",
                    bg=C_WARN, fg=C_WARNF, sz=9, h=20)
        fallbacks          = _build_fallback_charts(mosfets)
        # Wrap in a per-file list: first file "owns" all fallback charts
        per_file_fallback  = [fallbacks] + [[] for _ in range(n - 1)]
        groups             = _match_charts(per_file_fallback)
    else:
        merge_title(ws, 2, 1, ncols,
                    "Each device shown individually + overlay comparison on the right.  "
                    "Log-scale charts (C_oss/C_iss/C_rss vs V_DS, Z_th vs t_p): images only.",
                    bg=C_SUBHDR, fg="FFFFFF", sz=9, h=20)
        groups = _match_charts(all_charts_per_file)

    if not groups:
        c3 = ws.cell(row=3, column=1, value="No chart data available.")
        c3.font = fnt(11, italic=True, color="888888")
        return tmp_files

    current_row = 3

    for group in groups:
        chart_type   = group['chart_type']
        display_name = group['display_name']
        is_log       = group['is_log_scale']

        # ── Section header ────────────────────────────────────────────────────
        merge_title(ws, current_row, 1, ncols,
                    f"▶  {display_name}",
                    bg=C_SUBHDR, sz=11, h=28)
        current_row += 1

        if is_log:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=ncols)
            nc = ws.cell(row=current_row, column=1,
                         value="ℹ  Log-scale chart — comparison table omitted "
                               "(Coss/Ciss/Crss vs V_DS and Z_th vs t_p: images only)")
            nc.font = fnt(9, italic=True, color="555555")
            nc.fill = ff("FFFBEA"); nc.alignment = lft()
            current_row += 1

        # ── Comparison figure ─────────────────────────────────────────────────
        fig_buf = _make_comparison_fig(group, labels)

        if not HAS_MPL:
            ws.cell(row=current_row, column=1).value = (
                "⚠  matplotlib not installed — run:  pip install matplotlib")
            ws.cell(row=current_row, column=1).font = fnt(10, italic=True, color=C_WARNF)
            current_row += 2
        elif fig_buf is not None:
            try:
                tf = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                tf.write(fig_buf.getvalue()); tf.close()
                tmp_files.append(tf.name)

                from openpyxl.drawing.image import Image as XLImage
                img        = XLImage(tf.name)
                img.width  = min(360 * (n + 1), 1080)
                img.height = 310
                ws.add_image(img, f"A{current_row}")

                img_rows = 18
                for r in range(current_row, current_row + img_rows):
                    ws.row_dimensions[r].height = 19
                current_row += img_rows
            except Exception as ex:
                ws.cell(row=current_row, column=1).value = f"[Image error: {ex}]"
                ws.cell(row=current_row, column=1).font  = fnt(9, italic=True, color="AA0000")
                current_row += 2
        else:
            ws.cell(row=current_row, column=1).value = "[No chart data to plot]"
            ws.cell(row=current_row, column=1).font  = fnt(9, italic=True, color="888888")
            current_row += 2

        # ── Comparison table (non-log charts only) ────────────────────────────
        if not is_log:
            x_lbl, rows = _extract_table_points(group, labels, n_points=5)
            if rows:
                current_row += 1
                merge_title(ws, current_row, 1, 1 + n,
                            f"Extracted Comparison  —  {display_name}  (5 common x-points)",
                            bg=C_HDR, fg="FFFFFF", sz=9, h=20)
                current_row += 1

                # Header
                ws.row_dimensions[current_row].height = 22
                wc(ws, current_row, 1, x_lbl,
                   font=fnt(10, bold=True, color="FFFFFF"),
                   fill=ff(C_SUBHDR), align=ctr(), border=BT)
                y_unit = _TABLE_Y_UNITS.get(chart_type, '')
                for d_i, lbl in enumerate(labels):
                    wc(ws, current_row, 2 + d_i, f"{lbl}  ({y_unit})",
                       font=fnt(10, bold=True, color="FFFFFF"),
                       fill=ff(C_SUBHDR), align=ctr(), border=BT)
                current_row += 1

                # Rows (lower = better for most except output/transfer char)
                lower_better = chart_type not in ('transfer_char', 'output_char')
                for r_i, row_data in enumerate(rows):
                    ws.row_dimensions[current_row].height = 20
                    bg = C_ALT if r_i % 2 == 0 else C_WHITE
                    wc(ws, current_row, 1,
                       round(row_data[0], 4) if row_data[0] is not None else "N/A",
                       font=fnt(10, bold=True), fill=ff(bg),
                       align=ctr(), border=BT, nfmt="0.0####")
                    y_vals_d = {i: row_data[i + 1] for i in range(n)
                                if row_data[i + 1] is not None}
                    best_i = _winner(y_vals_d, lower_better)
                    for d_i in range(n):
                        v       = row_data[d_i + 1]
                        is_best = (d_i == best_i)
                        wc(ws, current_row, 2 + d_i,
                           round(v, 4) if v is not None else "N/A",
                           font=fnt(10, bold=is_best,
                                    color=C_BESTF if is_best else "000000"),
                           fill=ff(C_BEST if is_best else bg),
                           align=ctr(), border=BT, nfmt="0.0####")
                    current_row += 1

            current_row += 2
        else:
            current_row += 2

    ws.column_dimensions["A"].width = 30
    for d_i in range(n):
        ws.column_dimensions[gcol(2 + d_i)].width = 24

    return tmp_files


# ─── NORMALISATION ─────────────────────────────────────────────────────────────

def compute_normalised(mosfets, ref):
    vds_ref = ref["vds_ref"]
    id_ref  = ref["id_ref"]
    tj_ref  = ref["tj_ref"]
    results = []

    for d in mosfets:
        nm = {}

        # Vdss — direct
        nm["vdss"]  = (d.get("vdss"),  "Direct (rated value)",             d.get("vdss_tcond", ""))

        # Rds @ 25°C and High-T — direct with conditions
        rds25_cond  = f"@ 25°C, {d.get('rds_tcond','')}" if d.get('rds_tcond') else "@ 25°C"
        rdsht_cond  = (f"@ {d.get('_rds_ht_tj',125):.0f}°C, {d.get('rds_tcond','')}"
                       if d.get('rds_ht') else "Not in datasheet")
        nm["rds25"] = (d.get("rds25"), f"Typ @ 25°C;  {d.get('rds_tcond','')}",  rds25_cond)
        nm["rds_ht"]= (d.get("rds_ht"), rdsht_cond, rdsht_cond)

        # For normalised comparison: Rds @ Tj_ref via interpolation
        rds_tj_v, rds_tj_m = extrapolate_rds(d, tj_ref)
        nm["_rds_tj"]= (rds_tj_v, rds_tj_m, "")

        # Qg with test conditions
        qg_cond = d.get("qg_tcond", "")
        nm["qg"]    = (d.get("qg"), f"Direct; test cond: {qg_cond}", qg_cond)

        # Eon / Eoff — scaled to ID_ref
        eon_id  = d.get("_eon_id")
        eoff_id = d.get("_eoff_id")
        eon_v,  eon_m  = scale_energy(d.get("eon"),  eon_id,  id_ref)
        eoff_v, eoff_m = scale_energy(d.get("eoff"), eoff_id, id_ref)
        der_note = " [derived from t_r/t_f]" if d.get("_eon_derived") else ""
        nm["eon"]  = (eon_v,  eon_m  + der_note, d.get("eon_tcond",  ""))
        nm["eoff"] = (eoff_v, eoff_m + ((" [derived from t_f/t_d(off)]") if d.get("_eoff_derived") else ""),
                      d.get("eoff_tcond", ""))

        # Secondary — direct
        nm["id"]    = (d.get("id"),    "Direct @ 25°C",          d.get("id_tcond", ""))
        nm["vth"]   = (d.get("vth"),   f"Typ; test: {d.get('vth_tcond','')}",  d.get("vth_tcond", ""))
        nm["ciss"]  = (d.get("ciss"),  f"@ {d.get('_ciss_vds','?')}V;  {d.get('ciss_tcond','')}", d.get("ciss_tcond",""))
        # Coss @ vds_ref
        coss_v, coss_m = extrapolate_coss(d.get("coss_rows", []), vds_ref)
        nm["coss"]  = (coss_v, coss_m, d.get("coss_tcond", ""))
        nm["crss"]  = (d.get("crss"),  "Direct", "")
        nm["qgd"]   = (d.get("qgd"),   "Direct", "")
        der_eoss = " [derived: ½·Coss·VDS²]" if d.get("_eoss_derived") else ""
        nm["eoss"]  = (d.get("eoss"),  f"Direct{der_eoss}", "")
        nm["vsd"]   = (d.get("vsd"),   f"Direct; {d.get('vsd_tcond','')}", d.get("vsd_tcond",""))
        nm["trr"]   = (d.get("trr"),   f"Direct; {d.get('trr_tcond','')}", d.get("trr_tcond",""))
        nm["qrr"]   = (d.get("qrr"),   "Direct", "")
        nm["td_on"] = (d.get("td_on"), f"Direct; {d.get('td_on_tcond','')}", d.get("td_on_tcond",""))
        nm["tr"]    = (d.get("tr"),    "Direct", "")
        nm["td_off"]= (d.get("td_off"),"Direct", "")
        nm["tf"]    = (d.get("tf"),    "Direct", "")

        results.append(nm)
    return results


# ─── WINNER HELPER ────────────────────────────────────────────────────────────

def _winner(vals_d, lower_better):
    if not vals_d: return None
    return (min(vals_d, key=vals_d.__getitem__) if lower_better
            else max(vals_d, key=vals_d.__getitem__))


# ─── REPORT BUILDER ───────────────────────────────────────────────────────────

def build_report(paths, mosfets, norm_list, ref, iavg_def, irms_def,
                 fsw_def, vout_def, vgs_drive_def, out_path,
                 all_charts_per_file=None):
    n      = len(paths)
    labels = [mosfet_label(p, d) for p, d in zip(paths, mosfets)]
    wb     = openpyxl.Workbook()
    VDS_ref= ref["vds_ref"]
    ID_ref = ref["id_ref"]
    TJ_ref = ref["tj_ref"]

    def _cw(ws, widths):
        for c, w in widths.items():
            ws.column_dimensions[c].width = w

    def _param_section(ws, start_row, keys, values_fn, show_method=False,
                        method_fn=None, cond_fn=None, section_label=None,
                        section_bg=C_HDR, fdc=3):
        row  = start_row
        wins = {i: 0 for i in range(n)}
        ncols = fdc - 1 + n + 1 + (1 if show_method else 0)

        if section_label:
            merge_title(ws, row, 1, ncols, section_label, bg=section_bg, sz=10, h=22)
            row += 1

        for key in keys:
            display, unit, lower_better, is_pri, _ = PARAMS[key]
            bg = C_PRI_ROW if is_pri else (C_ALT if row % 2 == 0 else C_WHITE)
            ws.row_dimensions[row].height = 22

            wc(ws, row, 1, display,
               font=fnt(10, bold=is_pri), fill=ff(bg), align=lft(), border=BT)
            wc(ws, row, 2, unit,
               font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            vals   = values_fn(key)
            vals_d = {i: v for i, v in enumerate(vals) if v is not None}
            best_i = _winner(vals_d, lower_better)

            for d_i, v in enumerate(vals):
                is_best = (d_i == best_i)
                wc(ws, row, fdc + d_i,
                   round(v, 4) if isinstance(v, float) else (v if v is not None else "N/A"),
                   font=fnt(10, bold=is_best,
                            color=C_BESTF if is_best else "000000"),
                   fill=ff(C_BEST if is_best else bg),
                   align=ctr(), border=BT, nfmt="0.0####")

            best_col = fdc + n
            if best_i is not None:
                wc(ws, row, best_col, labels[best_i],
                   font=fnt(10, bold=True, color=C_BESTF),
                   fill=ff(C_BEST), align=ctr(), border=BT)
                wins[best_i] += 1
            else:
                wc(ws, row, best_col, "—",
                   font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            if show_method and method_fn:
                meths  = method_fn(key)
                unique = list(dict.fromkeys(m for m in meths if m))
                wc(ws, row, best_col + 1, " | ".join(unique),
                   font=fnt(8, italic=True, color="444444"),
                   fill=ff("F8F8F8"), align=lft(), border=BT)

            row += 1
        return row, wins

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — RAW PARAMETER COMPARISON  (as-extracted from datasheets)
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = "Raw Parameter Comparison"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "C4"

    merge_title(ws1, 1, 1, 4+n,
                "📋  RAW PARAMETER COMPARISON  (as extracted — conditions may differ between devices)",
                bg=C_TITLE, sz=13, h=36)
    merge_title(ws1, 2, 1, 4+n,
                "⚠  Values are as-measured per datasheet.  Use Sheet 2 for fair common-ground comparison.",
                bg=C_WARN, fg=C_WARNF, sz=10, h=20)

    ws1.row_dimensions[3].height = 30
    for ci, (txt, bg) in enumerate(
            [("Parameter", C_HDR), ("Unit", C_HDR)] +
            [(lbl, C_SUBHDR) for lbl in labels] +
            [("Best", C_HDR), ("Test Conditions (from datasheet)", C_HDR)], 1):
        wc(ws1, 3, ci, txt,
           font=fnt(11, bold=True, color="FFFFFF"), fill=ff(bg), align=ctr(), border=BT)

    def raw_vals(k):
        return [d.get(k) for d in mosfets]

    def raw_conds(k):
        # map param key → tcond field name
        cmap = {"vdss":"vdss_tcond","rds25":"rds_tcond","rds_ht":"rds_tcond",
                "qg":"qg_tcond","eon":"eon_tcond","eoff":"eoff_tcond",
                "id":"id_tcond","vth":"vth_tcond","ciss":"ciss_tcond",
                "coss":"coss_tcond","crss":"","qgd":"","eoss":"eoss_tcond",
                "vsd":"vsd_tcond","trr":"trr_tcond","qrr":"qrr_tcond",
                "td_on":"td_on_tcond","tr":"tr_tcond","td_off":"td_off_tcond","tf":"tf_tcond"}
        cf = cmap.get(k, "")
        if not cf: return [""] * n
        return [d.get(cf, "") or "" for d in mosfets]

    # Custom param section that includes conditions column
    def _section_with_conds(ws, start_row, keys, section_label, section_bg, fdc=3):
        row  = start_row
        wins = {i: 0 for i in range(n)}
        ncols = fdc - 1 + n + 2   # +2 = Best + Conditions

        merge_title(ws, row, 1, ncols, section_label, bg=section_bg, sz=10, h=22)
        row += 1

        for key in keys:
            display, unit, lower_better, is_pri, _ = PARAMS[key]
            bg = C_PRI_ROW if is_pri else (C_ALT if row % 2 == 0 else C_WHITE)
            ws.row_dimensions[row].height = 22

            wc(ws, row, 1, display,
               font=fnt(10, bold=is_pri), fill=ff(bg), align=lft(), border=BT)
            wc(ws, row, 2, unit,
               font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            vals   = raw_vals(key)
            vals_d = {i: v for i, v in enumerate(vals) if v is not None}
            best_i = _winner(vals_d, lower_better)

            for d_i, v in enumerate(vals):
                is_best = (d_i == best_i)
                # Mark derived values with a slightly different fill
                is_derived = (key in ("eon","eoff") and mosfets[d_i].get(f"_{key}_derived"))
                cell_fill = C_DERIVED if is_derived else (C_BEST if is_best else bg)
                wc(ws, row, fdc + d_i,
                   round(v, 4) if isinstance(v, float) else (v if v is not None else "N/A"),
                   font=fnt(10, bold=is_best,
                            color=C_BESTF if is_best else ("7F5C00" if is_derived else "000000")),
                   fill=ff(cell_fill),
                   align=ctr(), border=BT, nfmt="0.0####")

            best_col = fdc + n
            if best_i is not None:
                wc(ws, row, best_col, labels[best_i],
                   font=fnt(10, bold=True, color=C_BESTF),
                   fill=ff(C_BEST), align=ctr(), border=BT)
                wins[best_i] += 1
            else:
                wc(ws, row, best_col, "—",
                   font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            # Conditions column — merge all device conditions into one cell
            conds = raw_conds(key)
            unique_c = "; ".join(dict.fromkeys(c for c in conds if c))
            wc(ws, row, best_col + 1, unique_c,
               font=fnt(8, italic=True, color="444444"),
               fill=ff("F8F8F8"), align=lft(), border=BT)

            row += 1
        return row, wins

    next1, w1a = _section_with_conds(ws1, 4, PRIORITY_KEYS,
                                     "★  PRIORITY PARAMETERS", C_PRI_BG)
    next1, w1b = _section_with_conds(ws1, next1 + 1, SECONDARY_KEYS,
                                     "SECONDARY PARAMETERS", C_SEC_BG)

    all_w1 = {i: w1a[i] + w1b[i] for i in range(n)}
    mw = max(all_w1.values()) if all_w1 else 0
    wr = next1 + 1
    ws1.row_dimensions[wr].height = 26
    merge_title(ws1, wr, 1, 2, "Total Wins", bg=C_SEC_BG, sz=10, h=26)
    for d_i in range(n):
        wv = all_w1[d_i]
        wc(ws1, wr, 3 + d_i, wv,
           font=fnt(12, bold=True, color=C_BESTF if wv == mw else "000000"),
           fill=ff(C_BEST if wv == mw else C_ALT), align=ctr(), border=BM)
    if mw > 0:
        bdi = max(all_w1, key=all_w1.__getitem__)
        wc(ws1, wr, 3 + n, labels[bdi],
           font=fnt(10, bold=True, color=C_BESTF), fill=ff(C_BEST), align=ctr(), border=BM)

    # Legend row
    wr2 = wr + 2
    ws1.merge_cells(f"A{wr2}:{gcol(3+n+1)}{wr2}")
    lc = ws1.cell(row=wr2, column=1,
                  value="🟡 Orange-yellow cells = value derived from switching times (t_r, t_f) — not directly tabulated in datasheet")
    lc.font  = fnt(9, italic=True, color="7F5C00")
    lc.fill  = ff(C_DERIVED)
    lc.alignment = lft()

    _cw(ws1, {"A": 38, "B": 9})
    for d_i in range(n): ws1.column_dimensions[gcol(3 + d_i)].width = 18
    ws1.column_dimensions[gcol(3 + n)].width = 26
    ws1.column_dimensions[gcol(4 + n)].width = 52

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — COMMON-GROUND COMPARISON
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Common-Ground Comparison")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "C5"

    merge_title(ws2, 1, 1, 4 + n,
                (f"✅  COMMON-GROUND COMPARISON  "
                 f"(Rds @ Tj={TJ_ref:.0f}°C  |  Coss @ VDS={VDS_ref:.0f}V  |  Eon/Eoff @ ID={ID_ref:.1f}A)"),
                bg=C_TITLE, sz=13, h=36)
    merge_title(ws2, 2, 1, 4 + n,
                ("Rds: linear interp between 25°C & high-T datasheet points  |  "
                 "Coss: depletion model C0/√(1+VDS/Vbi)  |  Eon/Eoff: linear scaling with ID"),
                bg=C_SUBHDR, fg="FFFFFF", sz=9, h=18)
    merge_title(ws2, 3, 1, 4 + n,
                f"Reference basis: VDS_ref={VDS_ref:.0f}V  |  ID_ref={ID_ref:.1f}A  |  Tj_ref={TJ_ref:.0f}°C",
                bg=C_NORM, fg=C_GREEN, sz=10, h=20)

    ws2.row_dimensions[4].height = 28
    for ci, (txt, bg) in enumerate(
            [("Parameter", C_HDR), ("Unit", C_HDR)] +
            [(lbl, C_SUBHDR) for lbl in labels] +
            [("Best", C_HDR), ("Extrapolation / Scaling Method", C_HDR)], 1):
        wc(ws2, 4, ci, txt,
           font=fnt(11, bold=True, color="FFFFFF"), fill=ff(bg), align=ctr(), border=BT)

    # Use _rds_tj for priority comparison instead of rds25/rds_ht
    NORM_PRIORITY = ["vdss", "_rds_tj", "rds25", "rds_ht", "qg", "eon", "eoff"]
    NORM_SECONDARY = [k for k in SECONDARY_KEYS]

    NORM_DISP = {
        "_rds_tj": (f"R_DS(on) @ Tj={TJ_ref:.0f}°C  (extrapolated)", "mΩ", True, True, False),
        "rds25":   ("R_DS(on) @ 25°C  (direct)",                     "mΩ", True, True, False),
        "rds_ht":  ("R_DS(on) @ High-T  (direct)",                   "mΩ", True, True, False),
    }

    def _disp(k):
        if k in NORM_DISP: return NORM_DISP[k]
        return PARAMS[k]

    def norm_vals(k):
        return [norm_list[i].get(k, (None, "", ""))[0] for i in range(n)]

    def norm_meths(k):
        return [norm_list[i].get(k, (None, "", ""))[1] for i in range(n)]

    def _norm_section(ws, start_row, keys, section_label, section_bg, fdc=3):
        row  = start_row
        wins = {i: 0 for i in range(n)}
        ncols = fdc - 1 + n + 2

        merge_title(ws, row, 1, ncols, section_label, bg=section_bg, sz=10, h=22)
        row += 1

        for key in keys:
            disp_tuple = _disp(key)
            display, unit, lower_better = disp_tuple[0], disp_tuple[1], disp_tuple[2]
            is_pri = disp_tuple[3]
            bg = C_PRI_ROW if is_pri else (C_ALT if row % 2 == 0 else C_WHITE)
            ws.row_dimensions[row].height = 22

            wc(ws, row, 1, display,
               font=fnt(10, bold=is_pri), fill=ff(bg), align=lft(), border=BT)
            wc(ws, row, 2, unit,
               font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            vals   = norm_vals(key)
            vals_d = {i: v for i, v in enumerate(vals) if v is not None}
            best_i = _winner(vals_d, lower_better)

            for d_i, v in enumerate(vals):
                is_best = (d_i == best_i)
                is_derived = key in ("eon","eoff") and mosfets[d_i].get(f"_{key}_derived")
                cell_fill = C_DERIVED if is_derived else (C_BEST if is_best else bg)
                wc(ws, row, fdc + d_i,
                   round(v, 4) if isinstance(v, float) else (v if v is not None else "N/A"),
                   font=fnt(10, bold=is_best,
                            color=C_BESTF if is_best else ("7F5C00" if is_derived else "000000")),
                   fill=ff(cell_fill),
                   align=ctr(), border=BT, nfmt="0.0####")

            best_col = fdc + n
            if best_i is not None:
                wc(ws, row, best_col, labels[best_i],
                   font=fnt(10, bold=True, color=C_BESTF),
                   fill=ff(C_BEST), align=ctr(), border=BT)
                wins[best_i] += 1
            else:
                wc(ws, row, best_col, "—",
                   font=fnt(10), fill=ff(bg), align=ctr(), border=BT)

            meths  = norm_meths(key)
            unique = list(dict.fromkeys(m for m in meths if m))
            wc(ws, row, best_col + 1, " | ".join(unique),
               font=fnt(8, italic=True, color="444444"),
               fill=ff("F8F8F8"), align=lft(), border=BT)

            row += 1
        return row, wins

    next2, w2a = _norm_section(ws2, 5, NORM_PRIORITY,
                               "★  PRIORITY PARAMETERS  (normalised to common reference)", C_PRI_BG)
    next2, w2b = _norm_section(ws2, next2 + 1, NORM_SECONDARY,
                               "SECONDARY PARAMETERS", C_SEC_BG)

    all_w2 = {i: w2a[i] + w2b[i] for i in range(n)}
    mw2 = max(all_w2.values()) if all_w2 else 0
    wr2 = next2 + 1
    ws2.row_dimensions[wr2].height = 26
    merge_title(ws2, wr2, 1, 2, "Total Wins", bg=C_SEC_BG, sz=10, h=26)
    for d_i in range(n):
        wv = all_w2[d_i]
        wc(ws2, wr2, 3 + d_i, wv,
           font=fnt(12, bold=True, color=C_BESTF if wv == mw2 else "000000"),
           fill=ff(C_BEST if wv == mw2 else C_ALT), align=ctr(), border=BM)
    if mw2 > 0:
        bdi2 = max(all_w2, key=all_w2.__getitem__)
        wc(ws2, wr2, 3 + n, labels[bdi2],
           font=fnt(10, bold=True, color=C_BESTF), fill=ff(C_BEST), align=ctr(), border=BM)

    # ── Figure of Merit  FOM = Rds_on(Tj_ref) × Q_G ─────────────────────────
    # Uses the common-ground Rds value (extrapolated to Tj_ref) and the direct Qg.
    # Lower FOM = better switching + conduction trade-off.
    C_FOM_BG  = "1A1A2E"   # very dark navy — stands out at the bottom
    C_FOM_FG  = "FFD700"   # gold text
    C_FOM_VAL = "2C2C54"   # dark purple for value cells
    C_FOM_BEST= "FFD700"   # gold for best device

    fom_spacer = wr2 + 1
    ws2.row_dimensions[fom_spacer].height = 10   # visual gap

    fom_hdr = fom_spacer + 1
    ws2.row_dimensions[fom_hdr].height = 28
    merge_title(ws2, fom_hdr, 1, 4 + n,
                "⭐  FIGURE OF MERIT  (FOM = R_DS(on) @ Tj_ref  ×  Q_G)",
                bg=C_FOM_BG, fg=C_FOM_FG, sz=12, h=28)

    # Sub-header explaining the formula
    fom_desc = fom_hdr + 1
    ws2.row_dimensions[fom_desc].height = 20
    merge_title(ws2, fom_desc, 1, 4 + n,
                f"FOM = Rds_on [mΩ] × Q_G [nC]   "
                f"(Rds extrapolated to Tj_ref={TJ_ref:.0f}°C  |  Q_G at datasheet test conditions)   "
                f"Lower FOM = better switching–conduction trade-off",
                bg="2E2E5E", fg="CCCCFF", sz=9, h=20)

    # Column headers for FOM row
    fom_row = fom_desc + 1
    ws2.row_dimensions[fom_row].height = 30

    wc(ws2, fom_row, 1, "FOM = R_DS(on) × Q_G",
       font=Font(name="Calibri", size=11, bold=True, color=C_FOM_FG),
       fill=ff(C_FOM_VAL), align=lft(), border=BM)
    wc(ws2, fom_row, 2, "mΩ·nC",
       font=Font(name="Calibri", size=10, bold=True, color="AAAAFF"),
       fill=ff(C_FOM_VAL), align=ctr(), border=BM)

    # Compute FOM per device
    fom_vals = {}
    for d_i in range(n):
        rds_v = norm_list[d_i].get("_rds_tj", (None, "", ""))[0]
        qg_v  = norm_list[d_i].get("qg",      (None, "", ""))[0]
        if rds_v is not None and qg_v is not None:
            fom_vals[d_i] = round(rds_v * qg_v, 2)   # mΩ × nC

    best_fom_i = min(fom_vals, key=fom_vals.__getitem__) if fom_vals else None

    for d_i in range(n):
        fv = fom_vals.get(d_i)
        is_best = (d_i == best_fom_i)
        cell_bg = C_FOM_BEST if is_best else C_FOM_VAL
        cell_fg = C_FOM_BG   if is_best else "FFFFFF"
        c = ws2.cell(row=fom_row, column=3 + d_i,
                     value=fv if fv is not None else "N/A")
        c.font   = Font(name="Calibri", size=12, bold=True, color=cell_fg)
        c.fill   = ff(cell_bg)
        c.alignment = ctr(); c.border = BM
        if fv is not None:
            c.number_format = "0.00"

    # Best column
    fom_best_col = 3 + n
    if best_fom_i is not None:
        wc(ws2, fom_row, fom_best_col, f"🏆 {labels[best_fom_i]}",
           font=Font(name="Calibri", size=11, bold=True, color=C_FOM_BG),
           fill=ff(C_FOM_BEST), align=ctr(), border=BM)
    else:
        wc(ws2, fom_row, fom_best_col, "N/A",
           font=fnt(10), fill=ff(C_FOM_VAL), align=ctr(), border=BM)

    # Extrapolation method note
    fom_meth_col = 4 + n
    rds_meths = [norm_list[d_i].get("_rds_tj", (None, ""))[1] for d_i in range(n)]
    qg_meths  = [norm_list[d_i].get("qg",      (None, ""))[1] for d_i in range(n)]
    meth_parts = []
    unique_rds = list(dict.fromkeys(m for m in rds_meths if m))
    unique_qg  = list(dict.fromkeys(m for m in qg_meths  if m))
    if unique_rds: meth_parts.append("Rds: " + " | ".join(unique_rds))
    if unique_qg:  meth_parts.append("Qg: "  + " | ".join(unique_qg))
    wc(ws2, fom_row, fom_meth_col, "  ".join(meth_parts),
       font=fnt(8, italic=True, color="444444"),
       fill=ff("F8F8F8"), align=lft(), border=BM)

    # Footnote row explaining FOM interpretation
    fom_note = fom_row + 1
    ws2.row_dimensions[fom_note].height = 32
    ws2.merge_cells(start_row=fom_note, start_column=1, end_row=fom_note, end_column=4 + n)
    fn = ws2.cell(row=fom_note, column=1,
                  value=(
                      "📌  FOM Interpretation:  A lower FOM means the device has a better balance between "
                      "conduction loss (R_DS(on)) and switching loss (Q_G).  "
                      "Both Rds and Qg increase with gate-oxide thickness — the FOM product captures this "
                      "fundamental trade-off for a given technology node.  "
                      "Use FOM to compare devices across different manufacturers on equal footing, "
                      "independent of die size or current rating."
                  ))
    fn.font      = fnt(9, italic=True, color="1A1A2E")
    fn.fill      = ff("EAEAF8")
    fn.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    fn.border    = BT

    _cw(ws2, {"A": 40, "B": 9})
    for d_i in range(n): ws2.column_dimensions[gcol(3 + d_i)].width = 18
    ws2.column_dimensions[gcol(3 + n)].width = 26
    ws2.column_dimensions[gcol(4 + n)].width = 62

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — OPERATING CONDITIONS  (yellow editable inputs + device params)
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Operating Conditions")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A1"

    merge_title(ws3, 1, 1, 3, "⚙  OPERATING CONDITIONS",
                bg=C_TITLE, sz=13, h=36)
    merge_title(ws3, 2, 1, 3,
                "🟡 Yellow cells = edit these for your circuit  |  "
                "🔵 Blue cells = device parameters (auto-filled, editable)",
                bg=C_SUBHDR, fg="FFFFFF", sz=9, h=20)

    # ── Section A: Circuit operating point ────────────────────────────────────
    merge_title(ws3, 4, 1, 3, "▼  CIRCUIT OPERATING POINT  (edit yellow cells)",
                bg=C_SUBHDR, sz=11, h=24)

    ws3.row_dimensions[5].height = 22
    for ci, txt in enumerate(["Parameter", "Value", "Description"], 1):
        wc(ws3, 5, ci, txt,
           font=fnt(10, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=ctr(), border=BT)

    OP = [
        ("I_avg",     "A",   iavg_def,     "Average switch current — used for switching-energy scaling"),
        ("I_rms",     "A",   irms_def,     "RMS switch current — drives conduction loss (I²·Rds)"),
        ("F_sw",      "Hz",  fsw_def,      "Switching frequency — scales all switching losses"),
        ("V_out",     "V",   vout_def,     "DC bus / output voltage (drain-source voltage during switching)"),
        ("VGS_drive", "V",   vgs_drive_def,"Gate drive supply voltage — used for gate-charge loss"),
    ]
    op_rows = {}
    for i, (sym, unit, default, desc) in enumerate(OP):
        r = 6 + i
        ws3.row_dimensions[r].height = 26
        wc(ws3, r, 1, f"{sym}  ({unit})",
           font=fnt(11, bold=True), fill=ff(C_PRI_ROW), align=lft(), border=BT)
        inp = ws3.cell(row=r, column=2, value=default)
        inp.font = fnt(13, bold=True, color=C_INPUT.replace("FFF2CC","7F6000"))
        inp.fill = ff(C_INPUT); inp.alignment = ctr()
        inp.border = BM; inp.number_format = "0.###"
        wc(ws3, r, 3, desc,
           font=fnt(9, italic=True, color="555555"),
           fill=ff(C_WHITE), align=lft(), border=BT)
        op_rows[sym] = r

    # Named references (absolute) — Loss Calc sheet uses these cross-sheet refs
    IAVG = f"='Operating Conditions'!$B${op_rows['I_avg']}"
    IRMS = f"='Operating Conditions'!$B${op_rows['I_rms']}"
    FSW  = f"='Operating Conditions'!$B${op_rows['F_sw']}"
    VOUT = f"='Operating Conditions'!$B${op_rows['V_out']}"
    VGS  = f"='Operating Conditions'!$B${op_rows['VGS_drive']}"
    # Absolute cell addresses for use inside formulas on the Loss Calc sheet
    R_IAVG = f"'Operating Conditions'!$B${op_rows['I_avg']}"
    R_IRMS = f"'Operating Conditions'!$B${op_rows['I_rms']}"
    R_FSW  = f"'Operating Conditions'!$B${op_rows['F_sw']}"
    R_VOUT = f"'Operating Conditions'!$B${op_rows['V_out']}"
    R_VGS  = f"'Operating Conditions'!$B${op_rows['VGS_drive']}"

    # ── Section B: Device Parameters (one row per device param, one col per device)
    PB_START_ROW = 14
    merge_title(ws3, PB_START_ROW, 1, 2 + n,
                "▼  DEVICE PARAMETERS  (auto-filled from datasheets — you may override blue cells)",
                bg="4472C4", fg="FFFFFF", sz=11, h=24)

    ws3.row_dimensions[PB_START_ROW + 1].height = 30
    wc(ws3, PB_START_ROW + 1, 1, "Parameter  (unit)",
       font=fnt(10, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=lft(), border=BT)
    wc(ws3, PB_START_ROW + 1, 2, "Description",
       font=fnt(10, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=lft(), border=BT)
    for d_i, lbl in enumerate(labels):
        wc(ws3, PB_START_ROW + 1, 3 + d_i, lbl,
           font=fnt(10, bold=True, color="FFFFFF"), fill=ff(C_SUBHDR), align=ctr(), border=BT)

    DEV_BLOCK = [
        # (param_label, description, key_in_d)
        ("Rds_on @ Tj_ref  (mΩ)",    f"On-resistance extrapolated to Tj={ref['tj_ref']:.0f}°C",   "_rds_tj_val"),
        ("E_on  (µJ)",               "Turn-on switching energy (scaled to ID_ref)",               "_eon_scaled"),
        ("E_off  (µJ)",              "Turn-off switching energy (scaled to ID_ref)",              "_eoff_scaled"),
        ("E_oss  (µJ)",              "Output-cap stored energy (½·Coss·VDS²) per cycle",         "eoss"),
        ("Q_G  (nC)",                "Total gate charge",                                         "qg"),
        ("ID_test  (A)",             "Test current at which Eon/Eoff were measured",              "_id_test"),
    ]
    pb_rows = {}   # pkey → row number on THIS sheet (ws3)
    for i, (plbl, pdesc, pkey) in enumerate(DEV_BLOCK):
        r = PB_START_ROW + 2 + i
        ws3.row_dimensions[r].height = 22
        wc(ws3, r, 1, plbl,
           font=fnt(10, bold=True, color="1F3864"),
           fill=ff("EEF4FF"), align=lft(), border=BT)
        wc(ws3, r, 2, pdesc,
           font=fnt(9, italic=True, color="444444"),
           fill=ff("EEF4FF"), align=lft(), border=BT)
        for d_i, d in enumerate(mosfets):
            if pkey == "_rds_tj_val":
                val, _ = extrapolate_rds(d, ref["tj_ref"])
            elif pkey == "_eon_scaled":
                val, _ = scale_energy(d.get("eon"), d.get("_eon_id"), ref["id_ref"])
            elif pkey == "_eoff_scaled":
                val, _ = scale_energy(d.get("eoff"), d.get("_eoff_id"), ref["id_ref"])
            elif pkey == "_id_test":
                val = d.get("_eon_id") or d.get("_sw_id") or d.get("id") or ref["id_ref"]
            else:
                val = d.get(pkey)
            pv = round(float(val or 0), 6)
            wc(ws3, r, 3 + d_i, pv,
               font=fnt(10, color="0000CC"),
               fill=ff(C_DEVPARAM), align=ctr(), border=BT, nfmt="0.000###")
        pb_rows[pkey] = r   # row on ws3

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 52
    for d_i in range(n):
        ws3.column_dimensions[gcol(3 + d_i)].width = 22

    # ── Section C: Normalisation reference note ───────────────────────────────
    note_row = PB_START_ROW + 2 + len(DEV_BLOCK) + 2
    merge_title(ws3, note_row, 1, 2 + n,
                f"Normalisation reference:  VDS_ref={VDS_ref:.0f} V  |  ID_ref={ID_ref:.1f} A  |  Tj_ref={TJ_ref:.0f} °C",
                bg=C_NORM, fg=C_GREEN, sz=10, h=22)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 (was 3b) — LOSS CALCULATION  (live Excel formulas, separate sheet)
    # ══════════════════════════════════════════════════════════════════════════
    ws_lc = wb.create_sheet("Loss Calculation")
    ws_lc.sheet_view.showGridLines = False
    ws_lc.freeze_panes = "A4"

    merge_title(ws_lc, 1, 1, 3 + n + 1,
                "⚡  LOSS CALCULATION  —  All values are live Excel formulas",
                bg=C_TITLE, sz=13, h=36)
    merge_title(ws_lc, 2, 1, 3 + n + 1,
                "Change any yellow cell in 'Operating Conditions' sheet → all losses update here automatically",
                bg=C_SUBHDR, fg="FFFFFF", sz=9, h=20)

    # Helper: cross-sheet param cell address
    def pa(pkey, di):
        return f"'Operating Conditions'!${gcol(3+di)}${pb_rows[pkey]}"

    # Column headers  Col A=Loss Component | B=Formula | C..C+n-1=devices | last=Best
    ws_lc.row_dimensions[3].height = 32
    for ci, (txt, bg) in enumerate(
            [("Loss Component", C_HDR), ("Formula Used", C_HDR)] +
            [(lbl, C_SUBHDR) for lbl in labels] +
            [("Best (Lowest)", C_HDR)], 1):
        wc(ws_lc, 3, ci, txt,
           font=fnt(11, bold=True, color="FFFFFF"), fill=ff(bg), align=ctr(), border=BT)

    # ── Loss formula definitions ───────────────────────────────────────────────
    # Col A = label, Col B = human-readable formula, Col C+ = live Excel formula

    def lf(kind, di):
        """Return live Excel formula string for loss kind and device index di."""
        if kind == "cond":
            return (f"={R_IRMS}^2"
                    f"*{pa('_rds_tj_val',di)}/1000")
        if kind == "sw_on":
            return (f"=IF({pa('_id_test',di)}=0,0,"
                    f"{pa('_eon_scaled',di)}/1000000"
                    f"*({R_IAVG}/{pa('_id_test',di)})"
                    f"*{R_FSW})")
        if kind == "sw_off":
            return (f"=IF({pa('_id_test',di)}=0,0,"
                    f"{pa('_eoff_scaled',di)}/1000000"
                    f"*({R_IAVG}/{pa('_id_test',di)})"
                    f"*{R_FSW})")
        if kind == "coss":
            return (f"={pa('eoss',di)}/1000000"
                    f"*{R_FSW}")
        if kind == "gate":
            return (f"={pa('qg',di)}/1000000000"
                    f"*{R_VGS}*{R_FSW}")

    LOSS_DEFS = [
        # (row_offset, label, kind, human_formula, row_bg)
        (0, "P_cond   –  Conduction Loss",
             "cond",
             "I_rms²  ×  Rds_on(Tj) [Ω]",
             C_WHITE),
        (1, "P_sw,on  –  Turn-ON Switching Loss",
             "sw_on",
             "E_on [J]  ×  (I_avg / ID_test)  ×  F_sw",
             C_ALT),
        (2, "P_sw,off –  Turn-OFF Switching Loss",
             "sw_off",
             "E_off [J]  ×  (I_avg / ID_test)  ×  F_sw",
             C_WHITE),
        (3, "P_Coss   –  Output Cap Loss",
             "coss",
             "E_oss [J]  ×  F_sw",
             C_ALT),
        (4, "P_gate   –  Gate Drive Loss",
             "gate",
             "Q_G [C]  ×  V_GS_drive  ×  F_sw",
             C_WHITE),
    ]
    TOTAL_ROW_OFFSET = len(LOSS_DEFS)
    DATA_START_ROW   = 4    # first loss data row on ws_lc
    loss_excel_rows  = {}   # kind → excel row number

    for off, label, kind, formula_str, bg in LOSS_DEFS:
        lrow = DATA_START_ROW + off
        ws_lc.row_dimensions[lrow].height = 26
        loss_excel_rows[kind] = lrow

        wc(ws_lc, lrow, 1, label,
           font=fnt(10, bold=False), fill=ff(bg), align=lft(), border=BT)
        wc(ws_lc, lrow, 2, formula_str,
           font=fnt(9, italic=True, color="1F3864"),
           fill=ff("EEF4FB"), align=lft(), border=BT)

        for d_i in range(n):
            c = ws_lc.cell(row=lrow, column=3 + d_i, value=lf(kind, d_i))
            c.font = fnt(10); c.fill = ff(bg)
            c.alignment = ctr(); c.border = BT; c.number_format = "0.0000"

        mr = ",".join([f"{gcol(3+d_i)}{lrow}" for d_i in range(n)])
        la = "{" + ";".join([f'"{lb}"' for lb in labels]) + "}"
        wc(ws_lc, lrow, 3 + n,
           f'=INDEX({la},MATCH(MIN({mr}),{{{mr}}},0))',
           font=fnt(10, bold=True, color=C_BESTF),
           fill=ff(C_BEST), align=ctr(), border=BT)

    # Total row
    trow = DATA_START_ROW + TOTAL_ROW_OFFSET
    loss_excel_rows["total"] = trow
    ws_lc.row_dimensions[trow].height = 30
    wc(ws_lc, trow, 1, "P_total  –  TOTAL Device Loss",
       font=fnt(11, bold=True, color="FFFFFF"),
       fill=ff(C_HDR), align=lft(), border=BM)
    total_formula_str = "P_cond + P_sw,on + P_sw,off + P_Coss + P_gate"
    wc(ws_lc, trow, 2, total_formula_str,
       font=fnt(9, bold=True, italic=True, color="1F3864"),
       fill=ff("EEF4FB"), align=lft(), border=BM)
    for d_i in range(n):
        total_f = "=" + "+".join([f"{gcol(3+d_i)}{r}" for r in
                                   [loss_excel_rows[k] for k in ("cond","sw_on","sw_off","coss","gate")]])
        c = ws_lc.cell(row=trow, column=3 + d_i, value=total_f)
        c.font = fnt(12, bold=True); c.fill = ff(C_PRI_ROW)
        c.alignment = ctr(); c.border = BM; c.number_format = "0.0000"
    mr_t = ",".join([f"{gcol(3+d_i)}{trow}" for d_i in range(n)])
    la_t = "{" + ";".join([f'"{lb}"' for lb in labels]) + "}"
    wc(ws_lc, trow, 3 + n,
       f'=INDEX({la_t},MATCH(MIN({mr_t}),{{{mr_t}}},0))',
       font=fnt(11, bold=True, color=C_BESTF),
       fill=ff(C_BEST), align=ctr(), border=BM)

    # ── Units row ──────────────────────────────────────────────────────────────
    urow = trow + 1
    ws_lc.row_dimensions[urow].height = 18
    wc(ws_lc, urow, 1, "Units",
       font=fnt(9, italic=True, color="666666"), fill=ff("F5F5F5"), align=lft(), border=BT)
    wc(ws_lc, urow, 2, "",  fill=ff("F5F5F5"), border=BT)
    for d_i in range(n):
        wc(ws_lc, urow, 3 + d_i, "W",
           font=fnt(9, italic=True, color="666666"),
           fill=ff("F5F5F5"), align=ctr(), border=BT)
    wc(ws_lc, urow, 3 + n, "", fill=ff("F5F5F5"), border=BT)

    # ── Recommendation banner ──────────────────────────────────────────────────
    ws_lc.row_dimensions[trow + 3].height = 10
    merge_title(ws_lc, trow + 4, 1, 3 + n + 1,
                "🏆  BEST MOSFET FOR LOWEST TOTAL LOSS", bg=C_TITLE, sz=12, h=30)
    rrow = trow + 5
    ws_lc.row_dimensions[rrow].height = 34
    ws_lc.merge_cells(start_row=rrow, start_column=1, end_row=rrow, end_column=3 + n + 1)
    rc = ws_lc.cell(row=rrow, column=1,
                    value=(f'=INDEX({la_t},MATCH(MIN({mr_t}),{{{mr_t}}},0))'
                           f'&"  —  Total Loss = "&TEXT(MIN({mr_t}),"0.000")&" W"'))
    rc.font = Font(name="Calibri", bold=True, size=14, color=C_GOLD)
    rc.fill = ff(C_GREEN); rc.alignment = ctr(); rc.border = BM

    # ── Loss formula explanation block ─────────────────────────────────────────
    expl_start = rrow + 3
    merge_title(ws_lc, expl_start, 1, 3 + n + 1,
                "📐  FORMULA REFERENCE  —  How each loss is calculated",
                bg="1F3864", fg="FFFFFF", sz=11, h=26)

    FORMULA_DETAIL = [
        # (symbol, formula, derivation_note, unit_note)
        ("P_cond",
         "= I_rms²  ×  R_DS(on) [Ω]",
         "I_rms from Op. Conditions sheet (B6).  R_DS(on) is the value extrapolated to "
         "Tj_ref via linear interpolation between the 25°C and high-T datasheet points, "
         "then stored in Op. Conditions sheet device-param table.",
         "Watts (W)"),
        ("P_sw,on",
         "= E_on [J]  ×  (I_avg / ID_test)  ×  F_sw",
         "E_on is the turn-on energy from the datasheet (µJ→J ÷1e6).  "
         "It is linearly scaled by the ratio I_avg/ID_test to correct for the difference "
         "between the datasheet test current and the actual operating current.  "
         "If Eon was not tabulated it is estimated as ½·VDS_test·ID_test·(t_r + t_d(on)).",
         "Watts (W)"),
        ("P_sw,off",
         "= E_off [J]  ×  (I_avg / ID_test)  ×  F_sw",
         "Same scaling as P_sw,on but using E_off (turn-off energy).  "
         "Estimated as ½·VDS_test·ID_test·(t_f + t_d(off)) when not tabulated.",
         "Watts (W)"),
        ("P_Coss",
         "= E_oss [J]  ×  F_sw",
         "E_oss is the energy stored in C_oss at V_DS = V_out.  "
         "In hard switching this energy is dissipated every turn-on cycle.  "
         "When E_oss is not tabulated it is approximated as ½·C_oss·V_DS².",
         "Watts (W)"),
        ("P_gate",
         "= Q_G [C]  ×  V_GS_drive  ×  F_sw",
         "Q_G is the total gate charge (nC→C ÷1e9) from the datasheet at the stated "
         "V_DS and I_D test conditions.  V_GS_drive is the gate supply voltage entered "
         "in the Operating Conditions sheet.",
         "Watts (W)"),
        ("P_total",
         "= P_cond + P_sw,on + P_sw,off + P_Coss + P_gate",
         "Sum of all five loss components above.  Body-diode conduction loss during "
         "dead-time and reverse-recovery loss (P_rr = Q_rr·V_DS·F_sw) are NOT included "
         "here as they depend on dead-time which is circuit-topology-specific.",
         "Watts (W)"),
    ]

    for i, (sym, formula, note, unit_note) in enumerate(FORMULA_DETAIL):
        r = expl_start + 1 + i * 3
        ws_lc.row_dimensions[r].height   = 22
        ws_lc.row_dimensions[r+1].height = 42
        ws_lc.row_dimensions[r+2].height = 8   # spacer

        # Symbol + formula
        ws_lc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3+n+1)
        ch = ws_lc.cell(row=r, column=1,
                         value=f"  {sym}   {formula}   [{unit_note}]")
        ch.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ch.fill  = ff(C_SUBHDR); ch.alignment = lft(); ch.border = BT

        # Derivation note
        ws_lc.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=3+n+1)
        cn = ws_lc.cell(row=r+1, column=1, value=f"    ↳  {note}")
        cn.font  = fnt(9, italic=True, color="1F3864")
        cn.fill  = ff("EEF4FB"); cn.alignment = lft(); cn.border = BT

    ws_lc.column_dimensions["A"].width = 38
    ws_lc.column_dimensions["B"].width = 40
    for d_i in range(n): ws_lc.column_dimensions[gcol(3 + d_i)].width = 22
    ws_lc.column_dimensions[gcol(3 + n)].width = 26

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — SCORING DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Scoring Dashboard")
    ws4.sheet_view.showGridLines = False

    # ncols for scoring = Parameter + Score Formula + n device cols
    SC_NCOLS = 2 + n

    merge_title(ws4, 1, 1, SC_NCOLS,
                "📊  SCORING DASHBOARD  (0 = worst  |  100 = best  |  normalised to common-ground values)",
                bg=C_TITLE, sz=13, h=36)
    merge_title(ws4, 2, 1, SC_NCOLS,
                "Priority parameters (★) weighted ×2 in overall score.  "
                "Lower-is-better:  Score = (max − value) / (max − min) × 100.  "
                "Higher-is-better:  Score = (value − min) / (max − min) × 100.",
                bg=C_SUBHDR, fg="FFFFFF", sz=9, h=22)

    # Column headers: Parameter | Device1 | Device2 ... | Score Formula
    ws4.row_dimensions[3].height = 28
    wc(ws4, 3, 1, "Parameter",
       font=fnt(11, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=ctr(), border=BT)
    for d_i, lbl in enumerate(labels):
        wc(ws4, 3, 2 + d_i, lbl,
           font=fnt(11, bold=True, color="FFFFFF"), fill=ff(C_SUBHDR), align=ctr(), border=BT)
    wc(ws4, 3, 2 + n, "Score Formula Applied",
       font=fnt(11, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=ctr(), border=BT)

    scores  = [0.0] * n
    max_pts = 0

    def _score_block(ws4, start_row, keys, section_lbl, section_bg, weight):
        nonlocal max_pts
        row = start_row
        merge_title(ws4, row, 1, SC_NCOLS, section_lbl, bg=section_bg, sz=10, h=20)
        row += 1
        for key in keys:
            disp_t = _disp(key)
            display, unit, lower_better = disp_t[0], disp_t[1], disp_t[2]
            is_pri  = disp_t[3]
            bg = C_PRI_ROW if is_pri else (C_ALT if row % 2 == 0 else C_WHITE)
            ws4.row_dimensions[row].height = 22
            wc(ws4, row, 1,
               f"{display} ({unit})" + (" ★×2" if is_pri else ""),
               font=fnt(10, bold=is_pri), fill=ff(bg), align=lft(), border=BT)

            vals_d = {i: norm_list[i].get(key, (None, "", ""))[0]
                      for i in range(n)
                      if norm_list[i].get(key, (None, "", ""))[0] is not None}
            if not vals_d:
                for d_i in range(n):
                    wc(ws4, row, 2 + d_i, "—", font=fnt(10), fill=ff(bg), align=ctr(), border=BT)
                wc(ws4, row, 2 + n, "No data",
                   font=fnt(9, italic=True, color="AAAAAA"),
                   fill=ff(bg), align=lft(), border=BT)
                row += 1; continue

            mn, mx = min(vals_d.values()), max(vals_d.values())
            rng    = mx - mn
            max_pts += weight * 100
            best_v = mn if lower_better else mx

            for d_i in range(n):
                v = vals_d.get(d_i)
                if v is None:
                    wc(ws4, row, 2 + d_i, "—", font=fnt(10), fill=ff(bg), align=ctr(), border=BT)
                    continue
                s = 100.0 if rng == 0 else (
                    (mx - v) / rng * 100 if lower_better else (v - mn) / rng * 100)
                scores[d_i] += s * weight
                is_best = abs(v - best_v) < 1e-9
                wc(ws4, row, 2 + d_i, round(s, 1),
                   font=fnt(10, bold=is_best, color=C_BESTF if is_best else "000000"),
                   fill=ff(C_BEST if is_best else bg), align=ctr(), border=BT, nfmt="0.0")

            # Score formula column — show the actual numbers used
            direction = "lower-is-better" if lower_better else "higher-is-better"
            if rng == 0:
                score_formula_str = f"All equal → 100  [{direction}]"
            elif lower_better:
                score_formula_str = (f"({mx:.4g} − x) / ({mx:.4g} − {mn:.4g}) × 100"
                                     f"  [{direction}]  ×{weight}")
            else:
                score_formula_str = (f"(x − {mn:.4g}) / ({mx:.4g} − {mn:.4g}) × 100"
                                     f"  [{direction}]  ×{weight}")
            wc(ws4, row, 2 + n, score_formula_str,
               font=fnt(8, italic=True, color="1F3864"),
               fill=ff("EEF4FB"), align=lft(), border=BT)

            row += 1
        return row

    nr4 = _score_block(ws4, 4, PRIORITY_KEYS,  "★  PRIORITY  (weight ×2)", C_PRI_BG, 2)
    nr4 = _score_block(ws4, nr4 + 1, SECONDARY_KEYS, "SECONDARY  (weight ×1)", C_SEC_BG, 1)

    norm_scores = ([round(s / max(max_pts, 1) * 100, 1) for s in scores]
                   if max_pts > 0 else [0.0] * n)
    tot_row = nr4 + 1
    ws4.row_dimensions[tot_row - 1].height = 6
    ws4.row_dimensions[tot_row].height     = 32
    wc(ws4, tot_row, 1, "OVERALL SCORE  (0–100)",
       font=fnt(12, bold=True, color="FFFFFF"), fill=ff(C_HDR), align=lft(), border=BM)
    overall_formula = (f"= (Σ weighted_score_per_param) / {max_pts} × 100  "
                       f"   [max possible weighted points = {max_pts}]")
    wc(ws4, tot_row, 2 + n, overall_formula,
       font=fnt(9, italic=True, color="1F3864"),
       fill=ff("EEF4FB"), align=lft(), border=BM)
    mx_s = max(norm_scores)
    for d_i, s in enumerate(norm_scores):
        wc(ws4, tot_row, 2 + d_i, s,
           font=fnt(13, bold=True, color=C_BESTF if s == mx_s else "000000"),
           fill=ff(C_BEST if s == mx_s else C_ALT), align=ctr(), border=BM, nfmt="0.0")

    bdi4 = norm_scores.index(mx_s)
    ws4.row_dimensions[tot_row + 2].height = 32
    merge_title(ws4, tot_row + 2, 1, SC_NCOLS,
                f"🏆  Best MOSFET (overall score):  {labels[bdi4]}   ({norm_scores[bdi4]:.1f}/100)",
                bg=C_GREEN, fg=C_GOLD, sz=12, h=32)

    # ── Scoring methodology explanation ───────────────────────────────────────
    expl_row = tot_row + 5
    merge_title(ws4, expl_row, 1, SC_NCOLS,
                "📐  SCORING METHODOLOGY  —  How each score is calculated",
                bg="1F3864", fg="FFFFFF", sz=11, h=26)

    SCORING_NOTES = [
        ("Step 1 — Raw values",
         "All parameter values are taken from the Common-Ground Comparison sheet, "
         "meaning they have all been extrapolated / scaled to the same reference "
         f"conditions: VDS_ref={VDS_ref:.0f}V, ID_ref={ID_ref:.1f}A, Tj_ref={TJ_ref:.0f}°C."),
        ("Step 2 — Per-parameter normalisation",
         "For each parameter the best and worst values across all devices are found.  "
         "If lower is better (e.g. Rds, Eon):   Score = (max − value) / (max − min) × 100  "
         "→ the device with the lowest value gets 100, the highest gets 0.  "
         "If higher is better (e.g. VDSS, ID):  Score = (value − min) / (max − min) × 100  "
         "→ the device with the highest value gets 100, the lowest gets 0.  "
         "If all devices are equal on a parameter, all get 100."),
        ("Step 3 — Weighting",
         "Priority parameters (★) are multiplied by weight = 2.  "
         "Secondary parameters are multiplied by weight = 1.  "
         "This reflects that conduction/switching losses (Rds, Eon, Eoff, Qg, Vdss) "
         "have a larger impact on overall converter efficiency."),
        ("Step 4 — Overall score",
         f"Overall Score = (sum of all weighted scores) / (max possible weighted points) × 100.  "
         f"Max possible = {max_pts} points  "
         f"(= {sum(2 for k in PRIORITY_KEYS)} priority params × 2 × 100  +  "
         f"{sum(1 for k in SECONDARY_KEYS)} secondary params × 1 × 100, counting only params "
         f"where at least one device had data).  "
         f"Result is normalised back to 0–100 so the winning device always scores 100."),
    ]

    for i, (heading, body) in enumerate(SCORING_NOTES):
        r_h = expl_row + 1 + i * 3
        ws4.row_dimensions[r_h].height   = 22
        ws4.row_dimensions[r_h+1].height = 50
        ws4.row_dimensions[r_h+2].height = 6

        ws4.merge_cells(start_row=r_h, start_column=1, end_row=r_h, end_column=SC_NCOLS)
        ch = ws4.cell(row=r_h, column=1, value=f"  {heading}")
        ch.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ch.fill = ff(C_SUBHDR); ch.alignment = lft(); ch.border = BT

        ws4.merge_cells(start_row=r_h+1, start_column=1, end_row=r_h+1, end_column=SC_NCOLS)
        cb = ws4.cell(row=r_h+1, column=1, value=f"    {body}")
        cb.font = fnt(9, italic=True, color="1F3864")
        cb.fill = ff("EEF4FB"); cb.alignment = lft(); cb.border = BT

    _cw(ws4, {"A": 48})
    for d_i in range(n): ws4.column_dimensions[gcol(2 + d_i)].width = 20
    ws4.column_dimensions[gcol(2 + n)].width = 62

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — RAW EXTRACTED DATA + DETAILS
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Raw Extracted Data")
    ws5.sheet_view.showGridLines = False

    merge_title(ws5, 1, 1, 4 + n,
                "📋  RAW EXTRACTED DATA  +  Test Conditions  +  Normalisation Summary",
                bg=C_TITLE, sz=12, h=28)

    ws5.row_dimensions[3].height = 24
    for ci, (txt, bg) in enumerate(
            [("Symbol", C_HDR), ("Parameter", C_HDR), ("Unit", C_HDR)] +
            [(f"{lbl}\n(Typ@25°C)", C_SUBHDR) for lbl in labels] +
            [(f"{lbl}\n(Max@25°C)", C_SUBHDR) for lbl in labels] +
            [(f"{lbl}\n(Typ@High-T)", C_SUBHDR) for lbl in labels] +
            [("Test Conditions", C_HDR)], 1):
        wc(ws5, 3, ci, txt,
           font=fnt(10, bold=True, color="FFFFFF"), fill=ff(bg), align=ctr(), border=BT)

    # Build extended table with all raw values
    EXT_PARAMS = [
        ("V_DS",     "vdss",    "vdss_max",  None,          "V",   "vdss_tcond"),
        ("I_D",      "id",      None,        "id_highT",    "A",   "id_tcond"),
        ("R_DS(on)", "rds25",   "rds25_max", "rds_ht",      "mΩ",  "rds_tcond"),
        ("V_GS(th)", "vth",     "vth_max",   None,          "V",   "vth_tcond"),
        ("C_iss",    "ciss",    None,        None,           "pF",  "ciss_tcond"),
        ("C_oss",    "coss",    "coss_max",  None,           "pF",  "coss_tcond"),
        ("C_rss",    "crss",    None,        None,           "pF",  ""),
        ("Q_g",      "qg",      None,        None,           "nC",  "qg_tcond"),
        ("Q_gd",     "qgd",     None,        None,           "nC",  ""),
        ("E_oss",    "eoss",    None,        None,           "µJ",  "eoss_tcond"),
        ("E_on",     "eon",     None,        None,           "µJ",  "eon_tcond"),
        ("E_off",    "eoff",    None,        None,           "µJ",  "eoff_tcond"),
        ("t_d(on)",  "td_on",   None,        None,           "ns",  "td_on_tcond"),
        ("t_r",      "tr",      None,        None,           "ns",  ""),
        ("t_d(off)", "td_off",  None,        None,           "ns",  ""),
        ("t_f",      "tf",      None,        None,           "ns",  ""),
        ("V_SD",     "vsd",     None,        None,           "V",   "vsd_tcond"),
        ("t_rr",     "trr",     None,        None,           "ns",  "trr_tcond"),
        ("Q_rr",     "qrr",     None,        None,           "nC",  ""),
    ]

    for r_i, (sym, k_typ, k_max, k_ht, unit, k_tc) in enumerate(EXT_PARAMS, 1):
        row = 3 + r_i
        alt_bg = C_PRI_ROW if sym in ("V_DS","R_DS(on)","Q_g","E_on","E_off") else (C_ALT if r_i%2==0 else C_WHITE)
        ws5.row_dimensions[row].height = 20
        wc(ws5, row, 1, sym,       font=fnt(10, bold=True), fill=ff(alt_bg), align=ctr(), border=BT)
        wc(ws5, row, 2, PARAMS.get(k_typ, ("",))[0] if k_typ in PARAMS else sym,
           font=fnt(10), fill=ff(alt_bg), align=lft(), border=BT)
        wc(ws5, row, 3, unit,      font=fnt(10), fill=ff(alt_bg), align=ctr(), border=BT)

        # Typ @ 25°C
        for d_i, d in enumerate(mosfets):
            v = d.get(k_typ)
            is_der = (k_typ in ("eon","eoff") and d.get(f"_{k_typ}_derived"))
            cell_bg = C_DERIVED if is_der else alt_bg
            wc(ws5, row, 4 + d_i,
               round(v, 5) if isinstance(v, float) else (v if v is not None else "—"),
               font=fnt(9, color="7F5C00" if is_der else ("000000" if v is not None else "AAAAAA")),
               fill=ff(cell_bg), align=ctr(), border=BT,
               nfmt="0.00###" if isinstance(v, float) else None)
        # Max @ 25°C
        for d_i, d in enumerate(mosfets):
            v = d.get(k_max) if k_max else None
            wc(ws5, row, 4 + n + d_i,
               round(v, 5) if isinstance(v, float) else (v if v is not None else "—"),
               font=fnt(9, color="000000" if v is not None else "CCCCCC"),
               fill=ff(alt_bg), align=ctr(), border=BT,
               nfmt="0.00###" if isinstance(v, float) else None)
        # Typ @ High-T
        for d_i, d in enumerate(mosfets):
            v = d.get(k_ht) if k_ht else None
            tj_ht = d.get("_rds_ht_tj") if k_ht == "rds_ht" else None
            lbl_suffix = f" @{tj_ht:.0f}°C" if tj_ht else ""
            wc(ws5, row, 4 + 2*n + d_i,
               (f"{round(v,4)}{lbl_suffix}" if isinstance(v, float) else (f"{v}{lbl_suffix}" if v is not None else "—")),
               font=fnt(9, color="000000" if v is not None else "CCCCCC"),
               fill=ff(alt_bg), align=ctr(), border=BT)
        # Test conditions
        conds = [d.get(k_tc, "") or "" for d in mosfets] if k_tc else [""] * n
        unique_c = "; ".join(dict.fromkeys(c for c in conds if c))
        ws5.merge_cells(start_row=row, start_column=4+3*n, end_row=row, end_column=4+3*n)
        wc(ws5, row, 4 + 3*n, unique_c,
           font=fnt(8, italic=True, color="444444"),
           fill=ff("F8F8F8"), align=lft(), border=BT)

    # Normalisation summary
    ns = 3 + len(EXT_PARAMS) + 3
    merge_title(ws5, ns, 1, 3 + n,
                (f"NORMALISATION SUMMARY  "
                 f"(VDS_ref={VDS_ref:.0f}V · ID_ref={ID_ref:.1f}A · Tj_ref={TJ_ref:.0f}°C)"),
                bg=C_SUBHDR, sz=10, h=20)
    ws5.row_dimensions[ns + 1].height = 22
    for ci, (txt, bg) in enumerate([("Parameter", C_HDR)] + [(lbl, C_SUBHDR) for lbl in labels], 1):
        wc(ws5, ns + 1, ci, txt,
           font=fnt(10, bold=True, color="FFFFFF"), fill=ff(bg), align=ctr(), border=BT)

    for r_i, (key, disp_str) in enumerate([
        ("_rds_tj",  f"Rds_on @ {TJ_ref:.0f}°C [mΩ]"),
        ("eon",      f"E_on @ ID={ID_ref:.1f}A [µJ]"),
        ("eoff",     f"E_off @ ID={ID_ref:.1f}A [µJ]"),
        ("coss",     f"Coss @ VDS={VDS_ref:.0f}V [pF]"),
    ], 1):
        row = ns + 1 + r_i
        ws5.row_dimensions[row].height = 42
        wc(ws5, row, 1, disp_str,
           font=fnt(9, bold=True), fill=ff(C_ALT), align=lft(), border=BT)
        for d_i in range(n):
            v, m, _ = norm_list[d_i].get(key, (None, "", ""))
            wc(ws5, row, 2 + d_i,
               f"→ {round(v,4) if v is not None else 'N/A'}\n{m}",
               font=fnt(8), fill=ff(C_WHITE), align=lft(), border=BT)

    _cw(ws5, {"A": 12, "B": 42, "C": 7})
    for d_i in range(n):
        ws5.column_dimensions[gcol(4 + d_i)].width     = 16  # typ
        ws5.column_dimensions[gcol(4+n+d_i)].width     = 14  # max
        ws5.column_dimensions[gcol(4+2*n+d_i)].width   = 16  # high-T
    ws5.column_dimensions[gcol(4+3*n)].width            = 48  # conditions

    # ── Graph Comparison sheet ────────────────────────────────────────────────
    tmp_imgs = []
    if all_charts_per_file is not None:
        print("  Building Graph Comparison sheet …")
        tmp_imgs = build_graph_comparison_sheet(
            wb, paths, mosfets, labels, all_charts_per_file)

    wb.save(out_path)
    # Clean up temp PNG files used for embedding
    for tp in tmp_imgs:
        try:
            os.remove(tp)
        except OSError:
            pass
    print(f"\n✅  Report saved → {out_path}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def get_float(prompt, default=None):
    while True:
        raw = input(prompt).strip()
        if not raw and default is not None:
            return default
        try:
            return float(raw)
        except ValueError:
            print("  ✗  Please enter a number.")


def main():
    print("\n" + "═"*65)
    print("  MOSFET Comparison & Loss Analysis Tool  —  v3")
    print("═"*65)

    # ── 1. Collect file paths ─────────────────────────────────────────────────
    files = sys.argv[1:]
    if not files:
        print("\nEnter paths to v5 MOSFET parameter Excel files (minimum 2).")
        print("Press Enter on a blank line when done.\n")
        while True:
            raw = input("  File path: ").strip().strip('"').strip("'")
            if not raw:
                if len(files) >= 2:
                    break
                print("  ⚠  Please enter at least 2 files.")
                continue
            if not os.path.isfile(raw):
                print(f"  ✗  File not found: {raw}")
                continue
            files.append(raw)
            print(f"  ✓  Added  ({len(files)} file{'s' if len(files)>1 else ''} so far)")

    valid = [f for f in files if os.path.isfile(f)]
    if len(valid) < 2:
        sys.exit("❌  Need at least 2 valid Excel files.")

    # ── 2. Parse all MOSFET files ─────────────────────────────────────────────
    print("\n── Parsing MOSFET parameter files ──")
    mosfets            = []
    all_charts_per_file = []
    for f in valid:
        d   = read_v5_file(f)
        lbl = mosfet_label(f, d)
        print(f"\n  ✓  {lbl}")
        print(f"     VDSS={d.get('vdss')}V  ID={d.get('id')}A (25°C) / {d.get('id_highT')}A (high-T)")
        print(f"     Rds25={d.get('rds25')}mΩ (typ)  Rds_HT={d.get('rds_ht')}mΩ @ {d.get('_rds_ht_tj',125)}°C  tcond:[{d.get('rds_tcond','')}]")
        print(f"     Qg={d.get('qg')}nC [{d.get('qg_tcond','')}]")
        print(f"     Eon={d.get('eon')}µJ {'[DERIVED from t_r/t_d(on)]' if d.get('_eon_derived') else '[tabulated]'}  "
              f"Eoff={d.get('eoff')}µJ {'[DERIVED from t_f/t_d(off)]' if d.get('_eoff_derived') else '[tabulated]'}")
        print(f"     Ciss={d.get('ciss')}pF  Coss={d.get('coss')}pF  Crss={d.get('crss')}pF")
        print(f"     Eoss={d.get('eoss')}µJ {'[derived: ½CossV²]' if d.get('_eoss_derived') else '[tabulated]'}")
        print(f"     Vth={d.get('vth')}V  Vsd={d.get('vsd')}V  Qgd={d.get('qgd')}nC")
        print(f"     t_r={d.get('tr')}ns  t_f={d.get('tf')}ns  t_d(on)={d.get('td_on')}ns  t_d(off)={d.get('td_off')}ns")
        print(f"     trr={d.get('trr')}ns  Qrr={d.get('qrr')}nC")
        mosfets.append(d)
        charts = extract_charts_from_file(f)
        all_charts_per_file.append(charts)
        print(f"     Charts found: {len(charts)}"
              + (f" ({', '.join(dict.fromkeys(c['chart_type'] for c in charts))})"
                 if charts else " — will use synthetic charts from tabular data"))

    # ── 3. Reference conditions ───────────────────────────────────────────────
    vdss_all = [d.get("vdss") for d in mosfets if d.get("vdss")]
    vds_def  = round(min(vdss_all) * 0.6, 0) if vdss_all else 400.0
    id_all   = [d.get("_eon_id") or d.get("_sw_id") or d.get("id")
                for d in mosfets if (d.get("_eon_id") or d.get("_sw_id") or d.get("id"))]
    id_def   = round(min(id_all), 1) if id_all else 15.0

    print(f"\n── Reference Conditions for Common-Ground Normalisation ──")
    print("   (All devices will be compared at these identical conditions)")
    vds_ref = get_float(f"  VDS_ref (V)   [suggest {vds_def:.0f} V]: ", vds_def)
    id_ref  = get_float(f"  ID_ref  (A)   [suggest {id_def:.1f} A]: ",  id_def)
    tj_ref  = get_float(f"  Tj_ref  (°C)  [125 °C]: ",                  125.0)

    # ── 4. Operating conditions for loss ─────────────────────────────────────
    print("\n── Operating Conditions for Loss Analysis ──")
    iavg    = get_float("  I_avg  (A)   [average switch current]: ")
    irms    = get_float("  I_rms  (A)   [RMS switch current]: ")
    fsw     = get_float("  F_sw   (Hz)  [switching frequency]: ")
    vout    = get_float("  V_out  (V)   [DC output / bus voltage, 400]: ", 400.0)
    vgs_all = [d.get("vgs_drive") for d in mosfets if d.get("vgs_drive")]
    vgs_def = round(sum(vgs_all)/len(vgs_all), 1) if vgs_all else 10.0
    vgs_drv = get_float(f"  VGS_drive (V) [gate supply voltage, {vgs_def}]: ", vgs_def)

    # ── 5. Output filename ────────────────────────────────────────────────────
    out = input("\n  Output filename [mosfet_report.xlsx]: ").strip()
    if not out:       out = "mosfet_report.xlsx"
    if not out.endswith(".xlsx"): out += ".xlsx"

    ref = {"vds_ref": vds_ref, "id_ref": id_ref, "tj_ref": tj_ref}

    # ── 6. Normalise ──────────────────────────────────────────────────────────
    print("\n── Common-ground normalisation ──")
    norm_list = compute_normalised(mosfets, ref)
    for d_i, (d, nm) in enumerate(zip(mosfets, norm_list)):
        lbl = mosfet_label(valid[d_i], d)
        print(f"  {lbl}:")
        for k, disp in [("_rds_tj", f"Rds@{tj_ref:.0f}°C"),
                         ("eon",    f"Eon@ID={id_ref:.1f}A"),
                         ("eoff",   f"Eoff@ID={id_ref:.1f}A"),
                         ("coss",   f"Coss@VDS={vds_ref:.0f}V")]:
            v, m, _ = nm.get(k, (None, "", ""))
            print(f"    {disp:22} = {v}  [{str(m)[:70]}]")

    # ── 7. Build report ───────────────────────────────────────────────────────
    print("\n── Building 6-sheet Excel report ──")
    build_report(valid, mosfets, norm_list, ref, iavg, irms, fsw, vout, vgs_drv, out,
                 all_charts_per_file=all_charts_per_file)

    print("\n📌  Quick guide (7 sheets):")
    print("  Sheet 1 'Raw Parameter Comparison'    — all values as extracted from datasheet + conditions")
    print("  Sheet 2 'Common-Ground Comparison'    — all extrapolated to same VDS/ID/Tj reference")
    print("  Sheet 3 'Operating Conditions'        — YELLOW cells to edit + BLUE device params table")
    print("  Sheet 4 'Loss Calculation'            — live formula results + full formula reference")
    print("  Sheet 5 'Scoring Dashboard'           — 0-100 score per param + scoring methodology")
    print("  Sheet 6 'Raw Extracted Data'          — full typ/max/high-T table + test conditions")
    print("  Sheet 7 'Graph Comparison'            — side-by-side charts + overlay + 5-point comparison table\n")


if __name__ == "__main__":
    main()
